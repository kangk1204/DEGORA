#!/usr/bin/env python
"""Render a data-backed DEGORA evidence-card figure for selected benchmark genes."""

from __future__ import annotations

import argparse
import json
import sqlite3
import zipfile
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from degora.provenance import shell_command, write_source_sidecar


FIGURE_ID = "DEGORA_EVIDENCE_CARDS"


def _normalize_symbol(value: object) -> str:
    return str(value).strip().upper()


def parse_case(value: str) -> dict[str, Any]:
    parts = value.split("|")
    if len(parts) != 4:
        raise ValueError("--case must be formatted as corpus|score_csv|db_path|GENE1;GENE2")
    corpus, score_csv, db_path, genes = parts
    return {
        "corpus": corpus,
        "score_csv": Path(score_csv),
        "db_path": Path(db_path),
        "genes": [_normalize_symbol(gene) for gene in genes.split(";") if _normalize_symbol(gene)],
    }


def _score_metrics(case: dict[str, Any]) -> pd.DataFrame:
    score = pd.read_csv(case["score_csv"])
    score["_symbol"] = score["gene_symbol"].map(_normalize_symbol)
    rows = []
    for gene in case["genes"]:
        hit = score.loc[score["_symbol"].eq(gene)]
        if hit.empty:
            rows.append({"corpus": case["corpus"], "gene_symbol": gene, "present": False})
            continue
        row = hit.iloc[0].to_dict()
        rows.append(
            {
                "corpus": case["corpus"],
                "gene_symbol": gene,
                "present": True,
                "degora_rank": row.get("degora_rank", ""),
                "quality_weighted_degora_rank": row.get("quality_weighted_degora_rank", ""),
                "quality_weighted_top_percent": row.get("quality_weighted_top_percent", row.get("top_percent", "")),
                "consensus_direction": row.get("quality_weighted_consensus_direction", row.get("consensus_direction", "")),
                "n_source_units": row.get("n_source_units", ""),
                "sign_concordance": row.get("quality_weighted_sign_concordance", row.get("sign_concordance", "")),
                "source_quality_weight_sum": row.get("source_quality_weight_sum", ""),
                "loo_rank_stability_score": row.get("loo_rank_stability_score", ""),
                "loo_top100_fraction": row.get("loo_top100_fraction", ""),
                "source_units": row.get("source_units", ""),
            }
        )
    return pd.DataFrame.from_records(rows)


def _source_evidence(case: dict[str, Any]) -> pd.DataFrame:
    if not case["db_path"].exists():
        return pd.DataFrame()
    genes = case["genes"]
    if not genes:
        return pd.DataFrame(
            columns=[
                "gene_symbol",
                "source_unit_id",
                "lfc",
                "signed_z",
                "aggregate_pvalue",
                "source_quality_weight",
                "source_reliability_label",
                "direction",
            ]
        )
    placeholders = ",".join(["?"] * len(genes))
    query = f"""
    SELECT gene_symbol, source_unit_id, lfc, signed_z, aggregate_pvalue,
           source_quality_weight, source_reliability_label, direction
    FROM gene_evidence
    WHERE upper(gene_symbol) IN ({placeholders})
    """
    with sqlite3.connect(case["db_path"]) as connection:
        evidence = pd.read_sql_query(query, connection, params=genes)
    if evidence.empty:
        return evidence
    evidence["corpus"] = case["corpus"]
    evidence["gene_symbol"] = evidence["gene_symbol"].map(_normalize_symbol)
    return evidence


def build_source_data(cases: list[dict[str, Any]]) -> tuple[pd.DataFrame, pd.DataFrame]:
    metrics = pd.concat([_score_metrics(case) for case in cases], ignore_index=True)
    evidence_frames = [_source_evidence(case) for case in cases]
    evidence = pd.concat([frame for frame in evidence_frames if not frame.empty], ignore_index=True) if any(not frame.empty for frame in evidence_frames) else pd.DataFrame()
    metrics["quality_weighted_top_percent"] = pd.to_numeric(metrics["quality_weighted_top_percent"], errors="coerce")
    metrics["n_source_units"] = pd.to_numeric(metrics["n_source_units"], errors="coerce")
    metrics["sign_concordance"] = pd.to_numeric(metrics["sign_concordance"], errors="coerce")
    metrics["loo_rank_stability_score"] = pd.to_numeric(metrics["loo_rank_stability_score"], errors="coerce")
    if not evidence.empty:
        for column in ["lfc", "signed_z", "aggregate_pvalue", "source_quality_weight"]:
            evidence[column] = pd.to_numeric(evidence[column], errors="coerce")
    return metrics, evidence


def _write_xlsx(path: Path, metrics: pd.DataFrame, evidence: pd.DataFrame, metadata: dict[str, Any]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "gene_metrics"
    for row in dataframe_to_rows(metrics, index=False, header=True):
        sheet.append(row)
    evidence_sheet = workbook.create_sheet("source_evidence")
    for row in dataframe_to_rows(evidence, index=False, header=True):
        evidence_sheet.append(row)
    meta_sheet = workbook.create_sheet("metadata")
    meta_sheet.append(["key", "value"])
    for key, value in metadata.items():
        meta_sheet.append([key, json.dumps(value, sort_keys=True) if isinstance(value, (dict, list)) else value])
    workbook.save(path)


def _write_docx(path: Path, text: str) -> None:
    content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>"""
    rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>"""
    paragraphs = "".join(
        f"<w:p><w:r><w:t>{escape(line)}</w:t></w:r></w:p>"
        for line in text.splitlines()
    )
    document = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
<w:body>{paragraphs}<w:sectPr/></w:body>
</w:document>"""
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as docx:
        docx.writestr("[Content_Types].xml", content_types)
        docx.writestr("_rels/.rels", rels)
        docx.writestr("word/document.xml", document)


def plot(metrics: pd.DataFrame, evidence: pd.DataFrame, output_stem: Path, *, title: str) -> None:
    plt.rcParams.update(
        {
            "font.family": "DejaVu Sans",
            "font.size": 8.5,
            "axes.titlesize": 10,
            "axes.labelsize": 8.5,
            "legend.fontsize": 8,
            "svg.fonttype": "none",
        }
    )
    metrics = metrics.loc[metrics["present"].eq(True)].copy()
    metrics["label"] = metrics["corpus"].astype(str) + " | " + metrics["gene_symbol"].astype(str)
    metrics = metrics.sort_values(["corpus", "quality_weighted_degora_rank", "gene_symbol"])
    fig, axes = plt.subplots(1, 2, figsize=(12, 5.6), gridspec_kw={"width_ratios": [1.05, 1.25]})

    y_pos = range(len(metrics))
    scatter = axes[0].scatter(
        metrics["quality_weighted_top_percent"],
        list(y_pos),
        s=40 + 32 * metrics["n_source_units"].fillna(0),
        c=metrics["sign_concordance"].fillna(0),
        cmap="viridis",
        vmin=0,
        vmax=1,
        edgecolor="#17202a",
        linewidth=0.4,
    )
    axes[0].set_yticks(list(y_pos))
    axes[0].set_yticklabels(metrics["label"])
    axes[0].invert_yaxis()
    axes[0].set_xlabel("Quality-weighted top percent (lower is better)")
    axes[0].set_title("A. Gene-level evidence cards", loc="left", fontweight="bold")
    axes[0].grid(True, axis="x", color="#e5e7e9")
    fig.colorbar(scatter, ax=axes[0], label="Sign concordance")

    if evidence.empty:
        axes[1].text(0.5, 0.5, "No source evidence rows available", ha="center", va="center")
        axes[1].set_axis_off()
    else:
        evidence = evidence.copy()
        evidence["label"] = evidence["corpus"].astype(str) + " | " + evidence["gene_symbol"].astype(str)
        order = metrics["label"].tolist()
        evidence["x"] = evidence["label"].map({label: i for i, label in enumerate(order)})
        colors = evidence["direction"].map({"up": "#b03a2e", "down": "#1f618d"}).fillna("#566573")
        axes[1].scatter(
            evidence["x"],
            evidence["lfc"].clip(-8, 8),
            s=28 + 28 * evidence["source_quality_weight"].fillna(0.5),
            c=colors,
            alpha=0.78,
            edgecolor="#17202a",
            linewidth=0.25,
        )
        axes[1].axhline(0, color="#17202a", linewidth=0.8)
        axes[1].set_xticks(range(len(order)))
        axes[1].set_xticklabels([label.split(" | ", 1)[1] for label in order], rotation=90)
        axes[1].set_ylabel("Source-unit log2FC (clipped to +/-8)")
        axes[1].set_title("B. Source-unit directional support", loc="left", fontweight="bold")
        axes[1].grid(True, axis="y", color="#e5e7e9")
    fig.suptitle(title, x=0.02, y=0.995, ha="left", fontsize=12, fontweight="bold")
    fig.tight_layout(rect=(0, 0, 1, 0.965))
    for suffix in [".png", ".pdf", ".svg"]:
        fig.savefig(output_stem.with_suffix(suffix), dpi=300)
    plt.close(fig)


def write_package(cases: list[dict[str, Any]], output_dir: Path, *, title: str, command: str) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    metrics, evidence = build_source_data(cases)
    stem = output_dir / "degora_evidence_cards"
    plot(metrics, evidence, stem, title=title)
    source_data = output_dir / "degora_evidence_cards_source_data.xlsx"
    legend = output_dir / "degora_evidence_cards_legend.docx"
    manifest = output_dir / "degora_evidence_cards_manifest.json"
    validation = output_dir / "degora_evidence_cards_validation.txt"
    metadata = {
        "figure_id": FIGURE_ID,
        "generated_at": datetime.now(UTC).isoformat(),
        "cases": [
            {"corpus": case["corpus"], "score_csv": str(case["score_csv"]), "db_path": str(case["db_path"]), "genes": case["genes"]}
            for case in cases
        ],
    }
    _write_xlsx(source_data, metrics, evidence, metadata)
    _write_docx(
        legend,
        (
            "DEGORA evidence cards for representative benchmark genes.\n"
            "A, Quality-weighted rank percentile for selected canonical genes; point size encodes source-unit count and color encodes sign concordance. "
            "B, Source-unit log2 fold changes for the same genes, with each point representing an evidence row in the local DEGORA database."
        ),
    )
    outputs = [stem.with_suffix(suffix) for suffix in [".png", ".pdf", ".svg"]]
    outputs.extend([source_data, legend])
    manifest_record = {
        "figure_id": FIGURE_ID,
        "generated_at": metadata["generated_at"],
        "script": "outputs/code/figures/make_evidence_card_figure.py",
        "inputs": metadata["cases"],
        "outputs": [str(path) for path in outputs],
        "source_data": str(source_data),
        "panel_claims": {
            "A": "Selected genes have inspectable rank percentile, source-unit support count, and sign concordance.",
            "B": "Selected genes expose source-unit-level direction and effect-size support.",
        },
        "transformations": [
            "gene-level metrics read from degora_gene_scores.csv",
            "source-unit evidence read from SQLite gene_evidence",
            "log2FC values clipped to +/-8 for display only",
        ],
        "known_limitations": [
            "selected genes are representative evidence cards, not an unbiased full-benchmark summary",
            "source-unit log2FC values come from heterogeneous as-published or derived DEG tables",
        ],
    }
    manifest.write_text(json.dumps(manifest_record, indent=2, sort_keys=True) + "\n")
    outputs.append(manifest)
    validation_lines = []
    for path in outputs:
        validation_lines.append(f"{path.name}\texists={path.exists()}\tsize={path.stat().st_size if path.exists() else 0}")
    validation.write_text("\n".join(validation_lines) + "\n")
    for _ in range(5):
        current_size = validation.stat().st_size
        self_line = f"{validation.name}\texists={validation.exists()}\tsize={current_size}"
        validation.write_text("\n".join([*validation_lines, self_line]) + "\n")
        if validation.stat().st_size == current_size:
            break
    outputs.append(validation)
    for artifact in outputs:
        write_source_sidecar(
            artifact,
            command,
            inputs=[case["score_csv"] for case in cases] + [case["db_path"] for case in cases],
            metadata={"generator": "degora-evidence-card-figure"},
        )
    return {
        "figure_png": str(stem.with_suffix(".png")),
        "figure_pdf": str(stem.with_suffix(".pdf")),
        "figure_svg": str(stem.with_suffix(".svg")),
        "source_data": str(source_data),
        "legend": str(legend),
        "manifest": str(manifest),
        "validation": str(validation),
        "n_gene_rows": int(len(metrics)),
        "n_evidence_rows": int(len(evidence)),
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--case", action="append", required=True, help="corpus|score_csv|db_path|GENE1;GENE2")
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--title", default="DEGORA evidence cards")
    args = parser.parse_args(argv)
    cases = [parse_case(value) for value in args.case]
    command_parts: list[object] = [
        "env",
        "PYTHONPATH=outputs/code",
        "python",
        "outputs/code/figures/make_evidence_card_figure.py",
    ]
    for value in args.case:
        command_parts.extend(["--case", value])
    command_parts.extend(["--output-dir", args.output_dir, "--title", args.title])
    summary = write_package(cases, args.output_dir, title=args.title, command=shell_command(command_parts))
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
