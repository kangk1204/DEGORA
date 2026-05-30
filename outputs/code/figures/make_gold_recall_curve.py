#!/usr/bin/env python
"""Plot locked gold-panel recall curves through a configurable top-K cutoff."""

from __future__ import annotations

import argparse
import json
import math
import zipfile
from datetime import UTC, datetime
from pathlib import Path
from typing import Iterable
from xml.sax.saxutils import escape

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from degora.baselines import BASELINE_RESULT_COLUMNS, baseline_result_paths
from degora.provenance import shell_command, write_source_sidecar


DEFAULT_METHOD_ORDER = [
    ("degora_deg_score", "v1_2_source_unit_mean"),
    ("degora_quality_weighted_score", "quality_weighted_secondary"),
    ("weighted_stouffer", "default"),
    ("fisher", "default"),
    ("metavolcanor", "default"),
    ("robustrankaggreg", "default"),
    ("rank_product_approx", "default"),
]

METHOD_LABELS = {
    ("degora_deg_score", "v1_2_source_unit_mean"): "DEGORA primary",
    ("degora_quality_weighted_score", "quality_weighted_secondary"): "DEGORA quality-weighted",
    ("weighted_stouffer", "default"): "Weighted Stouffer",
    ("fisher", "default"): "Fisher",
    ("metavolcanor", "default"): "MetaVolcanoR",
    ("robustrankaggreg", "default"): "RobustRankAggreg",
    ("rank_product_approx", "default"): "Rank product approx.",
}

METHOD_STYLES = {
    "DEGORA primary": {"color": "#1b4f72", "linestyle": "-", "linewidth": 2.6},
    "DEGORA quality-weighted": {"color": "#117864", "linestyle": "-", "linewidth": 2.4},
    "Weighted Stouffer": {"color": "#7d3c98", "linestyle": "--", "linewidth": 2.0},
    "Fisher": {"color": "#b03a2e", "linestyle": "--", "linewidth": 2.0},
    "MetaVolcanoR": {"color": "#d35400", "linestyle": ":", "linewidth": 2.2},
    "RobustRankAggreg": {"color": "#566573", "linestyle": "-.", "linewidth": 2.0},
    "Rank product approx.": {"color": "#2e86c1", "linestyle": (0, (3, 1, 1, 1)), "linewidth": 2.0},
}


def _normalize_direction(value: object) -> str:
    text = str(value).strip().lower()
    if text in {"up", "+", "1", "increase", "increased", "upregulated", "up-regulated"}:
        return "up"
    if text in {"down", "-", "-1", "decrease", "decreased", "downregulated", "down-regulated"}:
        return "down"
    return ""


def _read_gold(path: Path, gene_column: str) -> tuple[set[str], dict[str, str]]:
    gold = pd.read_csv(path)
    if gene_column not in gold.columns:
        raise ValueError(f"gold file {path} missing gene column {gene_column!r}")
    gold = gold.copy()
    gold["_gene"] = gold[gene_column].dropna().astype(str).str.strip().str.upper()
    positives = {gene for gene in gold["_gene"] if gene}
    directions: dict[str, str] = {}
    if "expected_direction" in gold.columns:
        gold["_direction"] = gold["expected_direction"].map(_normalize_direction)
        directions = {
            str(gene): str(direction)
            for gene, direction in zip(gold["_gene"], gold["_direction"], strict=False)
            if str(gene) in positives and str(direction)
        }
    return positives, directions


def _ranked_records(frame: pd.DataFrame, symbol_column: str, rank_column: str, direction_column: str | None) -> list[tuple[str, str]]:
    if symbol_column not in frame.columns:
        raise ValueError(f"result frame missing symbol column {symbol_column!r}")
    if rank_column not in frame.columns:
        raise ValueError(f"result frame missing rank column {rank_column!r}")
    ranked = frame.dropna(subset=[symbol_column, rank_column]).copy()
    ranked[rank_column] = pd.to_numeric(ranked[rank_column], errors="coerce")
    ranked = ranked.dropna(subset=[rank_column]).sort_values([rank_column, symbol_column])
    symbols = ranked[symbol_column].astype(str).str.strip().str.upper().tolist()
    if direction_column and direction_column in ranked.columns:
        directions = ranked[direction_column].map(_normalize_direction).tolist()
    else:
        directions = [""] * len(symbols)
    return list(zip(symbols, directions, strict=False))


def _degora_rankings(score_csv: Path) -> dict[tuple[str, str], list[tuple[str, str]]]:
    frame = pd.read_csv(score_csv)
    rankings = {
        ("degora_deg_score", "v1_2_source_unit_mean"): _ranked_records(
            frame,
            "gene_symbol",
            "degora_rank",
            "consensus_direction",
        )
    }
    if {"quality_weighted_degora_rank", "quality_weighted_consensus_direction"}.issubset(frame.columns):
        rankings[("degora_quality_weighted_score", "quality_weighted_secondary")] = _ranked_records(
            frame,
            "gene_symbol",
            "quality_weighted_degora_rank",
            "quality_weighted_consensus_direction",
        )
    return rankings


def _baseline_rankings(baseline_dir: Path) -> dict[tuple[str, str], list[tuple[str, str]]]:
    rankings: dict[tuple[str, str], list[tuple[str, str]]] = {}
    required = set(BASELINE_RESULT_COLUMNS)
    for path in baseline_result_paths(baseline_dir):
        frame = pd.read_csv(path, sep="\t")
        if frame.empty or not required.issubset(frame.columns):
            continue
        key = (str(frame["method_id"].iloc[0]), str(frame["setting_id"].iloc[0]))
        rankings[key] = _ranked_records(frame, "symbol", "rank", "direction")
    return rankings


def _recall_rows(
    rankings: dict[tuple[str, str], list[tuple[str, str]]],
    positives: set[str],
    expected_directions: dict[str, str],
    *,
    max_k: int,
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    positives_count = len(positives)
    if positives_count == 0:
        raise ValueError("gold panel contains no positive genes")
    direction_count = len(expected_directions)
    for method_id, setting_id in DEFAULT_METHOD_ORDER:
        ranked = rankings.get((method_id, setting_id))
        if not ranked:
            continue
        seen: set[str] = set()
        direction_seen: set[str] = set()
        direction_mismatched: set[str] = set()
        for k, (symbol, direction) in enumerate(ranked[:max_k], start=1):
            if symbol in positives:
                seen.add(symbol)
            expected = expected_directions.get(symbol)
            if expected:
                if direction == expected:
                    direction_seen.add(symbol)
                    direction_mismatched.discard(symbol)
                elif symbol not in direction_seen:
                    direction_mismatched.add(symbol)
            rows.append(
                {
                    "method_id": method_id,
                    "setting_id": setting_id,
                    "method_label": METHOD_LABELS.get((method_id, setting_id), f"{method_id}/{setting_id}"),
                    "k": k,
                    "n_gold": positives_count,
                    "n_recovered": len(seen),
                    "recall": len(seen) / positives_count,
                    "n_direction_gold": direction_count,
                    "n_direction_recovered": len(direction_seen) if direction_count else "",
                    "direction_recall": len(direction_seen) / direction_count if direction_count else "",
                    "n_direction_mismatched": len(direction_mismatched) if direction_count else "",
                }
            )
    return pd.DataFrame(rows)


def _k_summary(source_data: pd.DataFrame, cutoffs: Iterable[int]) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for cutoff in cutoffs:
        subset = source_data.loc[source_data["k"].eq(cutoff)]
        if subset.empty:
            continue
        for row in subset.itertuples(index=False):
            method_curve = source_data.loc[
                source_data["method_label"].eq(row.method_label) & source_data["k"].le(cutoff)
            ]
            aurc = float(pd.to_numeric(method_curve["recall"], errors="coerce").mean())
            direction_values = pd.to_numeric(method_curve["direction_recall"], errors="coerce")
            direction_aurc = "" if direction_values.isna().all() else float(direction_values.mean())
            rows.append(
                {
                    "method_label": row.method_label,
                    "method_id": row.method_id,
                    "setting_id": row.setting_id,
                    "k": int(row.k),
                    "recall": float(row.recall),
                    "direction_recall": row.direction_recall,
                    "aurc_at_k": aurc,
                    "direction_aurc_at_k": direction_aurc,
                    "n_recovered": int(row.n_recovered),
                    "n_gold": int(row.n_gold),
                }
            )
    return pd.DataFrame(rows)


def _write_xlsx(path: Path, curve: pd.DataFrame, summary: pd.DataFrame, metadata: dict[str, object]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "recall_curve"
    for row in dataframe_to_rows(curve, index=False, header=True):
        worksheet.append(row)
    summary_sheet = workbook.create_sheet("selected_cutoffs")
    for row in dataframe_to_rows(summary, index=False, header=True):
        summary_sheet.append(row)
    meta_sheet = workbook.create_sheet("metadata")
    meta_sheet.append(["key", "value"])
    for key, value in metadata.items():
        meta_sheet.append([key, json.dumps(value, sort_keys=True) if isinstance(value, (dict, list)) else value])
    workbook.save(path)


def _plot(curve: pd.DataFrame, output_stem: Path, *, title: str, max_k: int, has_direction: bool) -> None:
    plt.rcParams.update(
        {
            "font.family": "DejaVu Sans",
            "font.size": 9,
            "axes.titlesize": 11,
            "axes.labelsize": 9,
            "legend.fontsize": 8,
            "svg.fonttype": "none",
        }
    )
    n_panels = 2 if has_direction else 1
    fig, axes = plt.subplots(1, n_panels, figsize=(10.5, 4.2), sharey=True)
    if n_panels == 1:
        axes = [axes]
    panels = [("recall", "A. Membership recall")]
    if has_direction:
        panels.append(("direction_recall", "B. Direction-aware recall"))
    for axis, (metric, panel_title) in zip(axes, panels, strict=True):
        for label, group in curve.groupby("method_label", sort=False):
            style = METHOD_STYLES.get(label, {})
            y = pd.to_numeric(group[metric], errors="coerce")
            axis.plot(group["k"], y, label=label, **style)
        axis.set_title(panel_title, loc="left", fontweight="bold")
        axis.set_xlabel("Top K genes")
        axis.set_xlim(1, max_k)
        axis.set_ylim(-0.02, 1.02)
        axis.grid(True, color="#e5e7e9", linewidth=0.8)
        axis.set_axisbelow(True)
        axis.spines["top"].set_visible(False)
        axis.spines["right"].set_visible(False)
    axes[0].set_ylabel("Recall of locked gold panel")
    axes[-1].legend(loc="lower right", frameon=False)
    fig.suptitle(title, x=0.02, y=0.995, ha="left", fontsize=12, fontweight="bold")
    fig.tight_layout(rect=(0, 0, 1, 0.95))
    for suffix in (".png", ".pdf", ".svg"):
        fig.savefig(output_stem.with_suffix(suffix), dpi=300)
    plt.close(fig)


def _docx_escape(text: str) -> str:
    return escape(text).replace("\n", "</w:t><w:br/><w:t>")


def _write_minimal_docx(path: Path, paragraphs: list[str]) -> None:
    content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>
"""
    rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>
"""
    body = "".join(f"<w:p><w:r><w:t>{_docx_escape(paragraph)}</w:t></w:r></w:p>" for paragraph in paragraphs)
    document = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
<w:body>{body}<w:sectPr/></w:body>
</w:document>
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types)
        archive.writestr("_rels/.rels", rels)
        archive.writestr("word/document.xml", document)


def _write_manifest(
    path: Path,
    *,
    figure_id: str,
    command: str,
    inputs: list[Path],
    outputs: list[Path],
    source_data: list[Path],
    max_k: int,
    selected_summary: pd.DataFrame,
) -> dict[str, object]:
    manifest = {
        "figure_id": figure_id,
        "generated_at": datetime.now(UTC).isoformat(),
        "script": "outputs/code/figures/make_gold_recall_curve.py",
        "command": command,
        "inputs": [str(path) for path in inputs],
        "outputs": [str(path) for path in outputs],
        "source_data": [str(path) for path in source_data],
        "panel_claims": [
            "Membership recall was computed cumulatively from each method's ranked gene list.",
            "Direction-aware recall counts locked genes only when the method direction matches expected_direction.",
        ],
        "transformations": {
            "max_k": max_k,
            "gene_ids": "upper-case gene symbols",
            "duplicates": "cumulative set recall; repeated symbols do not add extra recoveries",
            "plotted_methods": [label for label in selected_summary["method_label"].drop_duplicates().tolist()],
        },
        "validation": {
            "selected_cutoffs": selected_summary.to_dict(orient="records"),
        },
        "known_limitations": [
            "This IFN corpus is a derived-count pilot, not an as-published DEG-table benchmark.",
            "Recall is evaluated against a compact locked positive panel and does not measure false positives.",
        ],
    }
    path.write_text(json.dumps(manifest, indent=2, sort_keys=True) + "\n")
    return manifest


def _write_validation(path: Path, outputs: list[Path], selected_summary: pd.DataFrame) -> None:
    lines = ["Gold recall curve validation", ""]
    for artifact in outputs:
        size = artifact.stat().st_size if artifact.exists() else 0
        lines.append(f"{artifact}: exists={artifact.exists()} size_bytes={size}")
    lines.extend(["", "Selected cutoff values:"])
    for row in selected_summary.itertuples(index=False):
        direction = "" if row.direction_recall == "" or (isinstance(row.direction_recall, float) and math.isnan(row.direction_recall)) else f", direction_recall={float(row.direction_recall):.3f}"
        lines.append(f"{row.method_label} k={row.k}: recall={float(row.recall):.3f}{direction}")
    path.write_text("\n".join(lines) + "\n")


def build_figure(
    *,
    baseline_dir: Path,
    gold_path: Path,
    degora_score_csv: Path,
    output_dir: Path,
    figure_id: str,
    output_stem_name: str,
    title: str,
    gold_gene_column: str,
    max_k: int,
    command: str,
) -> dict[str, object]:
    positives, expected_directions = _read_gold(gold_path, gold_gene_column)
    rankings = {**_degora_rankings(degora_score_csv), **_baseline_rankings(baseline_dir)}
    curve = _recall_rows(rankings, positives, expected_directions, max_k=max_k)
    if curve.empty:
        raise ValueError("no plottable rankings were found")
    selected_cutoffs = sorted({10, 20, 50, 100, 200, 500, max_k})
    selected_summary = _k_summary(curve, selected_cutoffs)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_stem = output_dir / output_stem_name
    source_csv = output_dir / f"{output_stem_name}_source_data.csv"
    source_xlsx = output_dir / f"{output_stem_name}_source_data.xlsx"
    summary_csv = output_dir / f"{output_stem_name}_selected_cutoffs.csv"
    legend_docx = output_dir / f"{output_stem_name}_legend.docx"
    manifest_json = output_dir / f"{output_stem_name}_manifest.json"
    validation_txt = output_dir / f"{output_stem_name}_validation.txt"

    curve.to_csv(source_csv, index=False)
    selected_summary.to_csv(summary_csv, index=False)
    _write_xlsx(
        source_xlsx,
        curve,
        selected_summary,
        {
            "figure_id": figure_id,
            "gold_path": str(gold_path),
            "degora_score_csv": str(degora_score_csv),
            "baseline_dir": str(baseline_dir),
            "max_k": max_k,
        },
    )
    _plot(curve, output_stem, title=title, max_k=max_k, has_direction=bool(expected_directions))
    _write_minimal_docx(
        legend_docx,
        [
            f"{figure_id}. Locked gold-panel recall through top {max_k}.",
            "A, Cumulative recall of locked positive genes across ranked outputs. B, Direction-aware recall for panels with expected_direction annotations. Lines show DEGORA primary, DEGORA quality-weighted secondary, and runnable classical baselines.",
        ],
    )

    figure_outputs = [output_stem.with_suffix(suffix) for suffix in (".png", ".pdf", ".svg")]
    outputs = [*figure_outputs, source_csv, source_xlsx, summary_csv, legend_docx, manifest_json, validation_txt]
    _write_validation(validation_txt, [*figure_outputs, source_csv, source_xlsx, summary_csv, legend_docx], selected_summary)
    _write_manifest(
        manifest_json,
        figure_id=figure_id,
        command=command,
        inputs=[baseline_dir, gold_path, degora_score_csv],
        outputs=outputs,
        source_data=[source_csv, source_xlsx, summary_csv],
        max_k=max_k,
        selected_summary=selected_summary,
    )
    for artifact in outputs:
        write_source_sidecar(
            artifact,
            command,
            inputs=[baseline_dir, gold_path, degora_score_csv],
            metadata={"generator": "gold-recall-curve", "figure_id": figure_id, "max_k": max_k},
        )
    return {
        "figure_id": figure_id,
        "curve_rows": int(len(curve)),
        "plotted_methods": curve["method_label"].drop_duplicates().tolist(),
        "outputs": [str(path) for path in outputs],
        "selected_cutoffs": selected_summary.to_dict(orient="records"),
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--baseline-dir", type=Path, required=True)
    parser.add_argument("--gold", type=Path, required=True)
    parser.add_argument("--gold-gene-column", default="gene_symbol")
    parser.add_argument("--degora-score-csv", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--figure-id", required=True)
    parser.add_argument("--output-stem", required=True)
    parser.add_argument("--title", required=True)
    parser.add_argument("--max-k", type=int, default=1000)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    command = shell_command(
        [
            "PYTHONPATH=outputs/code",
            "python",
            "outputs/code/figures/make_gold_recall_curve.py",
            "--baseline-dir",
            args.baseline_dir,
            "--gold",
            args.gold,
            "--gold-gene-column",
            args.gold_gene_column,
            "--degora-score-csv",
            args.degora_score_csv,
            "--output-dir",
            args.output_dir,
            "--figure-id",
            args.figure_id,
            "--output-stem",
            args.output_stem,
            "--title",
            args.title,
            "--max-k",
            args.max_k,
        ]
    )
    summary = build_figure(
        baseline_dir=args.baseline_dir,
        gold_path=args.gold,
        degora_score_csv=args.degora_score_csv,
        output_dir=args.output_dir,
        figure_id=args.figure_id,
        output_stem_name=args.output_stem,
        title=args.title,
        gold_gene_column=args.gold_gene_column,
        max_k=args.max_k,
        command=command,
    )
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
