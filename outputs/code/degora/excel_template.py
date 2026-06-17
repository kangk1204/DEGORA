"""Excel template generation for beginner DEGORA runs."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


TEMPLATE_SHEETS = ["README", "Project", "Contrasts", "GoldPanel", "AdvancedSettings", "ColumnGuide"]
SHEET_NOTES = {
    "README": "# Start here: quick steps for editing and running this workbook.",
    "Project": "# Project-level settings. Defaults are OK for a first run.",
    "Contrasts": "# Main input: one row per DEG table or contrast; required columns are gene/effect/p-value mappings.",
    "GoldPanel": "# Optional: known marker genes for recall checks only; leave empty if unavailable.",
    "AdvancedSettings": "# Optional advanced settings. Most users can keep the defaults.",
    "ColumnGuide": "# Reference guide for required and optional columns.",
}


def _project_rows() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"field": "project_name", "value": "my_degora_project", "what_to_enter": "Short name for this analysis."},
            {"field": "topic", "value": "interferon response", "what_to_enter": "Biological topic or keyword."},
            {"field": "organism", "value": "Homo sapiens", "what_to_enter": "Main organism in the DEG tables."},
            {
                "field": "output_dir",
                "value": "outputs/results/degora-run",
                "what_to_enter": "Folder where DEGORA will write results.",
            },
            {
                "field": "harmonized_dir",
                "value": "data/deg/harmonized",
                "what_to_enter": "Folder for harmonized intermediate DEG tables.",
            },
            {
                "field": "min_studies",
                "value": 2,
                "what_to_enter": "Minimum number of independent source units needed for a gene to be scored.",
            },
        ]
    )


def _contrast_rows() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "study_id": "IFN_GSE001_4h",
                "source_unit_id": "GSE001",
                "source_path": "data/deg/raw/example_ifn_4h.csv",
                "condition": "IFN-beta vs untreated",
                "time_h": 4,
                "time_course_mode": "mean",
                "cell_system": "NHBE",
                "species": "Homo sapiens",
                "pipeline": "DESeq2",
                "assay_type": "RNA-seq",
                "source_input_type": "author_deg_table",
                "platform": "",
                "normalization": "DESeq2",
                "probe_id_column": "",
                "probe_collapse": "",
                "gene_column": "gene",
                "lfc_column": "log2FoldChange",
                "p_column": "pvalue",
                "padj_column": "padj",
                "table_scope": "auto",
                "rank_universe_size": "",
                "sep": "",
                "sheet_name": "",
                "gene_type_column": "",
                "gene_type_keep": "",
                "include": "yes",
                "notes": "Replace this example row with your DEG table.",
            },
            {
                "study_id": "IFN_GSE001_12h",
                "source_unit_id": "GSE001",
                "source_path": "data/deg/raw/example_ifn_12h.csv",
                "condition": "IFN-beta vs untreated",
                "time_h": 12,
                "time_course_mode": "mean",
                "cell_system": "NHBE",
                "species": "Homo sapiens",
                "pipeline": "DESeq2",
                "assay_type": "RNA-seq",
                "source_input_type": "author_deg_table",
                "platform": "",
                "normalization": "DESeq2",
                "probe_id_column": "",
                "probe_collapse": "",
                "gene_column": "gene",
                "lfc_column": "log2FoldChange",
                "p_column": "pvalue",
                "padj_column": "padj",
                "table_scope": "auto",
                "rank_universe_size": "",
                "sep": "",
                "sheet_name": "",
                "gene_type_column": "",
                "gene_type_keep": "",
                "include": "yes",
                "notes": "Same source_unit_id means same independent paper/dataset.",
            },
            {
                "study_id": "EXAMPLE_MICROARRAY_24h",
                "source_unit_id": "GSE_MICROARRAY",
                "source_path": "data/deg/raw/example_microarray_limma.csv",
                "condition": "drug vs vehicle",
                "time_h": 24,
                "time_course_mode": "mean",
                "cell_system": "cell line",
                "species": "Homo sapiens",
                "pipeline": "limma_microarray",
                "assay_type": "microarray",
                "source_input_type": "limma_full_table",
                "platform": "GPL570",
                "normalization": "RMA/log2",
                "probe_id_column": "probe_id",
                "probe_collapse": "min_pvalue_max_abs_lfc",
                "gene_column": "gene_symbol",
                "lfc_column": "logFC",
                "p_column": "P.Value",
                "padj_column": "adj.P.Val",
                "table_scope": "full_results",
                "rank_universe_size": "",
                "sep": "",
                "sheet_name": "",
                "gene_type_column": "",
                "gene_type_keep": "",
                "include": "no",
                "notes": "Microarray example only. Set include=yes after you create or obtain this full DEG table.",
            },
        ]
    )


def _gold_rows() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "gene_symbol": "ISG15",
                "expected_direction": "up",
                "role": "optional_marker",
                "evidence_basis": "Example IFN marker; set locked=yes only after replacing with your panel.",
                "locked": "no",
            },
            {
                "gene_symbol": "IFIT1",
                "expected_direction": "up",
                "role": "optional_marker",
                "evidence_basis": "Example IFN marker; set locked=yes only after replacing with your panel.",
                "locked": "no",
            },
        ]
    )


def _advanced_rows() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "setting": "score_version",
                "value": "degora_score_v1",
                "what_it_does": "Transparent prioritization score; not a probability.",
            },
            {"setting": "browser_port", "value": 8765, "what_it_does": "Local browser/API port used by degora serve."},
            {
                "setting": "collapse_independence_by",
                "value": "source_unit_id",
                "what_it_does": "Keep the same source_unit_id for related time points from one study.",
            },
        ]
    )


def _readme_rows() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"step": 1, "instruction": "Fill the Project sheet. The defaults are OK for a first run."},
            {"step": 2, "instruction": "Fill the Contrasts sheet. Use one row per DEG table or time point."},
            {"step": 3, "instruction": "Required Contrasts columns are marked in ColumnGuide; padj_column is optional."},
            {
                "step": 4,
                "instruction": "gene_column, lfc_column, and p_column must name exact columns inside each source DEG table.",
            },
            {"step": 5, "instruction": "Keep source_unit_id the same for rows from the same paper or dataset."},
            {"step": 6, "instruction": "If the paper only gives significant genes, set table_scope to deg_only."},
            {"step": 7, "instruction": "For microarray DEG tables, set assay_type=microarray and pipeline=limma_microarray."},
            {"step": 8, "instruction": "Run: degora validate degora_config.xlsx"},
            {"step": 9, "instruction": "Run: degora run degora_config.xlsx"},
            {"step": 10, "instruction": "Run: degora serve outputs/results/degora-run/degora_scores.db"},
        ]
    )


def _guide_rows() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "column": "study_id",
                "required": "yes",
                "checked_where": "Contrasts sheet",
                "meaning": "Unique ID for this contrast row.",
            },
            {
                "column": "source_unit_id",
                "required": "yes",
                "checked_where": "Contrasts sheet",
                "meaning": "Independent source ID. Reuse it for time points from the same paper/dataset.",
            },
            {
                "column": "source_path",
                "required": "yes",
                "checked_where": "Contrasts sheet",
                "meaning": "Path to your source DEG table file.",
            },
            {
                "column": "gene_column",
                "required": "yes",
                "checked_where": "Contrasts sheet and source DEG table",
                "meaning": "Exact source-table column containing gene symbols or IDs.",
            },
            {
                "column": "lfc_column",
                "required": "yes",
                "checked_where": "Contrasts sheet and source DEG table",
                "meaning": "Exact source-table column containing numeric log2 fold change.",
            },
            {
                "column": "p_column",
                "required": "yes",
                "checked_where": "Contrasts sheet and source DEG table",
                "meaning": "Exact source-table column containing p-values in [0, 1]; not -log10(p), a statistic, or a percent.",
            },
            {
                "column": "padj_column",
                "required": "no; checked if filled",
                "checked_where": "source DEG table only when filled",
                "meaning": "Optional exact source-table column containing adjusted p-values/FDR in [0, 1]. Leave blank if unavailable.",
            },
            {
                "column": "assay_type",
                "required": "no",
                "checked_where": "metadata",
                "meaning": "RNA-seq, microarray, or other assay label. Use microarray for array evidence.",
            },
            {
                "column": "source_input_type",
                "required": "no",
                "checked_where": "metadata",
                "meaning": "author_deg_table, limma_full_table, derived_count_table, or normalized_expression_matrix.",
            },
            {
                "column": "platform",
                "required": "no",
                "checked_where": "metadata",
                "meaning": "Microarray platform such as GPL570, GPL96, or blank.",
            },
            {
                "column": "normalization",
                "required": "no",
                "checked_where": "metadata",
                "meaning": "Normalization used by the source, such as RMA/log2 or quantile/log2.",
            },
            {
                "column": "probe_id_column",
                "required": "no",
                "checked_where": "metadata",
                "meaning": "Optional probe ID column for microarray-derived DEG tables.",
            },
            {
                "column": "probe_collapse",
                "required": "no",
                "checked_where": "metadata",
                "meaning": "How probes were collapsed to gene symbols.",
            },
            {
                "column": "table_scope",
                "required": "no",
                "checked_where": "validated setting",
                "meaning": "Use full_results for all tested genes, deg_only for significant-gene-only lists, or auto.",
            },
            {
                "column": "rank_universe_size",
                "required": "no",
                "checked_where": "validated setting if filled",
                "meaning": "For deg_only lists, number of genes originally tested if the paper reports it.",
            },
            {
                "column": "condition",
                "required": "no",
                "checked_where": "metadata",
                "meaning": "Human-readable contrast label.",
            },
            {
                "column": "time_h",
                "required": "no",
                "checked_where": "metadata",
                "meaning": "Time point in hours for time-course data.",
            },
            {
                "column": "time_course_mode",
                "required": "no",
                "checked_where": "validated setting if filled",
                "meaning": "mean, early, late, or peak_mean for rows sharing one source_unit_id.",
            },
            {
                "column": "sep",
                "required": "no",
                "checked_where": "source reader if filled",
                "meaning": "Delimiter override for text tables, such as \\t for TSV when auto-detection fails.",
            },
            {
                "column": "sheet_name",
                "required": "no",
                "checked_where": "source reader if filled",
                "meaning": "Sheet name for Excel source DEG tables.",
            },
            {
                "column": "gene_type_column",
                "required": "no",
                "checked_where": "source DEG table only when filled",
                "meaning": "Optional exact source-table column used with gene_type_keep to filter gene biotypes.",
            },
            {
                "column": "gene_type_keep",
                "required": "no",
                "checked_where": "source DEG table only when gene_type_column is filled",
                "meaning": "Pipe-separated values to keep from gene_type_column, such as protein_coding|lncRNA.",
            },
            {"column": "cell_system", "required": "no", "checked_where": "metadata", "meaning": "Cell type, tissue, or experimental system for the contrast."},
            {"column": "species", "required": "no", "checked_where": "metadata", "meaning": "Species of the source data, such as Homo sapiens."},
            {"column": "pipeline", "required": "no", "checked_where": "metadata", "meaning": "DEG analysis pipeline or method, such as DESeq2 or limma_microarray."},
            {"column": "notes", "required": "no", "checked_where": "metadata", "meaning": "Free-text note carried through to provenance; not used in scoring."},
            {"column": "include", "required": "no", "checked_where": "validated setting", "meaning": "yes to include, no to exclude."},
        ]
    )


def _autosize_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    note_fill = PatternFill("solid", fgColor="FFF2CC")
    for worksheet in workbook.worksheets:
        has_note = isinstance(worksheet["A1"].value, str) and worksheet["A1"].value.strip().startswith("#")
        header_row = 2 if has_note else 1
        worksheet.freeze_panes = "A3" if has_note else "A2"
        if has_note:
            worksheet["A1"].font = Font(italic=True, color="7F6000")
            worksheet["A1"].fill = note_fill
        for cell in worksheet[header_row]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for column_cells in worksheet.columns:
            values: list[Any] = [cell.value for cell in column_cells]
            width = min(max(len(str(value)) if value is not None else 0 for value in values) + 2, 60)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max(width, 12)
    workbook.save(path)


def write_note_sheet(writer: pd.ExcelWriter, sheet_name: str, frame: pd.DataFrame) -> None:
    """Write a config sheet with a human-readable '#'-note row above headers."""

    frame.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
    worksheet = writer.sheets[sheet_name]
    worksheet["A1"] = SHEET_NOTES.get(sheet_name, "# DEGORA configuration sheet.")


def write_template(path: str | Path, *, force: bool = False) -> Path:
    """Write a beginner-facing Excel config template."""

    output = Path(path)
    if output.exists() and not force:
        raise FileExistsError(f"Template already exists: {output}. Use --force to overwrite it.")
    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        write_note_sheet(writer, "README", _readme_rows())
        write_note_sheet(writer, "Project", _project_rows())
        write_note_sheet(writer, "Contrasts", _contrast_rows())
        write_note_sheet(writer, "GoldPanel", _gold_rows())
        write_note_sheet(writer, "AdvancedSettings", _advanced_rows())
        write_note_sheet(writer, "ColumnGuide", _guide_rows())
    _autosize_workbook(output)
    return output.resolve()
