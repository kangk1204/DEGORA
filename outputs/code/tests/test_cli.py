from __future__ import annotations

import sqlite3

import pandas as pd
from openpyxl import load_workbook

from degora.cli import _print_run_warnings, main
from degora.excel_template import TEMPLATE_SHEETS, write_template


def _write_source(path, genes, lfc_scale: float) -> None:
    pd.DataFrame(
        {
            "gene": genes,
            "log2FoldChange": [2.0 * lfc_scale, 1.5 * lfc_scale, 0.1],
            "pvalue": [1e-6, 1e-4, 0.8],
            "padj": [1e-5, 1e-3, 0.9],
        }
    ).to_csv(path, index=False)


def _write_config(path, source_a, source_b) -> None:
    project = pd.DataFrame(
        {
            "field": ["output_dir", "harmonized_dir", "min_studies"],
            "value": [str(path.parent / "results"), str(path.parent / "harmonized"), 1],
        }
    )
    contrasts = pd.DataFrame(
        {
            "study_id": ["S1_4h", "S2_4h"],
            "source_unit_id": ["P1", "P2"],
            "source_path": [str(source_a), str(source_b)],
            "gene_column": ["gene", "gene"],
            "lfc_column": ["log2FoldChange", "log2FoldChange"],
            "p_column": ["pvalue", "pvalue"],
            "padj_column": ["padj", "padj"],
            "include": ["yes", "yes"],
        }
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        project.to_excel(writer, sheet_name="Project", index=False)
        contrasts.to_excel(writer, sheet_name="Contrasts", index=False)


def test_template_command_writes_beginner_workbook(tmp_path) -> None:
    output = tmp_path / "DEGORA_template.xlsx"

    assert main(["template", str(output)]) == 0

    workbook = load_workbook(output)
    assert workbook.sheetnames == TEMPLATE_SHEETS
    assert workbook["Contrasts"]["A1"].value.startswith("#")
    assert workbook["Contrasts"]["A2"].value == "study_id"
    assert workbook["ColumnGuide"]["A1"].value.startswith("#")
    assert workbook["ColumnGuide"]["A2"].value == "column"
    assert workbook["ColumnGuide"]["C2"].value == "checked_where"
    assert workbook["Contrasts"].freeze_panes == "A3"
    contrast_headers = [cell.value for cell in workbook["Contrasts"][2]]
    guide_columns = [cell.value for cell in workbook["ColumnGuide"]["A"]]
    guide_required = {row[0].value: row[1].value for row in workbook["ColumnGuide"].iter_rows(min_row=3)}
    assert "time_course_mode" in contrast_headers
    assert "time_course_mode" in guide_columns
    assert guide_required["p_column"] == "yes"
    assert guide_required["padj_column"] == "no; checked if filled"


def test_demo_command_writes_runnable_workspace(tmp_path, capsys) -> None:
    demo_dir = tmp_path / "demo"

    assert main(["demo", str(demo_dir)]) == 0

    config = demo_dir / "degora_demo_config.xlsx"
    assert config.exists()
    assert (demo_dir / "deg_tables" / "demo_ifn_a_4h.csv").exists()
    assert (demo_dir / "README.md").exists()
    assert main(["validate", str(config)]) == 0
    assert main(["run", str(config)]) == 0
    assert "non-fatal input warnings" not in capsys.readouterr().err

    db = demo_dir / "results" / "degora_scores.db"
    assert db.exists()
    workbook_path = demo_dir / "results" / "DEGORA_output.xlsx"
    assert workbook_path.exists()
    workbook = load_workbook(workbook_path)
    assert workbook.sheetnames[:2] == ["Workbook_guide", "Column_dictionary"]
    assert {"Run_summary", "Gene_scores", "Gene_evidence", "Source_units"}.issubset(set(workbook.sheetnames))
    guide_rows = list(workbook["Workbook_guide"].iter_rows(min_row=2, values_only=True))
    assert any(row[0] == "Gene_scores" and "main prioritized gene list" in row[3] for row in guide_rows)
    dictionary_rows = list(workbook["Column_dictionary"].iter_rows(min_row=2, values_only=True))
    assert any(
        row[0] == "Gene_scores"
        and row[2] == "quality_weighted_degora_score"
        and "relative index, not a probability" in row[3]
        for row in dictionary_rows
    )
    assert workbook["Gene_scores"].freeze_panes == "A2"
    gene_score_headers = {cell.value: cell for cell in workbook["Gene_scores"][1]}
    assert gene_score_headers["quality_weighted_degora_score"].comment is not None
    assert "Values:" in gene_score_headers["quality_weighted_degora_score"].comment.text
    assert workbook["Column_dictionary"]["A1"].comment is not None
    with sqlite3.connect(db) as connection:
        top_genes = [
            row[0]
            for row in connection.execute("SELECT gene_symbol FROM genes ORDER BY degora_rank LIMIT 4").fetchall()
        ]
        source_units = connection.execute("SELECT COUNT(DISTINCT source_unit_id) FROM studies").fetchone()[0]
    assert top_genes[0] == "ISG15"
    assert {"IFIT1", "MX1"}.issubset(set(top_genes))
    assert source_units == 2


def test_validate_command_accepts_excel_config(tmp_path, capsys) -> None:
    source = tmp_path / "source.csv"
    _write_source(source, ["ISG15", "IFIT1", "RPL13A"], 1.0)
    config = tmp_path / "config.xlsx"
    _write_config(config, source, source)

    assert main(["validate", str(config)]) == 0
    captured = capsys.readouterr()
    assert "Required Contrasts columns" in captured.out
    assert "source_unit_id (or paper_id)" in captured.out
    assert "Required DEG-table mappings" in captured.out
    assert "Optional DEG-table mappings checked when filled" in captured.out
    assert "padj_column -> adjusted p-value/FDR" in captured.out


def test_run_command_builds_score_database_from_excel_config(tmp_path) -> None:
    source_a = tmp_path / "source_a.csv"
    source_b = tmp_path / "source_b.csv"
    _write_source(source_a, ["ISG15", "IFIT1", "RPL13A"], 1.0)
    _write_source(source_b, ["ISG15", "IFIT1", "RPL13A"], 0.8)
    config = tmp_path / "config.xlsx"
    _write_config(config, source_a, source_b)
    db = tmp_path / "results" / "degora_scores.db"

    assert main(["run", str(config), "--db", str(db)]) == 0

    assert db.exists()
    assert (tmp_path / "results" / "degora_gene_scores.csv").exists()
    assert (tmp_path / "results" / "degora_score_metadata.json").exists()
    assert (tmp_path / "results" / "DEGORA_output.xlsx").exists()
    with sqlite3.connect(db) as connection:
        top_gene = connection.execute("SELECT gene_symbol FROM genes ORDER BY degora_rank LIMIT 1").fetchone()[0]
        source_units = connection.execute("SELECT COUNT(DISTINCT source_unit_id) FROM studies").fetchone()[0]
    assert top_gene == "ISG15"
    assert source_units == 2


def test_run_command_can_skip_default_excel_export(tmp_path) -> None:
    source_a = tmp_path / "source_a.csv"
    source_b = tmp_path / "source_b.csv"
    _write_source(source_a, ["ISG15", "IFIT1", "RPL13A"], 1.0)
    _write_source(source_b, ["ISG15", "IFIT1", "RPL13A"], 0.8)
    config = tmp_path / "config.xlsx"
    _write_config(config, source_a, source_b)

    assert main(["run", str(config), "--no-excel"]) == 0

    assert (tmp_path / "results" / "degora_scores.db").exists()
    assert not (tmp_path / "results" / "DEGORA_output.xlsx").exists()


def test_validate_missing_config_returns_clean_error(tmp_path, capsys) -> None:
    exit_code = main(["validate", str(tmp_path / "does_not_exist.xlsx")])

    # Beginner-facing error contract: a clear message and exit code 2, not a raw traceback.
    assert exit_code == 2
    assert "config file was not found" in capsys.readouterr().err


def test_run_warnings_are_printed_to_stderr(tmp_path, capsys) -> None:
    metrics_path = tmp_path / "slice_metrics.json"

    _print_run_warnings(
        {
            "warnings": ["S1 produced zero harmonized rows", "S1 produced zero harmonized rows"],
            "rank_universe_warnings": ["DEG-only table without rank_universe_size"],
            "pvalue_clipped_rows": 2,
        },
        metrics_path=metrics_path,
    )

    captured = capsys.readouterr()
    assert captured.out == ""
    assert "DEGORA completed with non-fatal input warnings" in captured.err
    assert captured.err.count("S1 produced zero harmonized rows") == 1
    assert "DEG-only table without rank_universe_size" in captured.err
    assert "2 row(s) reported pvalue == 0" in captured.err
    assert str(metrics_path) in captured.err
