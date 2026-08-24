"""Species-scoped discovery of author-supplied GEO DEG-table candidates.

Discovery deliberately stops before scoring.  It searches GEO, records an audit
ledger, and can prepare an excluded draft catalog for manual contrast review.
Human and mouse are separate evidence universes; this module never pools them.
"""

from __future__ import annotations

import copy
import csv
import gzip
import hashlib
import http.client
import io
import json
import math
import os
import re
import shutil
import tempfile
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
import zipfile
import zlib
from collections import OrderedDict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable

from . import runtime_version_info


EUTILS_BASE = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils"
GEO_ACCESSION_URL = "https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi"
ALLOWED_NCBI_HOSTS = frozenset(
    {
        "eutils.ncbi.nlm.nih.gov",
        "www.ncbi.nlm.nih.gov",
        "ftp.ncbi.nlm.nih.gov",
    }
)
USER_AGENT = "DEGORA-discovery/0.4 (academic GEO candidate discovery)"
DEFAULT_LIMIT = 10
MAX_LIMIT = 20
DEFAULT_PAGE_SIZE = 10
MAX_PAGE = 500
MAX_SELECTED_STUDIES = 20
MAX_QUERY_LENGTH = 200
MAX_QUERY_TERMS = 12
MAX_SOFT_BYTES = 2 * 1024 * 1024
MAX_JSON_BYTES = 8 * 1024 * 1024
MAX_TEXT_PREFIX_BYTES = 768 * 1024
MAX_CANDIDATE_BYTES = 25 * 1024 * 1024
MAX_DECOMPRESSED_BYTES = 5 * 1024 * 1024
MAX_GZIP_UNCOMPRESSED_BYTES = 200 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_XLSX_MEMBER_BYTES = 50 * 1024 * 1024
MAX_XLSX_MEMBERS = 5_000
DEFAULT_INSPECTION_BUDGET = 6
DEFAULT_TIMEOUT_SECONDS = 30
DEFAULT_PACE_SECONDS = 0.35
API_KEY_PACE_SECONDS = 0.11
DEFAULT_RETRIES = 3
DEFAULT_CACHE_SIZE = 64
MAX_CACHE_ENTRY_BYTES = 2 * 1024 * 1024
EXACT_PAGE_FETCH_SIZE = 100
MAX_EXACT_PAGE_SCAN_RECORDS = 10_000
DEFAULT_GLOBAL_RANK_LIMIT = 1_000
MAX_GLOBAL_RANK_LIMIT = 1_000
GLOBAL_SEARCH_CACHE_SIZE = 8
GLOBAL_SEARCH_CACHE_TTL_SECONDS = 15 * 60
PUBMED_SUMMARY_BATCH_SIZE = 200
SEARCH_ASSESSMENT_VERSION = 1
DISCOVERY_BUNDLE_MARKER = ".degora-discovery-bundle.json"
DISCOVERY_BUNDLE_ARTIFACT_TYPE = "degora_discovery_prepared_bundle"
DISCOVERY_BUNDLE_FORMAT_VERSION = 1
SEARCH_ASSESSMENT_BASIS = (
    "Search-time GEO supplementary filename assessment only; "
    "Prepare selection downloads and validates file content."
)
DEG_INPUT_PRIORITIES = {
    "author_deg_likely": 4,
    "tabular_candidate": 3,
    "matrix_fallback": 2,
    "not_detected": 1,
    "unresolved": 0,
}
DEG_INPUT_LABELS = {
    "author_deg_likely": "Likely author DEG table",
    "tabular_candidate": "Table to inspect",
    "matrix_fallback": "Matrix fallback candidate",
    "not_detected": "No DEG-like file detected",
    "unresolved": "File availability unresolved",
}


class DiscoveryError(ValueError):
    """Invalid discovery input or unsafe candidate content."""


class DiscoveryUnsafeArchiveError(DiscoveryError):
    """An archive broke a safety rule: path escape, symlink, or a size/count cap.

    Kept distinct from an unreadable or missing file. A malformed download is
    one study's problem; a hostile archive is the operator's, and preparation
    still refuses the whole run rather than quietly skipping it.
    """


class DiscoveryUnavailableError(RuntimeError):
    """A remote discovery source could not be resolved after bounded retries."""


@dataclass(frozen=True)
class SpeciesSpec:
    key: str
    label: str
    scientific_name: str


@dataclass(frozen=True)
class NcbiRequestConfig:
    """Process-scoped NCBI identity and request-rate policy.

    Credentials are deliberately absent from browser state and generated
    artifacts.  ``api_key`` is excluded from ``repr`` so routine diagnostics
    cannot disclose it accidentally.
    """

    tool: str = "degora"
    email: str = ""
    api_key: str = field(default="", repr=False)

    def __post_init__(self) -> None:
        tool = str(self.tool).strip()
        email = str(self.email).strip()
        api_key = str(self.api_key).strip()
        if not re.fullmatch(r"[A-Za-z0-9_.-]{1,80}", tool):
            raise DiscoveryError("NCBI tool must contain only letters, numbers, dot, underscore, or hyphen")
        if email and (len(email) > 254 or email.count("@") != 1 or any(char.isspace() for char in email)):
            raise DiscoveryError("NCBI email must be a valid single contact address")
        if api_key and (len(api_key) > 256 or not re.fullmatch(r"[A-Za-z0-9_.-]+", api_key)):
            raise DiscoveryError("NCBI API key contains unsupported characters")
        object.__setattr__(self, "tool", tool)
        object.__setattr__(self, "email", email)
        object.__setattr__(self, "api_key", api_key)

    @classmethod
    def from_environment(cls) -> "NcbiRequestConfig":
        return cls(
            tool=os.environ.get("NCBI_TOOL", "degora"),
            email=os.environ.get("NCBI_EMAIL", ""),
            api_key=os.environ.get("NCBI_API_KEY", ""),
        )

    @property
    def pace_seconds(self) -> float:
        return API_KEY_PACE_SECONDS if self.api_key else DEFAULT_PACE_SECONDS

    @property
    def request_ceiling_per_second(self) -> int:
        return 10 if self.api_key else 3

    def eutils_params(self) -> dict[str, str]:
        params = {"tool": self.tool}
        if self.email:
            params["email"] = self.email
        if self.api_key:
            params["api_key"] = self.api_key
        return params


SPECIES_BY_KEY = {
    "human": SpeciesSpec("human", "Human", "Homo sapiens"),
    "mouse": SpeciesSpec("mouse", "Mouse", "Mus musculus"),
}
SPECIES_ALIASES = {
    "human": "human",
    "homo sapiens": "human",
    "mouse": "mouse",
    "mus musculus": "mouse",
}


def normalize_species(value: str) -> SpeciesSpec:
    key = SPECIES_ALIASES.get(str(value).strip().lower())
    if key is None:
        raise DiscoveryError("species must be exactly human or mouse; cross-species pooling is not supported")
    return SPECIES_BY_KEY[key]


def _query_terms(query: str) -> list[str]:
    text = str(query).strip()
    if not text:
        raise DiscoveryError("discovery query is required")
    if len(text) > MAX_QUERY_LENGTH:
        raise DiscoveryError(f"discovery query is too long; maximum length is {MAX_QUERY_LENGTH} characters")
    if any(ord(char) < 32 for char in text):
        raise DiscoveryError("discovery query contains control characters")
    # Natural-language search only.  Removing Entrez syntax characters prevents a
    # user query from weakening the mandatory organism and GSE filters.
    terms = re.findall(r"[^\W_]+(?:[-'][^\W_]+)*", text, flags=re.UNICODE)
    terms = [term[:64] for term in terms if term]
    if not terms:
        raise DiscoveryError("discovery query must contain at least one word or number")
    if len(terms) > MAX_QUERY_TERMS:
        raise DiscoveryError(f"discovery query has too many terms; maximum is {MAX_QUERY_TERMS}")
    return terms


def build_geo_query(query: str, species: str | SpeciesSpec) -> str:
    spec = species if isinstance(species, SpeciesSpec) else normalize_species(species)
    topic = " AND ".join(f'"{term}"[All Fields]' for term in _query_terms(query))
    assay = (
        '("Expression profiling by high throughput sequencing"[Filter] OR '
        '"Expression profiling by array"[Filter])'
    )
    return f"({topic}) AND gse[ETYP] AND \"{spec.scientific_name}\"[Organism] AND {assay}"


def normalize_ncbi_url(url: str) -> str:
    parsed = urllib.parse.urlsplit(str(url).strip())
    scheme = parsed.scheme.lower()
    host = (parsed.hostname or "").lower()
    if scheme == "ftp" and host == "ftp.ncbi.nlm.nih.gov":
        scheme = "https"
    if scheme != "https":
        raise DiscoveryError("discovery URLs must use HTTPS")
    if host not in ALLOWED_NCBI_HOSTS:
        raise DiscoveryError(f"discovery URL host is not allowed: {host or '(missing)'}")
    if parsed.username or parsed.password:
        raise DiscoveryError("discovery URLs must not contain user information")
    if parsed.port not in (None, 443):
        raise DiscoveryError("discovery URLs must use the default HTTPS port")
    return urllib.parse.urlunsplit((scheme, host, parsed.path or "/", parsed.query, ""))


class _ValidatedRedirectHandler(urllib.request.HTTPRedirectHandler):
    def redirect_request(self, req, fp, code, msg, headers, newurl):  # type: ignore[override]
        safe_url = normalize_ncbi_url(urllib.parse.urljoin(req.full_url, newurl))
        return super().redirect_request(req, fp, code, msg, headers, safe_url)


class SafeNcbiTransport:
    """Bounded HTTPS transport that cannot be redirected outside NCBI."""

    def __init__(self) -> None:
        self._opener = urllib.request.build_opener(_ValidatedRedirectHandler())

    def __call__(self, url: str, headers: dict[str, str], timeout: int, max_bytes: int) -> bytes:
        safe_url = normalize_ncbi_url(url)
        request = urllib.request.Request(safe_url, headers={"User-Agent": USER_AGENT, **headers})
        with self._opener.open(request, timeout=timeout) as response:
            normalize_ncbi_url(response.geturl())
            declared = response.headers.get("Content-Length")
            if declared and "Range" not in headers:
                try:
                    declared_bytes = int(declared)
                except ValueError:
                    declared_bytes = None
                if declared_bytes is not None and declared_bytes > max_bytes:
                    raise DiscoveryError(f"remote response exceeds the {max_bytes}-byte safety cap")
            chunks: list[bytes] = []
            total = 0
            while True:
                chunk = response.read(min(128 * 1024, max_bytes + 1 - total))
                if not chunk:
                    break
                total += len(chunk)
                if total > max_bytes:
                    raise DiscoveryError(f"remote response exceeds the {max_bytes}-byte safety cap")
                chunks.append(chunk)
            return b"".join(chunks)


Transport = Callable[[str, dict[str, str], int, int], bytes]


class NcbiGeoClient:
    """Small NCBI client with serialized pacing, retries, and success-only LRU cache."""

    def __init__(
        self,
        *,
        request_config: NcbiRequestConfig | None = None,
        transport: Transport | None = None,
        pace_seconds: float | None = None,
        retries: int = DEFAULT_RETRIES,
        cache_size: int = DEFAULT_CACHE_SIZE,
        sleep: Callable[[float], None] = time.sleep,
        monotonic: Callable[[], float] = time.monotonic,
    ) -> None:
        self.request_config = request_config or NcbiRequestConfig.from_environment()
        self.transport = transport or SafeNcbiTransport()
        requested_pace = self.request_config.pace_seconds if pace_seconds is None else pace_seconds
        if transport is None and float(requested_pace) < self.request_config.pace_seconds:
            raise DiscoveryError("NCBI pacing cannot exceed the configured E-utilities request ceiling")
        self.pace_seconds = max(float(requested_pace), 0.0)
        self.retries = max(int(retries), 1)
        self.cache_size = max(int(cache_size), 0)
        self.sleep = sleep
        self.monotonic = monotonic
        self._cache: OrderedDict[tuple[str, tuple[tuple[str, str], ...], int], bytes] = OrderedDict()
        self._lock = threading.Lock()
        self._last_request = 0.0
        self._global_search_cache: OrderedDict[
            tuple[str, str, int, int], tuple[float, dict[str, Any]]
        ] = OrderedDict()
        self._global_search_cache_lock = threading.Lock()
        self._global_search_build_lock = threading.Lock()

    def _pace(self) -> None:
        # NCBI's rate policy bounds how often requests are *issued*, so anchor the
        # interval to the moment this request starts. Stamping on completion (the
        # previous behaviour) spent an extra round trip of latency per request:
        # the effective gap became RTT + pace_seconds instead of pace_seconds.
        elapsed = self.monotonic() - self._last_request
        remaining = self.pace_seconds - elapsed
        if remaining > 0:
            self.sleep(remaining)
        self._last_request = self.monotonic()

    def get_bytes(
        self,
        url: str,
        *,
        headers: dict[str, str] | None = None,
        max_bytes: int,
        timeout: int = DEFAULT_TIMEOUT_SECONDS,
    ) -> bytes:
        safe_url = normalize_ncbi_url(url)
        request_headers = dict(headers or {})
        key = (safe_url, tuple(sorted(request_headers.items())), int(max_bytes))
        with self._lock:
            cached = self._cache.get(key)
            if cached is not None:
                self._cache.move_to_end(key)
                return cached

            last_error = "unknown transport error"
            for attempt in range(self.retries):
                self._pace()
                try:
                    payload = self.transport(safe_url, request_headers, timeout, max_bytes)
                    if self.cache_size and len(payload) <= MAX_CACHE_ENTRY_BYTES:
                        self._cache[key] = payload
                        self._cache.move_to_end(key)
                        while len(self._cache) > self.cache_size:
                            self._cache.popitem(last=False)
                    return payload
                except urllib.error.HTTPError as exc:
                    last_error = f"HTTP {exc.code}"
                    if exc.code not in {429, 500, 502, 503, 504}:
                        break
                except (urllib.error.URLError, TimeoutError, OSError, http.client.HTTPException) as exc:
                    last_error = type(exc).__name__
                if attempt < self.retries - 1:
                    self.sleep(min(2.0 * (attempt + 1), 6.0))
            raise DiscoveryUnavailableError(f"NCBI request unresolved after {self.retries} attempt(s): {last_error}")

    def get_json(self, url: str, *, max_bytes: int = MAX_JSON_BYTES) -> dict[str, Any]:
        payload = self.get_bytes(url, max_bytes=max_bytes)
        try:
            data = json.loads(payload.decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise DiscoveryUnavailableError("NCBI returned an invalid JSON response") from exc
        if not isinstance(data, dict):
            raise DiscoveryUnavailableError("NCBI returned an unexpected JSON response")
        if data.get("error"):
            raise DiscoveryUnavailableError("NCBI returned an E-utilities error response")
        return data

    def _eutils_url(self, endpoint: str, params: dict[str, Any]) -> str:
        if endpoint not in {"esearch.fcgi", "esummary.fcgi"}:
            raise DiscoveryError("unsupported NCBI E-utilities endpoint")
        merged = {**params, **self.request_config.eutils_params()}
        return f"{EUTILS_BASE}/{endpoint}?" + urllib.parse.urlencode(merged)

    def _gds_search(self, term: str, *, retstart: int, retmax: int) -> tuple[int, list[dict[str, Any]]]:
        esearch_url = self._eutils_url(
            "esearch.fcgi",
            {
                "db": "gds",
                "term": term,
                "retstart": retstart,
                "retmax": retmax,
                "retmode": "json",
                "sort": "relevance",
            },
        )
        search = self.get_json(esearch_url)
        result = search.get("esearchresult", {})
        ids = [str(value) for value in result.get("idlist", [])]
        try:
            total = int(result.get("count", 0))
        except (TypeError, ValueError):
            total = 0
        if not ids:
            return total, []
        summary_url = self._eutils_url(
            "esummary.fcgi",
            {"db": "gds", "id": ",".join(ids), "retmode": "json"},
        )
        summary = self.get_json(summary_url)
        summary_result = summary.get("result", {})
        records = [summary_result[uid] for uid in ids if isinstance(summary_result.get(uid), dict)]
        return total, records

    def search_summaries(
        self,
        query: str,
        species: SpeciesSpec,
        *,
        page: int,
        page_size: int,
    ) -> tuple[int, list[dict[str, Any]]]:
        """Return a stable page after excluding mixed-organism GEO Series.

        Entrez's organism filter also matches a multi-species Series that contains
        the requested organism.  Scan from the beginning in cached 100-record
        chunks so page boundaries are defined over exact, single-organism records
        and a rejected mixed study does not leave a short page or reappear later.
        The returned total remains NCBI's query count and is therefore a safe upper
        bound on the exact single-organism count.
        """

        total, records, _ = self.search_summaries_page(query, species, page=page, page_size=page_size)
        return total, records

    def search_summaries_page(
        self,
        query: str,
        species: SpeciesSpec,
        *,
        page: int,
        page_size: int,
    ) -> tuple[int, list[dict[str, Any]], bool]:
        """Return an exact-species page plus a verified next-page flag."""

        term = build_geo_query(query, species)
        exact_start = (page - 1) * page_size
        exact_end = exact_start + page_size
        exact_records: list[dict[str, Any]] = []
        raw_start = 0
        total = 0
        while len(exact_records) < exact_end + 1 and raw_start < MAX_EXACT_PAGE_SCAN_RECORDS:
            total, records = self._gds_search(term, retstart=raw_start, retmax=EXACT_PAGE_FETCH_SIZE)
            if not records:
                break
            exact_records.extend(
                record
                for record in records
                if re.fullmatch(r"GSE\d+", str(record.get("accession") or "").upper())
                and str(record.get("taxon") or "").strip() == species.scientific_name
            )
            raw_start += EXACT_PAGE_FETCH_SIZE
            if raw_start >= total:
                break
        return total, exact_records[exact_start:exact_end], len(exact_records) > exact_end

    def search_summaries_global(
        self,
        query: str,
        species: SpeciesSpec,
        *,
        limit: int,
    ) -> tuple[int, list[dict[str, Any]], bool]:
        """Collect the exact-species ranking universe without diagnostics."""

        total, records, truncated, _ = self.search_summaries_global_detailed(
            query,
            species,
            limit=limit,
        )
        return total, records, truncated

    def search_summaries_global_detailed(
        self,
        query: str,
        species: SpeciesSpec,
        *,
        limit: int,
    ) -> tuple[int, list[dict[str, Any]], bool, dict[str, Any]]:
        """Collect a complete, deduplicated exact-species ranking universe.

        One additional unique exact-species record is collected when available
        so truncation at ``limit`` is reported exactly without assessing the
        record beyond the cap. Diagnostics make Entrez pre-filter exclusions
        explicit without exposing excluded records as selectable studies.
        """

        if isinstance(limit, bool) or not isinstance(limit, int) or not 1 <= limit <= MAX_GLOBAL_RANK_LIMIT:
            raise DiscoveryError(f"global_limit must be a whole number from 1 to {MAX_GLOBAL_RANK_LIMIT}")
        term = build_geo_query(query, species)
        exact_records: list[dict[str, Any]] = []
        seen_accessions: set[str] = set()
        raw_start = 0
        total = 0
        raw_records_scanned = 0
        raw_records_fetched = 0
        excluded_reason_counts: dict[str, int] = {}
        excluded_records_sample: list[dict[str, str]] = []

        def record_exclusion(record: dict[str, Any], reason: str) -> None:
            excluded_reason_counts[reason] = excluded_reason_counts.get(reason, 0) + 1
            if len(excluded_records_sample) >= 25:
                return
            excluded_records_sample.append(
                {
                    "accession": str(record.get("accession") or "").strip().upper(),
                    "observed_taxon": str(record.get("taxon") or "").strip(),
                    "reason": reason,
                }
            )

        while len(exact_records) <= limit and raw_start < MAX_EXACT_PAGE_SCAN_RECORDS:
            total, records = self._gds_search(term, retstart=raw_start, retmax=EXACT_PAGE_FETCH_SIZE)
            if not records:
                if raw_start < total:
                    raise DiscoveryUnavailableError(
                        "NCBI returned an incomplete GEO summary page while building the global ranking"
                    )
                break
            raw_records_fetched += len(records)
            for record in records:
                raw_records_scanned += 1
                accession = str(record.get("accession") or "").upper()
                observed_taxon = str(record.get("taxon") or "").strip()
                if not re.fullmatch(r"GSE\d+", accession):
                    record_exclusion(record, "not_a_geo_series")
                    continue
                if observed_taxon != species.scientific_name:
                    record_exclusion(record, "mixed_or_mismatched_organism")
                    continue
                if accession in seen_accessions:
                    record_exclusion(record, "duplicate_exact_accession")
                    continue
                seen_accessions.add(accession)
                exact_records.append(record)
                if len(exact_records) > limit:
                    break
            raw_start += EXACT_PAGE_FETCH_SIZE
            if len(exact_records) > limit or raw_start >= total:
                break
        if len(exact_records) <= limit and raw_start < total:
            raise DiscoveryUnavailableError(
                f"exact-species GEO scan exceeded the {MAX_EXACT_PAGE_SCAN_RECORDS}-record safety ceiling"
            )
        exact_records_overflow = max(len(exact_records) - limit, 0)
        excluded_record_count = sum(excluded_reason_counts.values())
        diagnostics = {
            "scope": "NCBI summary records processed while building the exact single-organism ranking universe",
            "raw_records_fetched": raw_records_fetched,
            "raw_records_scanned": raw_records_scanned,
            "raw_records_processed": raw_records_scanned,
            "exact_records_retained": min(len(exact_records), limit),
            "exact_records_overflow": exact_records_overflow,
            "excluded_record_count": excluded_record_count,
            "records_accounted": min(len(exact_records), limit) + exact_records_overflow + excluded_record_count,
            "excluded_reason_counts": excluded_reason_counts,
            "excluded_records_sample": excluded_records_sample,
            "excluded_records_sample_limit": 25,
        }
        return total, exact_records[:limit], len(exact_records) > limit, diagnostics

    def get_or_build_global_search_snapshot(
        self,
        key: tuple[str, str, int, int],
        builder: Callable[[], dict[str, Any]],
    ) -> tuple[dict[str, Any], bool]:
        """Return a copied successful snapshot and whether it came from cache."""

        def cached_snapshot() -> dict[str, Any] | None:
            now = self.monotonic()
            with self._global_search_cache_lock:
                entry = self._global_search_cache.get(key)
                if entry is None:
                    return None
                created_at, snapshot = entry
                age = now - created_at
                if age < 0 or age > GLOBAL_SEARCH_CACHE_TTL_SECONDS:
                    self._global_search_cache.pop(key, None)
                    return None
                self._global_search_cache.move_to_end(key)
                return copy.deepcopy(snapshot)

        cached = cached_snapshot()
        if cached is not None:
            return cached, True
        with self._global_search_build_lock:
            cached = cached_snapshot()
            if cached is not None:
                return cached, True
            snapshot = builder()
            stored = copy.deepcopy(snapshot)
            with self._global_search_cache_lock:
                self._global_search_cache[key] = (self.monotonic(), stored)
                self._global_search_cache.move_to_end(key)
                while len(self._global_search_cache) > GLOBAL_SEARCH_CACHE_SIZE:
                    self._global_search_cache.popitem(last=False)
            return copy.deepcopy(stored), False

    def accession_summaries(
        self,
        accessions: list[str],
        species: SpeciesSpec,
    ) -> list[dict[str, Any]]:
        accession_term = " OR ".join(f'"{accession}"[ACCN]' for accession in accessions)
        term = f'({accession_term}) AND gse[ETYP] AND "{species.scientific_name}"[Organism]'
        _, records = self._gds_search(term, retstart=0, retmax=len(accessions))
        return records

    def publication_summaries(self, pmids: Iterable[str]) -> dict[str, dict[str, Any]]:
        ids = list(dict.fromkeys(str(value).strip() for value in pmids if str(value).strip()))
        if not ids:
            return {}
        publications: dict[str, dict[str, Any]] = {}
        for start in range(0, len(ids), PUBMED_SUMMARY_BATCH_SIZE):
            batch = ids[start : start + PUBMED_SUMMARY_BATCH_SIZE]
            summary_url = self._eutils_url(
                "esummary.fcgi",
                {"db": "pubmed", "id": ",".join(batch), "retmode": "json"},
            )
            summary = self.get_json(summary_url)
            result = summary.get("result", {})
            publications.update(
                {pmid: result[pmid] for pmid in batch if isinstance(result.get(pmid), dict)}
            )
        return publications

    def fetch_geo_soft(self, accession: str) -> str:
        if not re.fullmatch(r"GSE\d+", str(accession).upper()):
            raise DiscoveryError(f"invalid GEO Series accession: {accession}")
        url = GEO_ACCESSION_URL + "?" + urllib.parse.urlencode(
            {"acc": str(accession).upper(), "targ": "self", "form": "text", "view": "brief"}
        )
        payload = self.get_bytes(url, max_bytes=MAX_SOFT_BYTES)
        text = payload.decode("utf-8", "replace")
        if "^SERIES" not in text:
            raise DiscoveryUnavailableError(f"GEO returned no Series SOFT record for {accession}")
        return text

    def fetch_geo_sample_soft(self, accession: str) -> str:
        """Fetch the per-sample SOFT block so GSM ids can be shown with their labels."""

        if not re.fullmatch(r"GSE\d+", str(accession).upper()):
            raise DiscoveryError(f"invalid GEO Series accession: {accession}")
        url = GEO_ACCESSION_URL + "?" + urllib.parse.urlencode(
            {"acc": str(accession).upper(), "targ": "gsm", "form": "text", "view": "brief"}
        )
        payload = self.get_bytes(url, max_bytes=MAX_SOFT_BYTES)
        return payload.decode("utf-8", "replace")

    def fetch_candidate(self, url: str, *, full: bool) -> tuple[bytes, str]:
        lower = urllib.parse.urlsplit(url).path.lower()
        if full or lower.endswith(".xlsx"):
            return self.get_bytes(url, max_bytes=MAX_CANDIDATE_BYTES), "full"
        headers = {"Range": f"bytes=0-{MAX_TEXT_PREFIX_BYTES - 1}"}
        return self.get_bytes(url, headers=headers, max_bytes=MAX_TEXT_PREFIX_BYTES), "header_prefix"


DEFAULT_CLIENT = NcbiGeoClient()


HARD_REJECT_RE = re.compile(
    r"raw\.tar|\.mtx(?:\.gz)?$|\.h5(?:ad)?$|barcodes|features\.tsv|"
    r"\.bam(?:\.gz)?$|\.bed(?:\.gz)?$|\.bw$|\.bigwig$|fastq|readme|filelist",
    re.I,
)
STRONG_RE = re.compile(
    r"deseq|edger|limma|(?:^|[_\-.])deg(?:[_\-.]|$)|differential|diff[_\-.]?exp|"
    r"gene[_\-.]?exp\.diff|toptable|(?:^|[_\-.])dge(?:[_\-.]|$)|(?:^|[_\-.])dea(?:[_\-.]|$)",
    re.I,
)
UPSTREAM_COUNT_RE = re.compile(r"raw[_\-.]?counts?|count[_\-.]?matrix|counts?[_\-.]?(?:table|data)|_counts?\.", re.I)
UPSTREAM_EXPRESSION_RE = re.compile(
    r"series[_-]?matrix|normalized[_\-.]?(?:expression|intensity)|expression[_\-.]?matrix|rlog|vst|fpkm|rpkm|tpm",
    re.I,
)
WEAK_RE = re.compile(r"processed|compare|table[_\-.]?s?\d*|results?|supplement|analysis", re.I)
SUPPORTED_TEXT_RE = re.compile(r"\.(csv|tsv|txt)(\.gz)?$", re.I)
# Repositories serve supplementary tables as workbooks at least as often as text,
# and gzipped. The analysis path reads all four shapes, so refusing to inspect
# them here made DEGORA decline files it could have used.
SUPPORTED_WORKBOOK_RE = re.compile(r"\.(xlsx|xls)(\.gz)?$", re.I)


def classify_filename(name_or_url: str) -> dict[str, Any]:
    path = urllib.parse.urlsplit(str(name_or_url)).path
    name = urllib.parse.unquote(Path(path).name or str(name_or_url))
    lower = name.lower()
    supported = bool(SUPPORTED_TEXT_RE.search(lower) or SUPPORTED_WORKBOOK_RE.search(lower))
    if HARD_REJECT_RE.search(lower):
        return {
            "name": name,
            "tier": "reject",
            "role": "unsupported",
            "reason": "sequence, archive, sparse-matrix, or metadata file",
            "inspectable": False,
        }
    if UPSTREAM_COUNT_RE.search(lower):
        return {
            "name": name,
            "tier": "upstream",
            "role": "count_matrix",
            "reason": "filename indicates a public count matrix; contrast and sample groups are required",
            "inspectable": supported,
        }
    if UPSTREAM_EXPRESSION_RE.search(lower):
        return {
            "name": name,
            "tier": "upstream",
            "role": "normalized_expression_matrix",
            "reason": "filename indicates an expression matrix; scale, annotation, and sample groups are required",
            "inspectable": supported,
        }
    if STRONG_RE.search(lower):
        return {
            "name": name,
            "tier": "strong",
            "role": "deg_table",
            "reason": "filename contains a differential-expression analysis token",
            "inspectable": supported,
        }
    if supported and WEAK_RE.search(lower):
        return {
            "name": name,
            "tier": "weak",
            "role": "unknown_table",
            "reason": "generic processed or supplementary table name",
            "inspectable": True,
        }
    if supported:
        return {
            "name": name,
            "tier": "weak",
            "role": "unknown_table",
            "reason": "tabular file with an uninformative name",
            "inspectable": True,
        }
    return {
        "name": name,
        "tier": "unsupported",
        "role": "unsupported",
        "reason": "unsupported or non-tabular file type",
        "inspectable": False,
    }


GENE_EXACT_RE = re.compile(
    r"^(gene|gene[_. ]?id|gene[_. ]?name|gene[_. ]?symbol|symbol|hgnc.*|hugo.*|ensembl.*|geneid)$",
    re.I,
)
# row_name is DEGORA's own name for the identifier column it recovers from an R
# write.csv export, where the gene names arrive as an unnamed index. Leaving it
# out meant DEGORA could not recognise a column it had just created itself.
GENE_LOOSE_RE = re.compile(
    r"gene|symbol|hgnc|hugo|ensembl|entrez|probe|transcript|^row_name(_\d+)?$|(?:^|_)id$", re.I
)
LFC_HIGH_RE = re.compile(r"log2[\W_]*fold[\W_]*change|log2fc|log[_. ]?fc|log2ratio", re.I)
LFC_AMBIGUOUS_RE = re.compile(r"fold[_. ]?change|foldchange|(?:^|_)beta$|effect[_. ]?size", re.I)
PADJ_RE = re.compile(r"padj|adj[_. ]?p|adjusted[_. ]?p|fdr|q[_. ]?value|qvalue|qval", re.I)
P_RE = re.compile(r"p[_. ]?value|pvalue|pval|p_val|^p$", re.I)
NON_SAMPLE_RE = re.compile(
    r"^(?:gene[_. ]?)?(?:start|stop|end|length|strand|chrom(?:osome)?|chr)|"
    r"^(?:description|annotation|biotype|entrez|transcript|feature|base[_. ]?mean|mean|average|avg|"
    r"ave[_. ]?expr|age|batch|sex|gender|quality|qc|group|condition|phenotype|class)|"
    r"(?:log2?[_. ]?fc|fold[_. ]?change|p[_. ]?value|pval|padj|adj[_. ]?p[_. ]?val|fdr|q[_. ]?value|"
    r"stat(?:istic)?|score|stderr|std[_. ]?error|se|df|degrees[_. ]?of[_. ]?freedom)$|^(?:t|b)$",
    re.I,
)
DE_STAT_COLUMN_RE = re.compile(
    r"^(?:ave[_. ]?expr|t|b|stat(?:istic)?|score|log2?[_. ]?fc|fold[_. ]?change|"
    r"p[_. ]?value|pval|padj|adj[_. ]?p[_. ]?val|fdr|q[_. ]?value)$",
    re.I,
)
SAMPLE_NAME_RE = re.compile(
    r"^(?:GSM|SRR|ERR|DRR)\d+$|"
    r"(?:^|[_ .-])(?:ctrl|control|case|treat(?:ed|ment)?|vehicle|untreated|baseline|"
    r"wt|ko|knockout|hypox\w*|normox\w*|hox|nox|rep(?:licate)?|sample|donor|patient|subject|culture)"
    r"(?:[_ .-]?\w+)*$",
    re.I,
)


def _looks_like_sample_column(name: str) -> bool:
    value = str(name).strip()
    if not value or NON_SAMPLE_RE.search(value):
        return False
    return bool(SAMPLE_NAME_RE.search(value) or re.search(r"\d", value))


def _gene_column_priority(name: str) -> tuple[int, str]:
    value = str(name).strip().lower()
    if re.search(r"symbol|hgnc|hugo|gene[_. ]?name", value):
        return (0, value)
    if "ensembl" in value or re.search(r"gene[_. ]?id|geneid", value):
        return (1, value)
    if value == "gene":
        return (2, value)
    if value == "id_ref":
        return (3, value)
    return (4, value)


def classify_header(columns: Iterable[Any]) -> dict[str, Any]:
    names = [str(value).strip() for value in columns if str(value).strip() and str(value).strip().lower() != "nan"]
    genes = [name for name in names if GENE_EXACT_RE.search(name)]
    loose_genes = [name for name in names if GENE_LOOSE_RE.search(name)]
    lfc_high = [name for name in names if LFC_HIGH_RE.search(name)]
    lfc_ambiguous = [name for name in names if LFC_AMBIGUOUS_RE.search(name) and name not in lfc_high]
    lfc_candidates = [*lfc_high, *lfc_ambiguous]
    padj = [name for name in names if PADJ_RE.search(name)]
    pvalue = [name for name in names if P_RE.search(name) and name not in padj]
    gene = sorted(genes, key=_gene_column_priority)[0] if genes else (
        sorted(loose_genes, key=_gene_column_priority)[0] if loose_genes else ""
    )
    lfc = lfc_high[0] if lfc_high else (lfc_ambiguous[0] if lfc_ambiguous else "")
    status = "not_deg_table"
    reason = "missing gene, log2 fold-change, or nominal p-value column"
    if gene and lfc and pvalue:
        if len(lfc_candidates) > 1 or len(pvalue) > 1:
            status = "requires_column_mapping"
            reason = "multiple effect or nominal p-value columns require an explicit contrast mapping"
        elif lfc_high:
            status = "candidate_header"
            reason = "gene, log2 fold-change, and nominal p-value columns detected"
        else:
            status = "requires_lfc_confirmation"
            reason = "effect column detected but log2 scale is not explicit"
    elif gene and lfc and padj:
        status = "requires_pvalue_mapping"
        reason = "adjusted significance detected but nominal p-value is missing"
    return {
        "status": status,
        "reason": reason,
        "columns": names,
        "gene_columns": genes or loose_genes,
        "lfc_columns": lfc_candidates,
        "p_columns": pvalue,
        "padj_columns": padj,
        "mapping": {
            "gene_column": gene,
            "lfc_column": lfc,
            "p_column": pvalue[0] if pvalue else "",
            "padj_column": padj[0] if padj else "",
        },
        "lfc_scale_explicit": bool(lfc_high),
    }


def _numeric_fraction(values: Iterable[Any], *, unit_interval: bool = False) -> float:
    seen = 0
    valid = 0
    for value in values:
        text = str(value).strip()
        if not text or text.lower() in {"na", "nan", "null", "none"}:
            continue
        seen += 1
        try:
            number = float(text)
        except ValueError:
            continue
        if math.isfinite(number) and (not unit_interval or 0.0 <= number <= 1.0):
            valid += 1
    return valid / seen if seen else 0.0


def _inspect_rows(rows: list[list[Any]], *, sheet: str = "") -> dict[str, Any]:
    best: dict[str, Any] | None = None
    best_score = -1
    for index, row in enumerate(rows[:10]):
        header = classify_header(row)
        mapping = header["mapping"]
        score = sum(bool(mapping[key]) for key in ("gene_column", "lfc_column", "p_column", "padj_column"))
        if header["lfc_scale_explicit"]:
            score += 2
        if score <= best_score:
            continue
        following = rows[index + 1 : index + 21]
        name_to_index = {str(value).strip(): position for position, value in enumerate(row)}

        def column_values(name: str) -> list[Any]:
            position = name_to_index.get(name)
            if position is None:
                return []
            return [values[position] for values in following if position < len(values)]

        lfc_fraction = _numeric_fraction(column_values(mapping["lfc_column"])) if mapping["lfc_column"] else 0.0
        p_fraction = _numeric_fraction(column_values(mapping["p_column"]), unit_interval=True) if mapping["p_column"] else 0.0
        gene_fraction = (
            sum(bool(str(value).strip()) for value in column_values(mapping["gene_column"])) / max(len(following), 1)
            if mapping["gene_column"]
            else 0.0
        )
        header.update(
            {
                "header_row": index + 1,
                "sheet_name": sheet,
                "sample_validation": {
                    "gene_nonempty_fraction": round(gene_fraction, 3),
                    "lfc_numeric_fraction": round(lfc_fraction, 3),
                    "p_unit_interval_fraction": round(p_fraction, 3),
                    "sample_rows": len(following),
                },
            }
        )
        if (
            header["status"] == "candidate_header"
            and len(following) >= 2
            and gene_fraction >= 0.5
            and lfc_fraction >= 0.5
            and p_fraction >= 0.5
        ):
            header["status"] = "ready_for_review"
            header["reason"] = "header and sampled values match the DEGORA gene/log2FC/p-value contract"
            score += 3
        best = header
        best_score = score
    return best or classify_header([])


def _decode_text_payload(payload: bytes, *, compressed: bool) -> str:
    if not compressed:
        return payload[:MAX_DECOMPRESSED_BYTES].decode("utf-8-sig", "replace")
    decoder = zlib.decompressobj(31)
    try:
        decoded = decoder.decompress(payload, MAX_DECOMPRESSED_BYTES + 1)
    except zlib.error as exc:
        raise DiscoveryError(f"gzip header inspection failed: {exc}") from exc
    # Header inspection needs only a bounded prefix.  A valid full DEG table can
    # readily exceed 5 MB after decompression; truncating here is safe because the
    # full gzip stream is independently size-validated before materialization.
    return decoded[:MAX_DECOMPRESSED_BYTES].decode("utf-8-sig", "replace")


def _validate_full_gzip(payload: bytes) -> None:
    total = 0
    try:
        with gzip.GzipFile(fileobj=io.BytesIO(payload)) as handle:
            while chunk := handle.read(1024 * 1024):
                total += len(chunk)
                if total > MAX_GZIP_UNCOMPRESSED_BYTES:
                    raise DiscoveryError("gzip candidate exceeds the full uncompressed safety cap")
    except DiscoveryError:
        raise
    except (EOFError, OSError, zlib.error) as exc:
        raise DiscoveryError(f"gzip candidate is not a valid complete stream: {exc}") from exc


def _delimited_rows(text: str) -> list[list[str]]:
    lines = text.splitlines()[:80]
    candidates: list[tuple[int, list[list[str]]]] = []
    for separator in ("\t", ",", ";"):
        try:
            parsed = list(csv.reader(lines, delimiter=separator))
        except csv.Error:
            # A single field past csv's 128 KiB limit is not a DEG table, and this
            # runs on files fetched from public repositories - so it has to read as
            # "no usable table here", not as a traceback out of a whole preparation.
            continue
        width = max((len(row) for row in parsed[:10]), default=0)
        candidates.append((width, parsed))
    return max(candidates, key=lambda item: item[0])[1] if candidates else []


def _validate_xlsx_archive(payload: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            infos = archive.infolist()
            if len(infos) > MAX_XLSX_MEMBERS:
                raise DiscoveryError("XLSX candidate contains too many archive members")
            total = 0
            for info in infos:
                if info.file_size > MAX_XLSX_MEMBER_BYTES:
                    raise DiscoveryError("XLSX candidate contains an oversized archive member")
                total += info.file_size
                if total > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise DiscoveryError("XLSX candidate exceeds the uncompressed workbook safety cap")
    except zipfile.BadZipFile as exc:
        raise DiscoveryError("XLSX candidate is not a valid OOXML archive") from exc


def _decompressed_workbook(payload: bytes, lower_name: str) -> bytes:
    """Expand a gzipped workbook, bounded, so it can be inspected in memory."""

    if not lower_name.endswith(".gz"):
        return payload
    try:
        with gzip.GzipFile(fileobj=io.BytesIO(payload)) as handle:
            expanded = handle.read(MAX_GZIP_UNCOMPRESSED_BYTES + 1)
    except (EOFError, OSError, zlib.error) as exc:
        raise DiscoveryError(f"gzipped workbook could not be expanded: {exc}") from exc
    if len(expanded) > MAX_GZIP_UNCOMPRESSED_BYTES:
        raise DiscoveryError("gzipped workbook exceeds the uncompressed safety cap")
    return expanded


def _inspect_legacy_workbook(payload: bytes) -> dict[str, Any]:
    """Classify a legacy .xls workbook's best sheet, the way .xlsx is classified."""

    if not payload.startswith(b"\xd0\xcf\x11\xe0"):
        raise DiscoveryError("legacy workbook does not carry the OLE2 signature a .xls file has")
    try:
        import pandas as pd

        sheets = pd.read_excel(io.BytesIO(payload), sheet_name=None, header=None, nrows=30)
    except ImportError as exc:  # pragma: no cover - xlrd ships as a runtime dependency
        raise DiscoveryError(f"legacy workbook reader is unavailable: {exc}") from exc
    except Exception as exc:  # noqa: BLE001 - an unreadable candidate is not a crash
        raise DiscoveryError(f"legacy workbook could not be read: {exc}") from exc
    ranks = {
        "ready_for_review": 5,
        "candidate_header": 4,
        "requires_lfc_confirmation": 3,
        "requires_pvalue_mapping": 2,
        "not_deg_table": 1,
    }
    best: dict[str, Any] | None = None
    best_rank = -1
    for sheet_name, frame in list(sheets.items())[:8]:
        rows = [list(row) for row in frame.itertuples(index=False, name=None)]
        current = _inspect_rows(rows, sheet=str(sheet_name))
        current_rank = ranks.get(current["status"], 0)
        if current_rank > best_rank:
            best, best_rank = current, current_rank
    return best or classify_header([])


def inspect_candidate_bytes(name_or_url: str, payload: bytes) -> dict[str, Any]:
    lower = urllib.parse.urlsplit(str(name_or_url)).path.lower()
    if SUPPORTED_TEXT_RE.search(lower):
        text = _decode_text_payload(payload, compressed=lower.endswith(".gz"))
        result = _inspect_rows(_delimited_rows(text))
    elif SUPPORTED_WORKBOOK_RE.search(lower):
        payload = _decompressed_workbook(payload, lower)
        if lower.replace(".gz", "").endswith(".xls"):
            result = _inspect_legacy_workbook(payload)
            return {**result, "bytes": len(payload)}
        _validate_xlsx_archive(payload)
        from openpyxl import load_workbook

        workbook = None
        try:
            workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
            best: dict[str, Any] | None = None
            rank = -1
            for sheet_name in workbook.sheetnames[:8]:
                sheet = workbook[sheet_name]
                rows = [list(row) for row in sheet.iter_rows(max_row=30, max_col=60, values_only=True)]
                current = _inspect_rows(rows, sheet=sheet_name)
                current_rank = {
                    "ready_for_review": 5,
                    "candidate_header": 4,
                    "requires_lfc_confirmation": 3,
                    "requires_pvalue_mapping": 2,
                    "not_deg_table": 1,
                }.get(current["status"], 0)
                if current_rank > rank:
                    best = current
                    rank = current_rank
            result = best or classify_header([])
        except DiscoveryError:
            raise
        except Exception as exc:  # openpyxl raises several format-specific exception types
            raise DiscoveryError(f"XLSX header inspection failed: {type(exc).__name__}: {exc}") from exc
        finally:
            if workbook is not None:
                workbook.close()
    else:
        raise DiscoveryError("candidate type is not supported for automatic header inspection")
    result["payload_sha256"] = hashlib.sha256(payload).hexdigest()
    result["payload_bytes"] = len(payload)
    return result


def _matrix_rows_from_text(text: str) -> list[list[str]]:
    lines = text.splitlines()
    for index, line in enumerate(lines):
        if line.strip().lower() == "!series_matrix_table_begin":
            lines = lines[index + 1 :]
            break
    data_lines = [line for line in lines if line.strip() and not line.startswith("!")][:40]
    candidates: list[tuple[int, list[list[str]]]] = []
    for separator in ("\t", ",", ";"):
        parsed = list(csv.reader(data_lines, delimiter=separator))
        width = max((len(row) for row in parsed[:10]), default=0)
        candidates.append((width, parsed))
    return max(candidates, key=lambda item: item[0])[1] if candidates else []


def _inspect_upstream_rows(rows: list[list[Any]], *, declared_role: str, sheet: str = "") -> dict[str, Any]:
    best: dict[str, Any] | None = None
    best_score = -1
    for index, row in enumerate(rows[:12]):
        columns = [str(value).strip().strip('"') for value in row]
        if len(columns) < 3:
            continue
        non_numeric_labels = 0
        for name in columns:
            try:
                float(name)
            except ValueError:
                non_numeric_labels += 1
        # Numeric data rows must never outrank the real header merely because
        # each value can also be parsed as a numeric "column name".
        if non_numeric_labels < 2:
            continue
        following = rows[index + 1 : index + 21]
        numeric_columns: list[str] = []
        numeric_fractions: dict[str, float] = {}
        for position, name in enumerate(columns):
            values = [values[position] for values in following if position < len(values)]
            fraction = _numeric_fraction(values)
            numeric_fractions[name] = round(fraction, 3)
            if fraction >= 0.7:
                numeric_columns.append(name)
        gene_candidates = [name for name in columns if GENE_EXACT_RE.search(name) or name.upper() == "ID_REF"]
        if not gene_candidates:
            continue
        gene_column = sorted(gene_candidates, key=_gene_column_priority)[0]
        statistic_columns = [name for name in numeric_columns if DE_STAT_COLUMN_RE.search(name)]
        sample_columns = [name for name in numeric_columns if name != gene_column and _looks_like_sample_column(name)]
        blocked_unknown_statistics = declared_role == "unknown_matrix" and bool(statistic_columns)
        if blocked_unknown_statistics:
            sample_columns = []
        score = len(sample_columns) + (3 if gene_candidates else 0)
        if score <= best_score:
            continue
        status = "not_upstream_matrix"
        reason = "a gene/probe identifier and at least four numeric sample columns are required"
        if blocked_unknown_statistics:
            reason = "unknown table contains differential-expression statistic columns; sample identity is not confirmed"
        elif gene_column and len(sample_columns) >= 4 and len(following) >= 2:
            status = "upstream_matrix_ready_for_contrast"
            reason = "matrix columns were detected; choose at least two control and two treatment samples"
        best = {
            "status": status,
            "reason": reason,
            "declared_role": declared_role,
            "header_row": index + 1,
            "sheet_name": sheet,
            "gene_column": gene_column,
            "sample_columns": sample_columns,
            "statistic_columns": statistic_columns,
            "numeric_fractions": numeric_fractions,
            "sample_rows": len(following),
            "requires": ["control_samples", "treatment_samples", "contrast_direction"],
        }
        best_score = score
    return best or {
        "status": "not_upstream_matrix",
        "reason": "no tabular matrix header was detected",
        "declared_role": declared_role,
        "gene_column": "",
        "sample_columns": [],
        "requires": ["control_samples", "treatment_samples", "contrast_direction"],
    }


def inspect_upstream_bytes(name_or_url: str, payload: bytes, *, declared_role: str) -> dict[str, Any]:
    lower = urllib.parse.urlsplit(str(name_or_url)).path.lower()
    if SUPPORTED_TEXT_RE.search(lower):
        text = _decode_text_payload(payload, compressed=lower.endswith(".gz"))
        result = _inspect_upstream_rows(_matrix_rows_from_text(text), declared_role=declared_role)
    elif lower.endswith(".xlsx"):
        _validate_xlsx_archive(payload)
        from openpyxl import load_workbook

        workbook = None
        try:
            workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
            best: dict[str, Any] | None = None
            best_rank = -1
            for sheet_name in workbook.sheetnames[:8]:
                rows = [list(row) for row in workbook[sheet_name].iter_rows(max_row=32, max_col=120, values_only=True)]
                current = _inspect_upstream_rows(rows, declared_role=declared_role, sheet=sheet_name)
                current_rank = len(current.get("sample_columns", [])) + (
                    100 if current.get("status") == "upstream_matrix_ready_for_contrast" else 0
                )
                if current_rank > best_rank:
                    best = current
                    best_rank = current_rank
            result = best or _inspect_upstream_rows([], declared_role=declared_role)
        except DiscoveryError:
            raise
        except Exception as exc:
            raise DiscoveryError(f"XLSX matrix inspection failed: {type(exc).__name__}: {exc}") from exc
        finally:
            if workbook is not None:
                workbook.close()
    else:
        raise DiscoveryError("upstream candidate type is not supported for automatic inspection")
    result["payload_sha256"] = hashlib.sha256(payload).hexdigest()
    result["payload_bytes"] = len(payload)
    return result


def parse_geo_soft(text: str) -> dict[str, Any]:
    supplementary: list[str] = []
    taxa: set[str] = set()
    metadata: dict[str, Any] = {"title": "", "summary": "", "design": "", "pubmed_ids": []}
    for line in str(text).splitlines():
        if "=" not in line:
            continue
        key, value = (part.strip() for part in line.split("=", 1))
        if key == "!Series_supplementary_file":
            supplementary.append(value)
        elif key == "!Series_title":
            metadata["title"] = value
        elif key == "!Series_summary":
            metadata["summary"] = (metadata["summary"] + " " + value).strip()
        elif key == "!Series_overall_design":
            metadata["design"] = (metadata["design"] + " " + value).strip()
        elif key == "!Series_pubmed_id":
            metadata["pubmed_ids"].append(value)
        elif key.endswith("_organism_ch1") and value:
            taxa.add(value)
    metadata["supplementary_files"] = list(dict.fromkeys(supplementary))
    metadata["pubmed_ids"] = list(dict.fromkeys(metadata["pubmed_ids"]))
    metadata["taxa"] = sorted(taxa)
    return metadata


def _normalized_label_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).lower())


def match_sample_labels(columns: Iterable[str], labels: dict[str, Any]) -> dict[str, Any]:
    """Attach each matrix column to the GEO sample it came from.

    A series matrix is keyed by GSM, but an author-supplied matrix uses the
    submitter's own column names - `4641CERM6M24M_S2` and the like - so keying
    on the accession alone leaves every column unlabelled, which is where this
    started. Titles and source names are matched too, and anything ambiguous is
    left unmatched rather than guessed: a wrong label here flips a contrast.
    """

    index: dict[str, list[str]] = {}
    for accession, entry in labels.items():
        for value in (accession, entry.get("title", ""), entry.get("source", "")):
            key = _normalized_label_key(value)
            if len(key) < 3:
                continue
            index.setdefault(key, [])
            if accession not in index[key]:
                index[key].append(accession)

    unique = {key: found[0] for key, found in index.items() if len(found) == 1}
    resolved: dict[str, Any] = {}
    for column in columns:
        text = str(column)
        accession = ""
        if text.upper() in labels:
            accession = text.upper()
        else:
            key = _normalized_label_key(text)
            if not key:
                continue
            accession = unique.get(key, "")
            if not accession and len(key) >= 4:
                prefixes = [
                    value
                    for candidate, value in unique.items()
                    if len(candidate) >= 4 and (key.startswith(candidate) or candidate.startswith(key))
                ]
                if len(set(prefixes)) == 1:
                    accession = prefixes[0]
        if accession:
            resolved[text] = {"accession": accession, **labels[accession]}
    return resolved


def _sample_labels_for(client: Any, accession: str) -> dict[str, Any]:
    """Best-effort sample labels. Never let a label lookup fail a preparation."""

    fetch = getattr(client, "fetch_geo_sample_soft", None)
    if not callable(fetch):
        return {}
    try:
        return parse_geo_samples(fetch(accession))
    except (DiscoveryError, DiscoveryUnavailableError, OSError):
        return {}


def parse_geo_samples(text: str) -> dict[str, dict[str, Any]]:
    """Map each GSM in a Series sample SOFT record to its title and characteristics.

    Assigning control and treatment from bare GSM accessions means leaving the
    tool to look every one of them up in GEO, which is exactly where a group
    gets mis-assigned. The labels the submitter wrote are what make the choice
    checkable.
    """

    samples: dict[str, dict[str, Any]] = {}
    current = ""
    for line in str(text).splitlines():
        if "=" not in line:
            continue
        key, value = (part.strip() for part in line.split("=", 1))
        if key == "^SAMPLE":
            current = value.upper()
            samples.setdefault(current, {"title": "", "characteristics": [], "source": ""})
        elif not current:
            continue
        elif key == "!Sample_title":
            samples[current]["title"] = value
        elif key == "!Sample_source_name_ch1":
            samples[current]["source"] = value
        elif key.startswith("!Sample_characteristics_ch") and value:
            samples[current]["characteristics"].append(value)
    return samples


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _record_pmids(record: dict[str, Any]) -> list[str]:
    raw = record.get("pubmedids", [])
    if isinstance(raw, (str, int)):
        raw = [raw]
    values = [str(item).strip() for item in raw]
    return list(dict.fromkeys(value for value in values if re.fullmatch(r"\d+", value)))


def _merged_pmids(*collections: Iterable[Any]) -> list[str]:
    merged: list[str] = []
    for collection in collections:
        merged.extend(str(value).strip() for value in collection)
    return list(dict.fromkeys(value for value in merged if re.fullmatch(r"\d+", value)))


def _pmid_order(value: str) -> tuple[int, int | str]:
    return (0, int(value)) if value.isdigit() else (1, value)


def _assign_source_unit_components(studies: list[dict[str, Any]]) -> None:
    """Conservatively collapse studies connected by any linked PubMed ID."""

    parent = list(range(len(studies)))

    def find(index: int) -> int:
        while parent[index] != index:
            parent[index] = parent[parent[index]]
            index = parent[index]
        return index

    def union(left: int, right: int) -> None:
        left_root = find(left)
        right_root = find(right)
        if left_root != right_root:
            parent[right_root] = left_root

    owner: dict[str, int] = {}
    for index, study in enumerate(studies):
        pmids = _merged_pmids(study.get("pubmed_ids", []))
        study["pubmed_ids"] = pmids
        for pmid in pmids:
            previous = owner.setdefault(pmid, index)
            union(index, previous)

    component_pmids: dict[int, set[str]] = {}
    for index, study in enumerate(studies):
        component_pmids.setdefault(find(index), set()).update(study.get("pubmed_ids", []))

    for index, study in enumerate(studies):
        linked_pmids = sorted(component_pmids.get(find(index), set()), key=_pmid_order)
        accession = str(study.get("accession") or "").strip().upper()
        if linked_pmids:
            study["source_unit_id"] = f"PMID:{linked_pmids[0]}"
            study["source_unit_basis"] = (
                "shared linked PubMed component" if len(linked_pmids) > 1 else "linked PubMed paper"
            )
            study["source_unit_pubmed_ids"] = linked_pmids
        else:
            study["source_unit_id"] = accession
            study["source_unit_basis"] = "GEO accession without PubMed linkage"
            study["source_unit_pubmed_ids"] = []


def _publication_year(publication: dict[str, Any], record: dict[str, Any]) -> int | None:
    for value in (
        publication.get("sortpubdate"),
        publication.get("pubdate"),
        record.get("pdat"),
    ):
        match = re.search(r"(?:19|20)\d{2}", str(value or ""))
        if match:
            return int(match.group(0))
    return None


def _publication_authors(publication: dict[str, Any]) -> list[str]:
    raw = publication.get("authors", [])
    if not isinstance(raw, list):
        return []
    authors: list[str] = []
    for item in raw:
        if isinstance(item, dict):
            value = item.get("name", "")
        else:
            value = item
        text = str(value).strip()
        if text:
            authors.append(text)
    return authors


def _study_search_record(
    record: dict[str, Any],
    species: SpeciesSpec,
    publications: dict[str, dict[str, Any]],
) -> dict[str, Any] | None:
    accession = str(record.get("accession") or "").upper()
    if not re.fullmatch(r"GSE\d+", accession):
        return None
    if str(record.get("taxon") or "").strip() != species.scientific_name:
        return None
    pmids = _record_pmids(record)
    publication = publications.get(pmids[0], {}) if pmids else {}
    authors = _publication_authors(publication)
    paper_title = str(publication.get("title") or "").strip()
    dataset_title = str(record.get("title") or "").strip()
    journal = str(publication.get("fulljournalname") or publication.get("source") or "").strip()
    year = _publication_year(publication, record)
    return {
        "species": species.key,
        "scientific_name": species.scientific_name,
        "accession": accession,
        "paper_title": paper_title or dataset_title,
        "dataset_title": dataset_title,
        "authors": authors,
        "authors_display": ", ".join(authors) if authors else "Not linked in PubMed",
        "journal": journal or "Not linked in PubMed",
        "year": year,
        "publication_date": str(publication.get("pubdate") or record.get("pdat") or ""),
        "pubmed_ids": pmids,
        "pubmed_url": f"https://pubmed.ncbi.nlm.nih.gov/{pmids[0]}/" if pmids else "",
        "summary": str(record.get("summary") or ""),
        "study_type": str(record.get("gdstype") or ""),
        "release_date": str(record.get("pdat") or ""),
        "n_samples": _safe_int(record.get("n_samples")),
        "source_unit_id": f"PMID:{pmids[0]}" if pmids else accession,
        "source_unit_basis": "linked PubMed paper" if pmids else "GEO accession without PubMed linkage",
        "source_url": f"https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc={accession}",
    }


def _search_deg_input_assessment(accession: str, client: NcbiGeoClient | Any) -> dict[str, Any]:
    """Estimate DEG-input availability from GEO SOFT filenames without downloading payloads."""

    try:
        soft = parse_geo_soft(client.fetch_geo_soft(accession))
    except (DiscoveryError, DiscoveryUnavailableError, OSError, TimeoutError, urllib.error.URLError) as exc:
        tier = "unresolved"
        return {
            "tier": tier,
            "label": DEG_INPUT_LABELS[tier],
            "priority": DEG_INPUT_PRIORITIES[tier],
            "basis": SEARCH_ASSESSMENT_BASIS,
            "candidate_files": [],
            "counts": {"deg_like": 0, "tabular": 0, "matrix": 0},
            "error": f"SOFT request failed ({type(exc).__name__})",
        }

    classified = [classify_filename(url) for url in soft.get("supplementary_files", [])]
    deg_like = [item for item in classified if item.get("tier") == "strong" and item.get("inspectable")]
    tabular = [item for item in classified if item.get("tier") == "weak" and item.get("inspectable")]
    matrices = [item for item in classified if item.get("tier") == "upstream" and item.get("inspectable")]
    if deg_like:
        tier = "author_deg_likely"
    elif tabular:
        tier = "tabular_candidate"
    elif matrices:
        tier = "matrix_fallback"
    else:
        tier = "not_detected"
    relevant = [*deg_like, *tabular, *matrices]
    return {
        "tier": tier,
        "label": DEG_INPUT_LABELS[tier],
        "priority": DEG_INPUT_PRIORITIES[tier],
        "basis": SEARCH_ASSESSMENT_BASIS,
        "candidate_files": list(dict.fromkeys(str(item.get("name") or "") for item in relevant if item.get("name"))),
        "counts": {
            "deg_like": len(deg_like),
            "tabular": len(tabular),
            "matrix": len(matrices),
        },
        "error": "",
    }


GLOBAL_SEARCH_SORT_ALIASES = {
    "deg_input_priority": "deg_input_priority",
    "relevance": "ncbi_relevance_rank",
    "ncbi_relevance_rank": "ncbi_relevance_rank",
    "accession": "accession",
    "paper_title": "paper_title",
    "authors_display": "authors_display",
    "journal": "journal",
    "year": "year",
    "n_samples": "n_samples",
}


def _normalize_global_search_sort(sort_by: str | None, sort_order: str | None) -> tuple[str, str, str]:
    public_field = str(sort_by or "deg_input_priority").strip().lower()
    canonical_field = GLOBAL_SEARCH_SORT_ALIASES.get(public_field)
    if canonical_field is None:
        allowed = ", ".join(sorted(GLOBAL_SEARCH_SORT_ALIASES))
        raise DiscoveryError(f"sort_by must be one of: {allowed}")
    public_field = "relevance" if canonical_field == "ncbi_relevance_rank" else canonical_field
    order = str(sort_order or ("asc" if canonical_field == "ncbi_relevance_rank" else "desc")).strip().lower()
    if order not in {"asc", "desc"}:
        raise DiscoveryError("sort_order must be exactly asc or desc")
    return public_field, canonical_field, order


def _global_study_sort_value(study: dict[str, Any], field: str) -> Any:
    if field == "deg_input_priority":
        return _safe_int((study.get("deg_input_assessment") or {}).get("priority"), 0)
    if field in {"ncbi_relevance_rank", "year", "n_samples"}:
        value = study.get(field)
        return None if value is None else _safe_int(value)
    value = str(study.get(field) or "").strip()
    return value.casefold() if value else None


def _sort_global_studies(
    studies: list[dict[str, Any]],
    *,
    field: str,
    order: str,
) -> list[dict[str, Any]]:
    relevance_order = sorted(
        studies,
        key=lambda study: (
            _safe_int(study.get("ncbi_relevance_rank"), MAX_GLOBAL_RANK_LIMIT + 1),
            str(study.get("accession") or ""),
        ),
    )
    if field == "ncbi_relevance_rank":
        return relevance_order if order == "asc" else list(reversed(relevance_order))
    present = [study for study in relevance_order if _global_study_sort_value(study, field) is not None]
    missing = [study for study in relevance_order if _global_study_sort_value(study, field) is None]
    present.sort(key=lambda study: _global_study_sort_value(study, field), reverse=order == "desc")
    return [*present, *missing]


def _build_global_search_snapshot(
    clean_query: str,
    spec: SpeciesSpec,
    geo_client: NcbiGeoClient | Any,
    *,
    global_limit: int,
) -> dict[str, Any]:
    detailed_search = getattr(geo_client, "search_summaries_global_detailed", None)
    global_search = getattr(geo_client, "search_summaries_global", None)
    if not callable(global_search):
        raise DiscoveryError("global ranking requires a GEO client with search_summaries_global support")
    started = time.monotonic()
    if callable(detailed_search):
        total_hits, records, ranking_truncated, filter_diagnostics = detailed_search(
            clean_query,
            spec,
            limit=global_limit,
        )
    else:
        total_hits, records, ranking_truncated = global_search(clean_query, spec, limit=global_limit)
        filter_diagnostics = {
            "scope": "exact single-organism ranking universe",
            "raw_records_fetched": None,
            "raw_records_scanned": None,
            "raw_records_processed": None,
            "exact_records_retained": len(records),
            "exact_records_overflow": None,
            "excluded_record_count": None,
            "records_accounted": None,
            "excluded_reason_counts": {},
            "excluded_records_sample": [],
            "excluded_records_sample_limit": 0,
        }
    pmids = [pmid for record in records for pmid in _record_pmids(record)]
    publications = geo_client.publication_summaries(pmids)
    studies: list[dict[str, Any]] = []
    for relevance_rank, record in enumerate(records, start=1):
        study = _study_search_record(record, spec, publications)
        if study is None:
            continue
        study["ncbi_relevance_rank"] = relevance_rank
        study["deg_input_assessment"] = _search_deg_input_assessment(study["accession"], geo_client)
        studies.append(study)
    return {
        "total_hits": total_hits,
        "ranking_truncated": bool(ranking_truncated),
        "studies": studies,
        "filter_diagnostics": filter_diagnostics,
        "ranking_build_seconds": round(max(time.monotonic() - started, 0.0), 3),
    }


def search_geo(
    query: str,
    species: str,
    *,
    page: int = 1,
    page_size: int = DEFAULT_PAGE_SIZE,
    assess_files: bool = False,
    global_rank: bool = False,
    global_limit: int = DEFAULT_GLOBAL_RANK_LIMIT,
    sort_by: str | None = None,
    sort_order: str | None = None,
    client: NcbiGeoClient | Any | None = None,
) -> dict[str, Any]:
    """Return one species-scoped, publication-enriched GEO result page.

    With ``global_rank=True``, filename-only assessment ranks up to 1,000 exact
    studies before pagination. Candidate payloads remain untouched until
    explicit preparation.
    """

    spec = normalize_species(species)
    clean_query = " ".join(_query_terms(query))
    version_info = runtime_version_info()
    if isinstance(page, bool) or not isinstance(page, int) or not 1 <= page <= MAX_PAGE:
        raise DiscoveryError(f"page must be a whole number from 1 to {MAX_PAGE}")
    if isinstance(page_size, bool) or not isinstance(page_size, int) or not 1 <= page_size <= MAX_LIMIT:
        raise DiscoveryError(f"page_size must be a whole number from 1 to {MAX_LIMIT}")
    geo_client = client or DEFAULT_CLIENT
    if global_rank:
        if not assess_files:
            raise DiscoveryError("global_rank requires assess_files=True")
        if (
            isinstance(global_limit, bool)
            or not isinstance(global_limit, int)
            or not 1 <= global_limit <= MAX_GLOBAL_RANK_LIMIT
        ):
            raise DiscoveryError(f"global_limit must be a whole number from 1 to {MAX_GLOBAL_RANK_LIMIT}")
        public_sort, canonical_sort, normalized_order = _normalize_global_search_sort(sort_by, sort_order)
        cache_key = (clean_query.casefold(), spec.key, global_limit, SEARCH_ASSESSMENT_VERSION)
        builder = lambda: _build_global_search_snapshot(
            clean_query,
            spec,
            geo_client,
            global_limit=global_limit,
        )
        snapshot_cache = getattr(geo_client, "get_or_build_global_search_snapshot", None)
        if callable(snapshot_cache):
            snapshot, cache_hit = snapshot_cache(cache_key, builder)
        else:
            snapshot, cache_hit = builder(), False
        ranked_studies = _sort_global_studies(
            snapshot["studies"],
            field=canonical_sort,
            order=normalized_order,
        )
        evaluated_studies = len(ranked_studies)
        total_pages = math.ceil(evaluated_studies / page_size) if evaluated_studies else 0
        page_start = (page - 1) * page_size
        studies = ranked_studies[page_start : page_start + page_size]
        return {
            **version_info,
            "query": clean_query,
            "ncbi_query": build_geo_query(clean_query, spec),
            "species": {"key": spec.key, "label": spec.label, "scientific_name": spec.scientific_name},
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "source": "NCBI GEO DataSets and PubMed E-utilities",
            "total_hits": snapshot["total_hits"],
            "total_hits_scope": "NCBI query count before exact single-organism filtering",
            "exact_single_organism_filter": True,
            "filter_diagnostics": snapshot["filter_diagnostics"],
            "page": page,
            "page_size": page_size,
            "total_pages": total_pages,
            "total_pages_upper_bound": total_pages,
            "pagination_scope": "globally ranked exact single-organism studies",
            "has_previous": page > 1,
            "has_next": page < total_pages,
            "returned_studies": len(studies),
            "studies": studies,
            "ranking_scope": f"global_first_{global_limit}_exact_single_organism_studies",
            "ranking_limit": global_limit,
            "evaluated_studies": evaluated_studies,
            "ranking_truncated": bool(snapshot["ranking_truncated"]),
            "cache_hit": bool(cache_hit),
            "ranking_build_seconds": snapshot["ranking_build_seconds"],
            "sort_by": public_sort,
            "sort_order": normalized_order,
            "sortable_fields": list(GLOBAL_SEARCH_SORT_ALIASES),
            "search_assessment": {
                "enabled": True,
                "basis": SEARCH_ASSESSMENT_BASIS,
                "default_sort": "deg_input_priority_desc",
                "tier_order": list(DEG_INPUT_PRIORITIES),
                "ranking_scope": "complete evaluated universe before pagination",
                "assessment_version": SEARCH_ASSESSMENT_VERSION,
                "confirmation_gate": (
                    "Only Prepare selection may inspect content and mark a candidate ready for contrast review."
                ),
            },
            "analysis_policy": {
                "cross_species_pooling": False,
                "selection_is_analysis": False,
                "note": "Human and Mouse selections, files, catalogs, and DEGORA runs remain independent.",
            },
        }
    page_search = getattr(geo_client, "search_summaries_page", None)
    if callable(page_search):
        total_hits, records, has_next = page_search(clean_query, spec, page=page, page_size=page_size)
    else:
        total_hits, records = geo_client.search_summaries(clean_query, spec, page=page, page_size=page_size)
        has_next = len(records) >= page_size and page * page_size < total_hits
    pmids = [pmid for record in records for pmid in _record_pmids(record)]
    publications = geo_client.publication_summaries(pmids)
    studies = [
        study
        for record in records
        if (study := _study_search_record(record, spec, publications)) is not None
    ]
    page_offset = (page - 1) * page_size
    for index, study in enumerate(studies, start=1):
        study["ncbi_relevance_rank"] = page_offset + index
    if assess_files:
        for study in studies:
            study["deg_input_assessment"] = _search_deg_input_assessment(study["accession"], geo_client)
        studies.sort(
            key=lambda study: (
                -_safe_int((study.get("deg_input_assessment") or {}).get("priority"), 0),
                _safe_int(study.get("ncbi_relevance_rank"), MAX_EXACT_PAGE_SCAN_RECORDS + 1),
            )
        )
    total_pages_upper_bound = max(1, math.ceil(total_hits / page_size)) if total_hits else 0
    return {
        **version_info,
        "query": clean_query,
        "ncbi_query": build_geo_query(clean_query, spec),
        "species": {"key": spec.key, "label": spec.label, "scientific_name": spec.scientific_name},
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "NCBI GEO DataSets and PubMed E-utilities",
        "total_hits": total_hits,
        "total_hits_scope": "NCBI query count before exact single-organism filtering",
        "exact_single_organism_filter": True,
        "page": page,
        "page_size": page_size,
        "total_pages": None,
        "total_pages_upper_bound": total_pages_upper_bound,
        "pagination_scope": "exact single-organism pages; total page count is intentionally not estimated from mixed-organism hits",
        "has_previous": page > 1,
        "has_next": has_next,
        "returned_studies": len(studies),
        "studies": studies,
        "sortable_fields": [
            "deg_input_priority",
            "ncbi_relevance_rank",
            "accession",
            "paper_title",
            "authors_display",
            "journal",
            "year",
            "n_samples",
        ],
        "search_assessment": {
            "enabled": bool(assess_files),
            "basis": SEARCH_ASSESSMENT_BASIS if assess_files else "Not requested.",
            "default_sort": "deg_input_priority_desc" if assess_files else "ncbi_relevance_rank_asc",
            "tier_order": list(DEG_INPUT_PRIORITIES),
            "confirmation_gate": "Only Prepare selection may inspect content and mark a candidate ready for contrast review.",
        },
        "analysis_policy": {
            "cross_species_pooling": False,
            "selection_is_analysis": False,
            "note": "Human and Mouse selections, files, catalogs, and DEGORA runs remain independent.",
        },
    }


def _exact_taxon(record: dict[str, Any], soft: dict[str, Any], species: SpeciesSpec) -> tuple[bool, str]:
    summary_taxon = str(record.get("taxon", "")).strip()
    soft_taxa = {str(value).strip() for value in soft.get("taxa", []) if str(value).strip()}
    observed = {summary_taxon} if summary_taxon else set()
    observed.update(soft_taxa)
    if observed == {species.scientific_name}:
        return True, species.scientific_name
    if not observed:
        return False, "organism metadata is missing"
    return False, "mixed or mismatched organism metadata: " + ", ".join(sorted(observed))


def _safe_materialized_name(accession: str, url: str) -> str:
    source_name = urllib.parse.unquote(Path(urllib.parse.urlsplit(url).path).name)
    clean = re.sub(r"[^A-Za-z0-9._-]+", "_", source_name).strip("._") or "candidate.dat"
    digest = hashlib.sha256(url.encode("utf-8")).hexdigest()[:10]
    return f"{accession}_{digest}_{clean}"[:220]


def _atomic_write_bytes(path: Path, payload: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary_name = tempfile.mkstemp(prefix=f".{path.name}.", suffix=".tmp", dir=path.parent)
    temporary = Path(temporary_name)
    try:
        with os.fdopen(fd, "wb") as handle:
            handle.write(payload)
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _atomic_write_text(path: Path, text: str) -> None:
    _atomic_write_bytes(path, text.encode("utf-8"))


def _series_matrix_url(accession: str) -> str:
    digits = accession.removeprefix("GSE")
    bucket = f"GSE{digits[:-3]}nnn" if len(digits) > 3 else "GSEnnn"
    return f"https://ftp.ncbi.nlm.nih.gov/geo/series/{bucket}/{accession}/matrix/{accession}_series_matrix.txt.gz"


def _validated_accessions(values: Iterable[str]) -> list[str]:
    accessions = list(dict.fromkeys(str(value).strip().upper() for value in values if str(value).strip()))
    if not accessions:
        raise DiscoveryError("at least one GEO Series accession must be selected")
    if len(accessions) > MAX_SELECTED_STUDIES:
        raise DiscoveryError(f"at most {MAX_SELECTED_STUDIES} studies can be prepared at once")
    invalid = [value for value in accessions if not re.fullmatch(r"GSE\d+", value)]
    if invalid:
        raise DiscoveryError("invalid GEO Series accession(s): " + ", ".join(invalid))
    return accessions


def _file_priority(item: dict[str, Any]) -> tuple[int, str]:
    return ({"strong": 0, "weak": 1, "upstream": 2}.get(str(item.get("tier")), 9), str(item.get("name", "")))


def _inspect_preparation_file(
    file_record: dict[str, Any],
    *,
    payload: bytes,
    fetch_scope: str,
    accession: str,
    target_dir: Path | None,
) -> dict[str, Any]:
    source_url = str(file_record["source_url"])
    if fetch_scope == "full" and urllib.parse.urlsplit(source_url).path.lower().endswith(".gz"):
        _validate_full_gzip(payload)
    role = str(file_record.get("role", "unknown_table"))
    if role in {"count_matrix", "normalized_expression_matrix"}:
        inspection = inspect_upstream_bytes(source_url, payload, declared_role=role)
    else:
        inspection = inspect_candidate_bytes(source_url, payload)
        if inspection.get("status") == "not_deg_table":
            upstream = inspect_upstream_bytes(source_url, payload, declared_role="unknown_matrix")
            if upstream.get("status") == "upstream_matrix_ready_for_contrast":
                inspection = upstream
                file_record["role"] = "unknown_matrix"
                file_record["tier"] = "upstream"
    inspection["fetch_scope"] = fetch_scope
    if target_dir is not None:
        local_path = target_dir / _safe_materialized_name(accession, source_url)
        _atomic_write_bytes(local_path, payload)
        inspection["local_path"] = str(local_path)
        inspection["full_file_sha256"] = hashlib.sha256(payload).hexdigest()
    return inspection


def _prepare_geo_studies_in_place(
    accessions: Iterable[str],
    species: str,
    *,
    query: str = "",
    inspection_budget: int = 40,
    max_files_per_study: int = 6,
    materialize_dir: str | Path | None = None,
    client: NcbiGeoClient | Any | None = None,
    _export_bundle: bool = True,
) -> dict[str, Any]:
    """Download and inspect selected GEO records without silently activating them."""

    spec = normalize_species(species)
    selected = _validated_accessions(accessions)
    if isinstance(inspection_budget, bool) or not isinstance(inspection_budget, int) or inspection_budget < 0:
        raise DiscoveryError("inspection_budget must be a non-negative whole number")
    if isinstance(max_files_per_study, bool) or not isinstance(max_files_per_study, int) or not 1 <= max_files_per_study <= 12:
        raise DiscoveryError("max_files_per_study must be a whole number from 1 to 12")
    target_dir = Path(materialize_dir).resolve() if materialize_dir is not None else None
    if target_dir is not None:
        target_dir.mkdir(parents=True, exist_ok=True)
    geo_client = client or DEFAULT_CLIENT
    records = geo_client.accession_summaries(selected, spec)
    by_accession = {str(record.get("accession") or "").upper(): record for record in records}
    pmids = [pmid for record in records for pmid in _record_pmids(record)]
    publications = geo_client.publication_summaries(pmids)
    studies: list[dict[str, Any]] = []
    excluded: list[dict[str, str]] = []
    inspections_used = 0

    for accession in selected:
        record = by_accession.get(accession)
        if record is None:
            excluded.append({"accession": accession, "reason": f"not found as an exact {spec.label} GEO Series"})
            continue
        try:
            soft = parse_geo_soft(geo_client.fetch_geo_soft(accession))
        except DiscoveryUnavailableError as exc:
            excluded.append({"accession": accession, "reason": f"unresolved SOFT record: {exc}"})
            continue
        taxon_ok, taxon_reason = _exact_taxon(record, soft, spec)
        if not taxon_ok:
            excluded.append({"accession": accession, "reason": taxon_reason})
            continue

        merged_pmids = _merged_pmids(_record_pmids(record), soft.get("pubmed_ids", []))
        missing_publications = [pmid for pmid in merged_pmids if pmid not in publications]
        if missing_publications:
            publications.update(geo_client.publication_summaries(missing_publications))
        enriched_record = {**record, "pubmedids": merged_pmids}

        files: list[dict[str, Any]] = []
        seen_source_urls: set[str] = set()
        for raw_url in soft.get("supplementary_files", []):
            try:
                safe_url = normalize_ncbi_url(raw_url)
            except DiscoveryError as exc:
                files.append(
                    {
                        "candidate_id": hashlib.sha256(f"{spec.key}|{accession}|{raw_url}".encode()).hexdigest()[:16],
                        "name": Path(urllib.parse.urlsplit(str(raw_url)).path).name,
                        "source_url": str(raw_url),
                        "tier": "reject",
                        "role": "unsupported",
                        "reason": str(exc),
                        "inspectable": False,
                        "inspection": {"status": "not_inspected", "reason": "unsafe source URL"},
                    }
                )
                continue
            if safe_url in seen_source_urls:
                continue
            seen_source_urls.add(safe_url)
            classified = classify_filename(safe_url)
            files.append(
                {
                    "candidate_id": hashlib.sha256(f"{spec.key}|{accession}|{safe_url}".encode()).hexdigest()[:16],
                    "source_url": safe_url,
                    **classified,
                    "inspection": {"status": "not_inspected", "reason": "not selected by the bounded inspection pass"},
                }
            )

        if not any(item.get("role") in {"count_matrix", "normalized_expression_matrix"} for item in files):
            matrix_url = _series_matrix_url(accession)
            classified = classify_filename(matrix_url)
            files.append(
                {
                    "candidate_id": hashlib.sha256(f"{spec.key}|{accession}|{matrix_url}".encode()).hexdigest()[:16],
                    "source_url": matrix_url,
                    **classified,
                    "reason": "GEO processed series-matrix fallback; use only after scale, annotation, and groups are verified",
                    "inspection": {"status": "not_inspected", "reason": "not selected by the bounded inspection pass"},
                }
            )

        inspectable = sorted((item for item in files if item.get("inspectable")), key=_file_priority)
        for file_record in inspectable[:max_files_per_study]:
            if inspections_used >= inspection_budget:
                file_record["inspection"] = {"status": "not_inspected", "reason": "inspection budget exhausted"}
                continue
            inspections_used += 1
            try:
                payload, fetch_scope = geo_client.fetch_candidate(str(file_record["source_url"]), full=target_dir is not None)
                file_record["inspection"] = _inspect_preparation_file(
                    file_record,
                    payload=payload,
                    fetch_scope=fetch_scope,
                    accession=accession,
                    target_dir=target_dir,
                )
            except (DiscoveryError, DiscoveryUnavailableError) as exc:
                file_record["inspection"] = {"status": "unresolved", "reason": str(exc)}

        ready_author = [item for item in files if item.get("inspection", {}).get("status") == "ready_for_review"]
        review_author = [
            item
            for item in files
            if item.get("inspection", {}).get("status")
            in {"candidate_header", "requires_lfc_confirmation", "requires_pvalue_mapping"}
        ]
        upstream = [
            item
            for item in files
            if item.get("inspection", {}).get("status") == "upstream_matrix_ready_for_contrast"
        ]
        if ready_author:
            preparation_status = "author_deg_ready_for_contrast_review"
        elif review_author:
            preparation_status = "author_table_requires_mapping_review"
        elif upstream:
            preparation_status = "upstream_matrix_requires_contrast"
        else:
            preparation_status = "no_usable_table_resolved"

        search_record = _study_search_record(enriched_record, spec, publications) or {}
        # Only the labeled fallback asks a reader to assign groups, so the extra
        # per-sample request is made only when one is actually on the table.
        sample_labels: dict[str, Any] = {}
        if any(item.get("inspection", {}).get("sample_columns") for item in files):
            sample_labels = _sample_labels_for(geo_client, accession)
            for item in files:
                columns = item.get("inspection", {}).get("sample_columns") or []
                if columns and sample_labels:
                    item["inspection"]["sample_labels"] = match_sample_labels(columns, sample_labels)
        studies.append(
            {
                **search_record,
                "title": str(record.get("title") or soft.get("title") or ""),
                "design": str(soft.get("design") or ""),
                "sample_labels": sample_labels,
                "files": files,
                "candidate_file_count": sum(item.get("tier") in {"strong", "weak", "upstream"} for item in files),
                "ready_for_review_count": len(ready_author),
                "upstream_matrix_count": len(upstream),
                "preparation_status": preparation_status,
            }
        )

    _assign_source_unit_components(studies)

    result = {
        "query": " ".join(_query_terms(query)) if str(query).strip() else "",
        "species": {"key": spec.key, "label": spec.label, "scientific_name": spec.scientific_name},
        "selected_accessions": selected,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "NCBI GEO DataSets, PubMed E-utilities, and GEO SOFT",
        "returned_studies": len(studies),
        "inspection_budget": inspection_budget,
        "inspections_used": inspections_used,
        "materialize_dir": str(target_dir) if target_dir is not None else "",
        "studies": studies,
        "excluded_studies": excluded,
        "analysis_policy": {
            "auto_run": False,
            "draft_rows_included": False,
            "cross_species_pooling": False,
            "minimum_independent_source_units": 2,
            "source_unit_rule": "GEO accessions linked to the same PubMed paper count as one source unit.",
            "author_table_gate": "Confirm contrast direction and table scope before activation.",
            "upstream_gate": (
                "Choose disjoint control/treatment samples with at least two biological replicates per group; "
                "count matrices use logCPM + Welch and normalized expression matrices require an explicit "
                "log2 or linear scale confirmation before the labeled Welch fallback."
            ),
            "note": "Human and Mouse catalogs and DEGORA runs are always generated separately.",
        },
    }
    if target_dir is not None and _export_bundle:
        result["exports"] = export_discovery_bundle(result, target_dir, force=True)
    return result


def _recognized_prepared_bundle(output: Path) -> bool:
    marker = output / DISCOVERY_BUNDLE_MARKER
    try:
        payload = json.loads(marker.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError):
        return False
    return bool(
        isinstance(payload, dict)
        and payload.get("artifact_type") == DISCOVERY_BUNDLE_ARTIFACT_TYPE
        and payload.get("format_version") == DISCOVERY_BUNDLE_FORMAT_VERSION
    )


def _validate_preparation_target(target: Path, *, force: bool) -> None:
    resolved = target.resolve()
    if resolved.parent == resolved or resolved == Path.home().resolve() or resolved == Path.cwd().resolve():
        raise DiscoveryError("preparation output must be a dedicated subdirectory")
    if target.exists() and not target.is_dir():
        raise FileExistsError(f"preparation output exists and is not a directory: {target}")
    if not target.exists() or not any(target.iterdir()):
        return
    if not force:
        raise FileExistsError(f"preparation output already exists and is not empty: {target}")
    if not _recognized_prepared_bundle(target):
        raise DiscoveryError("refusing --force because the output is not a recognized DEGORA prepared bundle")


def _retarget_materialized_paths(result: dict[str, Any], staging: Path, target: Path) -> None:
    for study in result.get("studies", []):
        for candidate in study.get("files", []):
            inspection = candidate.get("inspection") or {}
            local_path = str(inspection.get("local_path") or "")
            if not local_path:
                continue
            try:
                relative = Path(local_path).resolve().relative_to(staging.resolve())
            except ValueError as exc:
                raise DiscoveryError("prepared candidate path escaped the staging directory") from exc
            inspection["local_path"] = str(target / relative)


def _publish_prepared_bundle(staging: Path, target: Path, *, force: bool) -> None:
    _validate_preparation_target(target, force=force)
    existed_empty = target.exists() and not any(target.iterdir())
    backup: Path | None = None
    if target.exists() and not existed_empty:
        backup = Path(tempfile.mkdtemp(prefix=f".{target.name}.backup-", dir=target.parent))
        backup.rmdir()
        target.replace(backup)
    elif existed_empty:
        target.rmdir()
    try:
        staging.replace(target)
    except BaseException:
        if target.exists():
            shutil.rmtree(target)
        if backup is not None and backup.exists():
            backup.replace(target)
        elif existed_empty:
            target.mkdir(parents=True, exist_ok=True)
        raise
    if backup is not None and backup.exists():
        shutil.rmtree(backup)


def prepare_geo_studies(
    accessions: Iterable[str],
    species: str,
    *,
    query: str = "",
    inspection_budget: int = 40,
    max_files_per_study: int = 6,
    materialize_dir: str | Path | None = None,
    client: NcbiGeoClient | Any | None = None,
    force: bool = False,
) -> dict[str, Any]:
    """Prepare a bundle transactionally, publishing files only after all checks pass."""

    if materialize_dir is None:
        return _prepare_geo_studies_in_place(
            accessions,
            species,
            query=query,
            inspection_budget=inspection_budget,
            max_files_per_study=max_files_per_study,
            materialize_dir=None,
            client=client,
        )

    target = Path(materialize_dir).resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    _validate_preparation_target(target, force=force)
    staging = Path(tempfile.mkdtemp(prefix=f".{target.name}.prepare-", dir=target.parent))
    try:
        result = _prepare_geo_studies_in_place(
            accessions,
            species,
            query=query,
            inspection_budget=inspection_budget,
            max_files_per_study=max_files_per_study,
            materialize_dir=staging,
            client=client,
            _export_bundle=False,
        )
        _retarget_materialized_paths(result, staging, target)
        result["materialize_dir"] = str(target)
        # Same reason as the federated path: the persisted audit has to describe
        # the published bundle, not the staging directory it is written in.
        result["exports"] = {
            "output_dir": str(target),
            "audit_json": str(target / "discovery_audit.json"),
            "candidates_csv": str(target / "discovery_candidates.csv"),
            "draft_catalog_csv": str(target / "DEGORA_discovery_draft_catalog.csv"),
        }
        export_discovery_bundle(result, staging, force=True)
        marker_payload = {
            "artifact_type": DISCOVERY_BUNDLE_ARTIFACT_TYPE,
            "format_version": DISCOVERY_BUNDLE_FORMAT_VERSION,
            "species": str(result.get("species", {}).get("key") or ""),
        }
        _atomic_write_text(
            staging / DISCOVERY_BUNDLE_MARKER,
            json.dumps(marker_payload, indent=2, sort_keys=True) + "\n",
        )
        _publish_prepared_bundle(staging, target, force=force)
        return result
    finally:
        if staging.exists():
            shutil.rmtree(staging)


def discover_geo(
    query: str,
    species: str,
    *,
    limit: int = DEFAULT_LIMIT,
    inspect: bool = False,
    inspection_budget: int = DEFAULT_INSPECTION_BUDGET,
    materialize_dir: str | Path | None = None,
    client: NcbiGeoClient | Any | None = None,
) -> dict[str, Any]:
    """Compatibility one-shot search; preparation remains explicitly non-active."""

    search = search_geo(query, species, page=1, page_size=limit, client=client)
    if not inspect:
        return search
    prepared = prepare_geo_studies(
        [study["accession"] for study in search["studies"]],
        species,
        query=query,
        inspection_budget=inspection_budget,
        materialize_dir=materialize_dir,
        client=client,
    )
    prepared["search"] = {key: value for key, value in search.items() if key != "studies"}
    prepared["total_hits"] = search["total_hits"]
    return prepared


def _candidate_rows(result: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for study in result.get("studies", []):
        for item in study.get("files", []):
            inspection = item.get("inspection", {})
            mapping = inspection.get("mapping", {})
            rows.append(
                {
                    "species": study.get("species", ""),
                    "scientific_name": study.get("scientific_name", ""),
                    "accession": study.get("accession", ""),
                    "source_unit_id": study.get("source_unit_id", ""),
                    "study_title": study.get("title", ""),
                    "study_type": study.get("study_type", ""),
                    "release_date": study.get("release_date", ""),
                    "source_url": item.get("source_url", ""),
                    "filename": item.get("name", ""),
                    "filename_tier": item.get("tier", ""),
                    "filename_reason": item.get("reason", ""),
                    "inspection_status": inspection.get("status", ""),
                    "inspection_reason": inspection.get("reason", ""),
                    "gene_column": mapping.get("gene_column", ""),
                    "lfc_column": mapping.get("lfc_column", ""),
                    "p_column": mapping.get("p_column", ""),
                    "padj_column": mapping.get("padj_column", ""),
                    "header_row": inspection.get("header_row", ""),
                    "sheet_name": inspection.get("sheet_name", ""),
                    "fetch_scope": inspection.get("fetch_scope", ""),
                    "payload_sha256": inspection.get("payload_sha256", ""),
                    "full_file_sha256": inspection.get("full_file_sha256", ""),
                    "source_path": inspection.get("local_path", ""),
                }
            )
    return rows


def _draft_catalog_rows(result: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index, candidate in enumerate(_candidate_rows(result), start=1):
        if candidate["inspection_status"] not in {
            "ready_for_review",
            "candidate_header",
            "requires_lfc_confirmation",
            "requires_pvalue_mapping",
        }:
            continue
        rows.append(
            {
                "study_id": f"{candidate['species']}_{candidate['accession']}_{index:03d}",
                "paper_id": candidate["source_unit_id"],
                "source_unit_id": candidate["source_unit_id"],
                "source_path": candidate["source_path"],
                "source_url": candidate["source_url"],
                "gene_column": candidate["gene_column"],
                "lfc_column": candidate["lfc_column"],
                "p_column": candidate["p_column"],
                "padj_column": candidate["padj_column"],
                "sheet_name": candidate["sheet_name"],
                "species": candidate["scientific_name"],
                "assay_type": candidate["study_type"],
                "source_input_type": "author_deg_table_candidate",
                "table_scope": "auto",
                "include_in_analysis": "no",
                "curation_status": "requires_manual_review",
                "curation_notes": (
                    "Confirm biological contrast, biological replicates, log2FC scale, source-unit independence, "
                    "table scope, and every inferred column before changing include_in_analysis."
                ),
            }
        )
    return rows


def _csv_text(rows: list[dict[str, Any]], columns: list[str]) -> str:
    def spreadsheet_safe(value: Any) -> Any:
        if not isinstance(value, str):
            return value
        stripped = value.lstrip(" \t\r\n")
        if stripped.startswith(("=", "+", "-", "@")):
            return "'" + value
        return value

    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(buffer, fieldnames=columns, extrasaction="ignore", lineterminator="\n")
    writer.writeheader()
    writer.writerows({key: spreadsheet_safe(value) for key, value in row.items()} for row in rows)
    return buffer.getvalue()


def export_search_page(result: dict[str, Any], output_dir: str | Path, *, force: bool = False) -> dict[str, str]:
    output = Path(output_dir).resolve()
    output.mkdir(parents=True, exist_ok=True)
    json_path = output / "geo_search_page.json"
    csv_path = output / "geo_search_page.csv"
    existing = [path for path in (json_path, csv_path) if path.exists()]
    if existing and not force:
        raise FileExistsError("search output already exists; use --force to replace: " + ", ".join(map(str, existing)))
    rows = []
    for study in result.get("studies", []):
        row = dict(study)
        row["authors"] = "; ".join(map(str, row.get("authors", [])))
        row["pubmed_ids"] = "; ".join(map(str, row.get("pubmed_ids", [])))
        assessment = row.pop("deg_input_assessment", {}) or {}
        counts = assessment.get("counts", {}) or {}
        row["deg_input_likelihood_tier"] = assessment.get("tier", "")
        row["deg_input_likelihood_label"] = assessment.get("label", "")
        row["deg_input_priority"] = assessment.get("priority", "")
        row["deg_input_likelihood_basis"] = assessment.get("basis", "")
        row["deg_input_candidate_filenames"] = "; ".join(map(str, assessment.get("candidate_files", [])))
        row["deg_input_deg_like_count"] = counts.get("deg_like", 0)
        row["deg_input_tabular_count"] = counts.get("tabular", 0)
        row["deg_input_matrix_count"] = counts.get("matrix", 0)
        row["deg_input_assessment_error"] = assessment.get("error", "")
        rows.append(row)
    columns = list(rows[0]) if rows else [
        "species",
        "scientific_name",
        "accession",
        "paper_title",
        "dataset_title",
        "authors",
        "authors_display",
        "journal",
        "year",
        "publication_date",
        "pubmed_ids",
        "pubmed_url",
        "study_type",
        "release_date",
        "n_samples",
        "source_url",
        "ncbi_relevance_rank",
        "deg_input_likelihood_tier",
        "deg_input_likelihood_label",
        "deg_input_priority",
        "deg_input_likelihood_basis",
        "deg_input_candidate_filenames",
        "deg_input_deg_like_count",
        "deg_input_tabular_count",
        "deg_input_matrix_count",
        "deg_input_assessment_error",
    ]
    _atomic_write_text(json_path, json.dumps(result, indent=2, sort_keys=True, ensure_ascii=False) + "\n")
    _atomic_write_text(csv_path, _csv_text(rows, columns))
    return {"output_dir": str(output), "search_json": str(json_path), "search_csv": str(csv_path)}


def export_discovery_bundle(result: dict[str, Any], output_dir: str | Path, *, force: bool = False) -> dict[str, str]:
    output = Path(output_dir).resolve()
    output.mkdir(parents=True, exist_ok=True)
    audit_path = output / "discovery_audit.json"
    candidates_path = output / "discovery_candidates.csv"
    catalog_path = output / "DEGORA_discovery_draft_catalog.csv"
    targets = (audit_path, candidates_path, catalog_path)
    existing = [path for path in targets if path.exists()]
    if existing and not force:
        raise FileExistsError("discovery output already exists; use --force to replace: " + ", ".join(map(str, existing)))

    candidates = _candidate_rows(result)
    catalog = _draft_catalog_rows(result)
    candidate_columns = list(candidates[0]) if candidates else [
        "species",
        "scientific_name",
        "accession",
        "source_unit_id",
        "study_title",
        "study_type",
        "release_date",
        "source_url",
        "filename",
        "filename_tier",
        "filename_reason",
        "inspection_status",
        "inspection_reason",
        "gene_column",
        "lfc_column",
        "p_column",
        "padj_column",
        "header_row",
        "sheet_name",
        "fetch_scope",
        "payload_sha256",
        "full_file_sha256",
        "source_path",
    ]
    catalog_columns = list(catalog[0]) if catalog else [
        "study_id",
        "paper_id",
        "source_unit_id",
        "source_path",
        "source_url",
        "gene_column",
        "lfc_column",
        "p_column",
        "padj_column",
        "sheet_name",
        "species",
        "assay_type",
        "source_input_type",
        "table_scope",
        "include_in_analysis",
        "curation_status",
        "curation_notes",
    ]
    _atomic_write_text(audit_path, json.dumps(result, indent=2, sort_keys=True, ensure_ascii=False) + "\n")
    _atomic_write_text(candidates_path, _csv_text(candidates, candidate_columns))
    _atomic_write_text(catalog_path, _csv_text(catalog, catalog_columns))
    return {
        "output_dir": str(output),
        "audit_json": str(audit_path),
        "candidates_csv": str(candidates_path),
        "draft_catalog_csv": str(catalog_path),
    }


__all__ = [
    "DiscoveryError",
    "DiscoveryUnavailableError",
    "NcbiGeoClient",
    "NcbiRequestConfig",
    "SpeciesSpec",
    "SPECIES_BY_KEY",
    "build_geo_query",
    "classify_filename",
    "classify_header",
    "discover_geo",
    "export_discovery_bundle",
    "export_search_page",
    "inspect_candidate_bytes",
    "inspect_upstream_bytes",
    "normalize_ncbi_url",
    "normalize_species",
    "parse_geo_soft",
    "prepare_geo_studies",
    "search_geo",
]
