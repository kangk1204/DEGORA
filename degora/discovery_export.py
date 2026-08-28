"""Audit-grade exports for source-neutral publication discovery snapshots."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import tempfile
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .formula_safety import neutralize_formula_cell
from .discovery_store import sanitize_discovery_payload
from .provenance import (
    apply_default_file_mode,
    output_directory_lock,
    publication_target_lock_path,
    publish_staged_artifacts,
)


SEARCH_JSON_NAME = "publication_search.json"
SEARCH_CSV_NAME = "publication_search.csv"
SEARCH_XLSX_NAME = "DEGORA_search_results.xlsx"
SEARCH_MANIFEST_NAME = "publication_search.manifest.json"
SEARCH_EXPORT_ARTIFACT_TYPE = "degora_publication_search_export_set"
SEARCH_EXPORT_FORMAT_VERSION = 1


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _export_set_manifest(paths: Iterable[Path]) -> dict[str, Any]:
    files = {
        path.name: {"sha256": _sha256(path), "size_bytes": path.stat().st_size}
        for path in paths
    }
    generation_text = "\n".join(f"{name}:{entry['sha256']}" for name, entry in sorted(files.items()))
    return {
        "artifact_type": SEARCH_EXPORT_ARTIFACT_TYPE,
        "format_version": SEARCH_EXPORT_FORMAT_VERSION,
        "generation_sha256": hashlib.sha256(generation_text.encode("utf-8")).hexdigest(),
        "files": files,
    }


def verify_publication_search_export(output_dir: str | Path) -> dict[str, Any]:
    """Verify that all fixed-name search exports belong to one published generation."""

    output = Path(output_dir).resolve()
    manifest_path = output / SEARCH_MANIFEST_NAME
    try:
        manifest = json.loads(
            manifest_path.read_text(encoding="utf-8"),
            parse_constant=lambda value: (_ for _ in ()).throw(ValueError(f"non-finite JSON value: {value}")),
        )
    except (OSError, UnicodeError, json.JSONDecodeError, ValueError) as exc:
        raise ValueError(f"publication search export manifest is missing or invalid: {manifest_path}") from exc
    if (
        not isinstance(manifest, dict)
        or manifest.get("artifact_type") != SEARCH_EXPORT_ARTIFACT_TYPE
        or manifest.get("format_version") != SEARCH_EXPORT_FORMAT_VERSION
    ):
        raise ValueError("publication search export manifest has an unsupported artifact type or format version")
    expected_names = {SEARCH_JSON_NAME, SEARCH_CSV_NAME, SEARCH_XLSX_NAME}
    files = manifest.get("files")
    if not isinstance(files, dict) or set(files) != expected_names:
        raise ValueError("publication search export manifest does not name the complete artifact set")
    for name in sorted(expected_names):
        entry = files.get(name)
        path = output / name
        if not isinstance(entry, dict) or not path.is_file():
            raise ValueError(f"publication search export artifact is missing: {name}")
        if entry.get("sha256") != _sha256(path) or entry.get("size_bytes") != path.stat().st_size:
            raise ValueError(f"publication search export artifact does not match its generation manifest: {name}")
    expected = _export_set_manifest([output / name for name in sorted(expected_names)])
    if manifest.get("generation_sha256") != expected["generation_sha256"]:
        raise ValueError("publication search export generation digest is invalid")
    return manifest


def _safe_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, (list, tuple, set)):
        value = "; ".join(str(item) for item in value)
    elif isinstance(value, dict):
        value = json.dumps(value, sort_keys=True, ensure_ascii=False)
    else:
        value = str(value)
    return neutralize_formula_cell(value)


def _atomic_bytes(path: Path, payload: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{path.name}.", suffix=".tmp", dir=path.parent)
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(payload)
        apply_default_file_mode(temporary)
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _atomic_text(path: Path, text: str) -> None:
    _atomic_bytes(path, text.encode("utf-8"))


def _as_list(value: Any) -> list[Any]:
    if isinstance(value, list):
        return value
    if value in (None, ""):
        return []
    return [value]


def _readiness(record: dict[str, Any]) -> dict[str, Any]:
    value = record.get("data_readiness") or record.get("deg_input_assessment") or {}
    return value if isinstance(value, dict) else {}


def publication_rows(snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in snapshot.get("records", snapshot.get("studies", [])):
        readiness = _readiness(record)
        rows.append(
            {
                "paper_title": record.get("paper_title") or record.get("title", ""),
                "authors": record.get("authors_display") or record.get("authors", []),
                "journal": record.get("journal", ""),
                "year": record.get("year", ""),
                "canonical_id": record.get("canonical_id") or record.get("source_unit_id") or record.get("accession", ""),
                "record_kind": record.get("record_kind", "publication"),
                "species": record.get("species", ""),
                "species_decision": record.get("species_decision", ""),
                "pubmed_ids": record.get("pubmed_ids", []),
                "doi": record.get("doi", ""),
                "pmcid": record.get("pmcid", ""),
                "geo_accessions": record.get("geo_accessions") or _as_list(record.get("accession")),
                "source_unit_id": record.get("source_unit_id", ""),
                "source_unit_conflict": record.get("source_unit_conflict", []),
                "shared_submission_units": record.get("shared_submission_units", []),
                "shared_submission_warning": record.get("shared_submission_warning", ""),
                "resolution_state": record.get("resolution_state", ""),
                "relevance_rank": record.get("relevance_rank", record.get("ncbi_relevance_rank", "")),
                "readiness_tier": readiness.get("tier", ""),
                "readiness_priority": readiness.get("priority", ""),
                "verification_state": readiness.get("verification_state", "likely"),
                "readiness_basis": readiness.get("basis", ""),
                "candidate_count": len(_as_list(record.get("candidates"))),
            }
        )
    return rows


def _identifier_rows(records: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in records:
        canonical = record.get("canonical_id") or record.get("source_unit_id") or record.get("accession", "")
        values: list[tuple[str, str]] = []
        values.extend(("PMID", str(value)) for value in _as_list(record.get("pubmed_ids")))
        values.extend(("DOI", str(value)) for value in _as_list(record.get("doi")))
        values.extend(("PMCID", str(value)) for value in _as_list(record.get("pmcid")))
        values.extend(("GEO", str(value)) for value in _as_list(record.get("geo_accessions") or record.get("accession")))
        provider_ids = record.get("provider_ids") or {}
        if isinstance(provider_ids, dict):
            for provider, provider_values in provider_ids.items():
                values.extend((str(provider), str(value)) for value in _as_list(provider_values))
        elif isinstance(provider_ids, list):
            for value in provider_ids:
                text = str(value).strip()
                provider, separator, identifier = text.partition(":")
                values.append((provider if separator else "provider", identifier if separator else text))
        seen: set[tuple[str, str]] = set()
        for kind, value in values:
            key = (kind, value.strip())
            if not key[1] or key in seen:
                continue
            seen.add(key)
            rows.append({"canonical_id": canonical, "identifier_type": key[0], "identifier": key[1]})
    return rows


def _dataset_rows(records: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in records:
        canonical = record.get("canonical_id") or record.get("source_unit_id") or record.get("accession", "")
        for accession in _as_list(record.get("geo_accessions") or record.get("accession")):
            if str(accession).strip():
                rows.append({"canonical_id": canonical, "provider": "NCBI GEO", "accession": accession})
        for source in _as_list(record.get("sources")):
            if isinstance(source, dict):
                rows.append(
                    {
                        "canonical_id": canonical,
                        "provider": source.get("provider", ""),
                        "accession": source.get("accession", source.get("id", "")),
                        "landing_url": source.get("landing_url", source.get("source_url", "")),
                    }
                )
    return rows


def _candidate_rows(records: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in records:
        canonical = record.get("canonical_id") or record.get("source_unit_id") or record.get("accession", "")
        for candidate in _as_list(record.get("candidates")):
            if not isinstance(candidate, dict):
                continue
            rows.append(
                {
                    "canonical_id": canonical,
                    "candidate_id": candidate.get("candidate_id", ""),
                    "provider": candidate.get("provider", ""),
                    "name": candidate.get("name", ""),
                    "role": candidate.get("role", ""),
                    "source_input_type": candidate.get("source_input_type", ""),
                    "verification_state": candidate.get("verification_state", "likely"),
                    "source_url": candidate.get("source_url", ""),
                    "landing_url": candidate.get("landing_url", ""),
                    "license": candidate.get("license", ""),
                    "reason": candidate.get("reason", ""),
                }
            )
    return rows


def _species_rows(records: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in records:
        rows.append(
            {
                "canonical_id": record.get("canonical_id") or record.get("source_unit_id") or record.get("accession", ""),
                "requested_species": record.get("species", ""),
                "species_decision": record.get("species_decision", ""),
                "target_species_verified": record.get("target_species_verified", False),
                "species_evidence": record.get("species_evidence", ""),
            }
        )
    return rows


def _event_rows(snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    diagnostics = snapshot.get("provider_events") or snapshot.get("provider_diagnostics") or []
    if isinstance(diagnostics, dict):
        diagnostics = [dict(provider=key, detail=value) for key, value in diagnostics.items()]
    return [value if isinstance(value, dict) else {"detail": value} for value in _as_list(diagnostics)]


def _write_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]], columns: list[str]) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(columns)
    for row in rows:
        worksheet.append([_safe_cell(row.get(column, "")) for column in columns])
    header_fill = PatternFill("solid", fgColor="18364A")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, column in enumerate(columns, start=1):
        values = [str(column)] + [str(_safe_cell(row.get(column, ""))) for row in rows[:200]]
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(max(map(len, values)) + 2, 11), 48)


def build_publication_search_workbook(snapshot: dict[str, Any]) -> bytes:
    records = list(snapshot.get("records", snapshot.get("studies", [])))
    publications = publication_rows(snapshot)
    species_value = snapshot.get("species", "")
    if isinstance(species_value, dict):
        species_value = species_value.get("key") or species_value.get("label") or ""
    query_rows = [
        {"field": "query", "value": snapshot.get("query", "")},
        {"field": "species", "value": species_value},
        {"field": "generated_at", "value": snapshot.get("generated_at", "")},
        {"field": "evaluated_records", "value": snapshot.get("evaluated_records", len(records))},
        {"field": "ranking_limit", "value": snapshot.get("ranking_limit", "")},
        {"field": "ranking_truncated", "value": snapshot.get("ranking_truncated", False)},
        {"field": "ranking_contract", "value": snapshot.get("ranking_contract", "")},
        {"field": "cross_species_pooling", "value": False},
    ]
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_sheet(workbook, "Query", query_rows, ["field", "value"])
    _write_sheet(workbook, "Publications", publications, list(publications[0]) if publications else [
        "canonical_id", "record_kind", "species", "species_decision", "paper_title", "authors", "journal",
        "year", "pubmed_ids", "doi", "pmcid", "geo_accessions", "source_unit_id", "source_unit_conflict",
        "shared_submission_units", "shared_submission_warning", "resolution_state", "relevance_rank",
        "readiness_tier", "readiness_priority", "verification_state", "readiness_basis", "candidate_count",
    ])
    identifiers = _identifier_rows(records)
    _write_sheet(workbook, "Identifiers", identifiers, list(identifiers[0]) if identifiers else ["canonical_id", "identifier_type", "identifier"])
    datasets = _dataset_rows(records)
    _write_sheet(workbook, "Linked datasets", datasets, list(datasets[0]) if datasets else ["canonical_id", "provider", "accession", "landing_url"])
    candidates = _candidate_rows(records)
    _write_sheet(workbook, "Candidate routes", candidates, list(candidates[0]) if candidates else [
        "canonical_id", "candidate_id", "provider", "name", "role", "source_input_type",
        "verification_state", "source_url", "landing_url", "license", "reason",
    ])
    species = _species_rows(records)
    _write_sheet(workbook, "Species decisions", species, list(species[0]) if species else [
        "canonical_id", "requested_species", "species_decision", "target_species_verified", "species_evidence",
    ])
    events = _event_rows(snapshot)
    event_keys = {key for row in events for key in row}
    event_columns = [key for key in ("provider", "event", "status", "message") if key in event_keys]
    event_columns.extend(sorted(event_keys.difference(event_columns)))
    if not event_columns:
        event_columns = ["provider", "event", "status", "message"]
    _write_sheet(workbook, "Provider events", events, event_columns)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _publication_csv(rows: list[dict[str, Any]]) -> str:
    columns = list(rows[0]) if rows else ["canonical_id", "paper_title", "species", "species_decision"]
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(buffer, fieldnames=columns, extrasaction="ignore", lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow({column: _safe_cell(row.get(column, "")) for column in columns})
    return buffer.getvalue()


def export_publication_search(snapshot: dict[str, Any], output_dir: str | Path, *, force: bool = False) -> dict[str, str]:
    output = Path(output_dir).resolve()
    json_path = output / SEARCH_JSON_NAME
    csv_path = output / SEARCH_CSV_NAME
    xlsx_path = output / SEARCH_XLSX_NAME
    manifest_path = output / SEARCH_MANIFEST_NAME
    safe_snapshot = sanitize_discovery_payload(snapshot)
    rows = publication_rows(safe_snapshot)
    with output_directory_lock(publication_target_lock_path(output)):
        # Preserve the public API contract: callers may name a new nested output
        # directory.  The stable publication lock lives beside the target and
        # therefore does not create the target itself.
        output.mkdir(parents=True, exist_ok=True)
        existing = [path for path in (json_path, csv_path, xlsx_path, manifest_path) if path.exists()]
        if existing and not force:
            raise FileExistsError(
                "federated search output already exists; use --force to replace: "
                + ", ".join(map(str, existing))
            )
        with tempfile.TemporaryDirectory(prefix=".degora-search-export-", dir=output) as staging_name:
            staging = Path(staging_name)
            staged_json = staging / json_path.name
            staged_csv = staging / csv_path.name
            staged_xlsx = staging / xlsx_path.name
            staged_manifest = staging / manifest_path.name
            _atomic_text(
                staged_json,
                json.dumps(
                    safe_snapshot,
                    indent=2,
                    sort_keys=True,
                    ensure_ascii=False,
                    allow_nan=False,
                )
                + "\n",
            )
            _atomic_text(staged_csv, _publication_csv(rows))
            _atomic_bytes(staged_xlsx, build_publication_search_workbook(safe_snapshot))
            _atomic_text(
                staged_manifest,
                json.dumps(
                    _export_set_manifest([staged_json, staged_csv, staged_xlsx]),
                    indent=2,
                    sort_keys=True,
                    allow_nan=False,
                )
                + "\n",
            )
            publish_staged_artifacts(
                {
                    staged_json: json_path,
                    staged_csv: csv_path,
                    staged_xlsx: xlsx_path,
                    # The generation manifest is intentionally published last.
                    # A reader racing publication sees either matching hashes or
                    # a detectable incomplete/mixed generation, never silent mix.
                    staged_manifest: manifest_path,
                }
            )
    return {
        "output_dir": str(output),
        "search_json": str(json_path),
        "search_csv": str(csv_path),
        "search_xlsx": str(xlsx_path),
        "manifest": str(manifest_path),
    }


__all__ = [
    "SEARCH_CSV_NAME",
    "SEARCH_JSON_NAME",
    "SEARCH_XLSX_NAME",
    "SEARCH_MANIFEST_NAME",
    "build_publication_search_workbook",
    "export_publication_search",
    "publication_rows",
    "verify_publication_search_export",
]
