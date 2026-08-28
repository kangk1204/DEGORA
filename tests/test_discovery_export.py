from __future__ import annotations

import csv
import io
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from degora.discovery_export import build_publication_search_workbook, export_publication_search


def _snapshot() -> dict:
    return {
        "query": "hypoxia",
        "species": {"key": "human", "scientific_name": "Homo sapiens"},
        "generated_at": "2026-08-09T00:00:00+00:00",
        "evaluated_records": 1,
        "ranking_limit": 1000,
        "ranking_truncated": False,
        "ranking_contract": "readiness_desc,relevance_asc,year_desc,canonical_id_asc",
        "provider_events": [{"provider": "Europe PMC", "status": "ok"}],
        "records": [
            {
                "canonical_id": "PMID:123",
                "record_kind": "publication",
                "species": "human",
                "species_decision": "exact",
                "species_evidence": "Homo sapiens",
                "target_species_verified": True,
                "paper_title": '=HYPERLINK("https://invalid.example","open")',
                "authors": ["Kim K", "Lee J"],
                "authors_display": "Kim K, Lee J",
                "journal": "Genome Biology",
                "year": 2025,
                "pubmed_ids": ["123"],
                "doi": "10.1000/example",
                "pmcid": "PMC123",
                "geo_accessions": ["GSE1"],
                "provider_ids": {"zenodo": ["999"]},
                "source_unit_id": "PMID:123",
                "shared_submission_units": ["GSE2"],
                "shared_submission_warning": "shares its title with another repository record",
                "relevance_rank": 4,
                "data_readiness": {
                    "tier": "author_deg_likely",
                    "priority": 60,
                    "verification_state": "likely",
                    "basis": "filename only",
                },
                "candidates": [
                    {
                        "candidate_id": "c1",
                        "provider": "NCBI GEO",
                        "name": "+results.csv",
                        "role": "deg_table",
                        "source_input_type": "author_deg_table",
                        "source_url": "https://ftp.ncbi.nlm.nih.gov/results.csv",
                    }
                ],
            }
        ],
    }


def test_workbook_contains_audit_sheets_and_neutralizes_formulas() -> None:
    workbook = load_workbook(io.BytesIO(build_publication_search_workbook(_snapshot())), data_only=False)
    assert workbook.sheetnames == [
        "Query",
        "Publications",
        "Identifiers",
        "Linked datasets",
        "Candidate routes",
        "Species decisions",
        "Provider events",
    ]
    publications = workbook["Publications"]
    title_column = [cell.value for cell in publications[1]].index("paper_title") + 1
    assert publications.cell(2, title_column).value.startswith("'=")
    publications_header = [cell.value for cell in publications[1]]
    assert "shared_submission_units" in publications_header
    warning_column = publications_header.index("shared_submission_warning") + 1
    assert publications.cell(2, warning_column).value == "shares its title with another repository record"
    candidates = workbook["Candidate routes"]
    name_column = [cell.value for cell in candidates[1]].index("name") + 1
    assert candidates.cell(2, name_column).value == "'+results.csv"
    assert workbook["Query"].freeze_panes == "A2"


def test_formula_neutralization_detects_prefix_after_whitespace(tmp_path: Path) -> None:
    snapshot = _snapshot()
    snapshot["records"][0]["paper_title"] = " \t=HYPERLINK(\"https://invalid.example\",\"open\")"
    snapshot["records"][0]["candidates"][0]["name"] = "\r\n+results.csv"

    workbook = load_workbook(io.BytesIO(build_publication_search_workbook(snapshot)), data_only=False)
    publications = workbook["Publications"]
    title_column = [cell.value for cell in publications[1]].index("paper_title") + 1
    assert publications.cell(2, title_column).value.startswith("' \t=")
    candidates = workbook["Candidate routes"]
    name_column = [cell.value for cell in candidates[1]].index("name") + 1
    assert candidates.cell(2, name_column).value.startswith("'\n+")

    with Path(export_publication_search(snapshot, tmp_path)["search_csv"]).open(
        newline="",
        encoding="utf-8",
    ) as handle:
        row = next(csv.DictReader(handle))
    assert row["paper_title"].startswith("' \t=")


@pytest.mark.parametrize("literal", ["#VALUE!", "#REF!", " \t#VALUE!", "\r\n#REF!"])
def test_formula_neutralization_preserves_excel_error_literals_as_text(literal: str) -> None:
    snapshot = _snapshot()
    snapshot["records"][0]["paper_title"] = literal

    workbook = load_workbook(io.BytesIO(build_publication_search_workbook(snapshot)), data_only=False)
    publications = workbook["Publications"]
    title_column = [cell.value for cell in publications[1]].index("paper_title") + 1
    cell = publications.cell(2, title_column)

    assert cell.data_type == "s"
    assert cell.value.startswith("'")
    assert cell.value[1:].lstrip(" \t\r\n") == literal.lstrip(" \t\r\n")


def test_export_writes_json_csv_and_xlsx_atomically_and_refuses_overwrite(tmp_path: Path) -> None:
    result = export_publication_search(_snapshot(), tmp_path)
    assert Path(result["search_json"]).is_file()
    assert Path(result["search_csv"]).is_file()
    assert Path(result["search_xlsx"]).is_file()
    assert json.loads(Path(result["search_json"]).read_text(encoding="utf-8"))["query"] == "hypoxia"
    with Path(result["search_csv"]).open(newline="", encoding="utf-8") as handle:
        row = next(csv.DictReader(handle))
    assert row["paper_title"].startswith("'=")
    assert row["shared_submission_units"] == "GSE2"
    assert row["shared_submission_warning"] == "shares its title with another repository record"
    with pytest.raises(FileExistsError, match="--force"):
        export_publication_search(_snapshot(), tmp_path)
    export_publication_search(_snapshot(), tmp_path, force=True)
    assert not list(tmp_path.glob(".*.tmp"))


def test_export_creates_a_new_nested_output_directory(tmp_path: Path) -> None:
    output = tmp_path / "new" / "nested" / "search"

    result = export_publication_search(_snapshot(), output)

    assert Path(result["search_json"]).is_file()
    assert Path(result["search_csv"]).is_file()
    assert Path(result["search_xlsx"]).is_file()
    assert Path(result["manifest"]).is_file()
