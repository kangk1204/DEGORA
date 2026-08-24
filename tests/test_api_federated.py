from __future__ import annotations

import io
import json
import re
import sys
import threading
import time
import types
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook
import pytest

import degora.discovery as discovery
from degora.api import INDEX_HTML, create_server


def _request_json(
    url: str,
    *,
    payload: dict | None = None,
    action: bool = False,
    token: str | None = None,
) -> tuple[int, dict]:
    data = None
    headers = {}
    if payload is not None:
        data = json.dumps(payload).encode()
        headers["Content-Type"] = "application/json"
    if action:
        headers["X-DEGORA-Action"] = "1"
    if token:
        headers["X-DEGORA-Token"] = token
    request = urllib.request.Request(url, data=data, headers=headers)
    with urllib.request.urlopen(request, timeout=10) as response:
        return response.status, json.loads(response.read().decode())


def _request_bytes(url: str, *, token: str | None = None) -> tuple[int, dict[str, str], bytes]:
    headers = {"X-DEGORA-Token": token} if token else {}
    request = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(request, timeout=10) as response:
        return response.status, dict(response.headers.items()), response.read()


def _start_server(tmp_path: Path):
    db = tmp_path / "degora.db"
    db.touch()
    server = create_server(db, port=0, quiet=True, discovery_root=tmp_path / "discovery")
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    host, port = server.server_address
    return server, thread, f"http://{host}:{port}"


def _stop_server(server, thread) -> None:
    server.shutdown()
    server.server_close()
    thread.join(timeout=5)


def _install_federated_module(monkeypatch, *, calls: list[dict]) -> None:
    module = types.ModuleType("degora.discovery_federated")

    def search_publications(*, query, species, limit):
        records = []
        for index in range(120):
            records.append(
                {
                    "publication_id": f"{species}-{index:03d}",
                    "title": "=formula risk" if index == 0 else f"{species} paper {index:03d}",
                    "authors": f"Author {index}",
                    "journal": "Journal",
                    "year": 2000 + (index % 20),
                    "readiness": "ready" if index % 2 == 0 else "candidate",
                    "relevance": 120 - index,
                    "data_sources": ["GEO", f"GSE{index:04d}"],
                    "candidate_routes": ["author_deg_table"],
                    "species_decision": species,
                }
            )
        return {
            "query": query,
            "species": {"key": species},
            "limit": limit,
            "records": records,
            "total": len(records),
            "provider_events": [{"provider": "fake", "event": "search", "status": "ok", "message": "+neutralize"}],
        }

    def page_publication_snapshot(snapshot, *, page, page_size, sort_by, sort_order, text_filter=""):
        calls.append(
            {
                "page": page,
                "page_size": page_size,
                "sort_by": sort_by,
                "sort_order": sort_order,
                "text_filter": text_filter,
            }
        )
        records = list(snapshot["records"])
        if text_filter:
            wanted = text_filter.lower()
            records = [record for record in records if wanted in str(record.get("title", "")).lower()]
        reverse = sort_order == "desc"
        if sort_by == "year":
            records.sort(key=lambda record: record["year"], reverse=reverse)
        elif sort_by == "title":
            records.sort(key=lambda record: record["title"], reverse=reverse)
        elif sort_by == "relevance":
            records.sort(key=lambda record: record["relevance"], reverse=reverse)
        offset = (page - 1) * page_size
        return {
            "records": records[offset : offset + page_size],
            "total": len(records),
            "page": page,
            "page_size": page_size,
            "total_pages": 6,
            "sort_by": sort_by,
            "sort_order": sort_order,
            "has_next": page < 6,
        }

    module.search_publications = search_publications
    module.page_publication_snapshot = page_publication_snapshot
    module.resolve_publication_records = lambda records, species: list(records)
    monkeypatch.setitem(sys.modules, "degora.discovery_federated", module)


def _create_search(base: str, species: str) -> dict:
    status, created = _request_json(
        f"{base}/api/discovery/searches",
        payload={"query": "hypoxia", "species": species, "limit": 120},
        action=True,
    )
    assert status == 202
    deadline = time.time() + 5
    while time.time() < deadline:
        _, job_payload = _request_json(f"{base}/api/discovery/jobs/{created['job_id']}")
        if job_payload["job"]["status"] == "complete":
            return created
        time.sleep(0.05)
    raise AssertionError("search job did not complete")


def test_federated_search_async_pagination_sort_restart_and_export(tmp_path: Path, monkeypatch) -> None:
    page_calls: list[dict] = []
    _install_federated_module(monkeypatch, calls=page_calls)
    server, thread, base = _start_server(tmp_path)
    try:
        created = _create_search(base, "human")
        _, search_payload = _request_json(f"{base}/api/discovery/searches/{created['search_id']}")
        _, first_page = _request_json(f"{base}/api/discovery/searches/{created['search_id']}/records?page=1&page_size=20")
        _, sorted_page = _request_json(
            f"{base}/api/discovery/searches/{created['search_id']}/records?"
            f"{urllib.parse.urlencode({'page': 1, 'page_size': 20, 'sort_by': 'year', 'sort_order': 'asc'})}"
        )
        status, headers, xlsx = _request_bytes(f"{base}/api/discovery/searches/{created['search_id']}/export.xlsx")
    finally:
        _stop_server(server, thread)

    assert search_payload["search"]["status"] == "complete"
    assert search_payload["search"]["total"] == 120
    assert len(first_page["records"]) == 20
    assert first_page["records"][0]["publication_id"] == "human-000"
    def paged_with(**expected):
        return any(all(call.get(key) == value for key, value in expected.items()) for call in page_calls)

    assert paged_with(page=1, page_size=20, sort_by="readiness", sort_order="desc")
    assert sorted_page["records"][0]["year"] == 2000
    assert paged_with(page=1, page_size=20, sort_by="year", sort_order="asc")
    # An unfiltered request must not invent a filter on the way through.
    assert all(call.get("text_filter", "") == "" for call in page_calls)
    assert status == 200
    assert headers["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    workbook = load_workbook(io.BytesIO(xlsx), data_only=False)
    assert workbook.sheetnames == [
        "Query",
        "Publications",
        "Identifiers",
        "Linked datasets",
        "Candidate routes",
        "Species decisions",
        "Provider events",
    ]
    assert workbook["Publications"]["A2"].value == "'=formula risk"
    assert workbook["Provider events"]["D2"].value == "'+neutralize"

    restarted, restarted_thread, restarted_base = _start_server(tmp_path)
    try:
        _, persisted = _request_json(f"{restarted_base}/api/discovery/searches/{created['search_id']}")
        _, persisted_page = _request_json(
            f"{restarted_base}/api/discovery/searches/{created['search_id']}/records?page=6&page_size=20"
        )
    finally:
        _stop_server(restarted, restarted_thread)
    assert persisted["search"]["status"] == "complete"
    assert len(persisted_page["records"]) == 20


def test_federated_search_keeps_human_and_mouse_isolated(tmp_path: Path, monkeypatch) -> None:
    _install_federated_module(monkeypatch, calls=[])
    server, thread, base = _start_server(tmp_path)
    try:
        human = _create_search(base, "human")
        mouse = _create_search(base, "mouse")
        _, human_page = _request_json(f"{base}/api/discovery/searches/{human['search_id']}/records?page=1&page_size=20")
        _, mouse_page = _request_json(f"{base}/api/discovery/searches/{mouse['search_id']}/records?page=1&page_size=20")
    finally:
        _stop_server(server, thread)

    assert human["search_id"] != mouse["search_id"]
    assert human_page["records"][0]["publication_id"].startswith("human-")
    assert mouse_page["records"][0]["publication_id"].startswith("mouse-")
    assert "Search papers and linked repositories" in INDEX_HTML
    assert "Run separate Human + Mouse searches" in INDEX_HTML
    assert "Human and Mouse never pooled" in INDEX_HTML
    assert "Publication + linked public data discovery" in INDEX_HTML
    assert 'sort: { key: "readiness", order: "desc" }' in INDEX_HTML
    assert "Sort: DEG readiness · relevance tie-break" in INDEX_HTML
    assert 'id="downloadSearchExcel"' in INDEX_HTML
    assert '`/api/discovery/searches/${state.searchId}/export.xlsx`' in INDEX_HTML
    assert "complete snapshot" in INDEX_HTML
    assert "source-unit conflict" in INDEX_HTML
    assert 'class="mobile-field-label">DEG readiness' in INDEX_HTML
    assert 'role="tabpanel"' in INDEX_HTML


def test_federated_prepare_uses_persisted_canonical_selection_and_persists_bundle(tmp_path: Path, monkeypatch) -> None:
    _install_federated_module(monkeypatch, calls=[])
    captured: dict = {}
    prepare_module = types.ModuleType("degora.discovery_prepare")

    def fake_prepare(records, species, *, query, materialize_dir, before_publish):
        captured.update(
            records=records,
            species=species,
            query=query,
            materialize_dir=str(materialize_dir),
            before_publish=before_publish,
        )
        return {
            "species": {"key": species},
            "query": query,
            "studies": [],
            "excluded_studies": [],
            "returned_studies": 0,
            "materialize_dir": str(materialize_dir),
            "exports": {},
        }

    prepare_module.prepare_publication_records = fake_prepare
    monkeypatch.setitem(sys.modules, "degora.discovery_prepare", prepare_module)
    server, thread, base = _start_server(tmp_path)
    try:
        created = _create_search(base, "human")
        status, prepared = _request_json(
            f"{base}/api/discovery/prepare",
            payload={
                "species": "human",
                "query": "hypoxia",
                "search_id": created["search_id"],
                "record_ids": ["human-000"],
            },
            action=True,
        )
    finally:
        _stop_server(server, thread)

    assert status == 201
    assert captured["records"][0]["publication_id"] == "human-000"
    assert captured["species"] == "human"
    assert captured["before_publish"] is None
    assert prepared["search_id"] == created["search_id"]

    restarted, restarted_thread, _base = _start_server(tmp_path)
    try:
        persisted = restarted.discovery_search_store.get_artifact("bundle", prepared["bundle_id"])
    finally:
        _stop_server(restarted, restarted_thread)
    assert persisted["bundle_id"] == prepared["bundle_id"]
    assert persisted["search_id"] == created["search_id"]


def test_federated_prepare_rejects_query_relabeling(tmp_path: Path, monkeypatch) -> None:
    _install_federated_module(monkeypatch, calls=[])
    server, thread, base = _start_server(tmp_path)
    try:
        created = _create_search(base, "human")
        request = urllib.request.Request(
            f"{base}/api/discovery/prepare",
            data=json.dumps(
                {
                    "species": "human",
                    "query": "unrelated label",
                    "search_id": created["search_id"],
                    "record_ids": ["human-000"],
                }
            ).encode(),
            headers={"Content-Type": "application/json", "X-DEGORA-Action": "1"},
        )
        try:
            urllib.request.urlopen(request, timeout=10)
        except urllib.error.HTTPError as exc:
            assert exc.code == 400
            payload = json.loads(exc.read().decode())
        else:  # pragma: no cover - the request must be rejected.
            raise AssertionError("query relabeling was accepted")
    finally:
        _stop_server(server, thread)

    assert "does not match" in payload["error"]


def test_legacy_discovery_search_endpoint_remains_synchronous_geo_wrapper(tmp_path: Path, monkeypatch) -> None:
    def fake_search_geo(query, species, **kwargs):
        return {
            "query": query,
            "species": {"key": species},
            "page": kwargs["page"],
            "page_size": kwargs["page_size"],
            "studies": [{"accession": "GSE1", "paper_title": "legacy"}],
        }

    monkeypatch.setattr(discovery, "search_geo", fake_search_geo)
    server, thread, base = _start_server(tmp_path)
    try:
        status, payload = _request_json(f"{base}/api/discovery/search?q=hypoxia&species=human&page=1", action=True)
    finally:
        _stop_server(server, thread)

    assert status == 200
    assert payload["studies"] == [{"accession": "GSE1", "paper_title": "legacy"}]


def test_discovery_workspace_has_a_single_live_server_owner(tmp_path: Path) -> None:
    db = tmp_path / "degora.db"
    db.touch()
    root = tmp_path / "discovery"
    first = create_server(db, port=0, quiet=True, discovery_root=root)
    try:
        with pytest.raises(RuntimeError, match="already using this discovery workspace"):
            create_server(db, port=0, quiet=True, discovery_root=root)
    finally:
        first.server_close()

    restarted = create_server(db, port=0, quiet=True, discovery_root=root)
    restarted.server_close()


def test_internal_type_error_is_not_retried_as_a_second_provider_call(tmp_path: Path, monkeypatch) -> None:
    calls: list[tuple[str, str, int]] = []
    module = types.ModuleType("degora.discovery_federated")

    def search_publications(*, query, species, limit):
        calls.append((query, species, limit))
        raise TypeError("provider implementation bug")

    module.search_publications = search_publications
    monkeypatch.setitem(sys.modules, "degora.discovery_federated", module)
    server, thread, base = _start_server(tmp_path)
    try:
        status, created = _request_json(
            f"{base}/api/discovery/searches",
            payload={"query": "hypoxia", "species": "human", "limit": 10},
            action=True,
        )
        assert status == 202
        deadline = time.time() + 5
        while time.time() < deadline:
            _, payload = _request_json(f"{base}/api/discovery/jobs/{created['job_id']}")
            if payload["job"]["status"] == "failed":
                break
            time.sleep(0.05)
        else:
            raise AssertionError("failing search job did not reach terminal state")
    finally:
        _stop_server(server, thread)

    assert calls == [("hypoxia", "human", 10)]
    assert payload["job"]["error"] == "provider implementation bug"


def test_persistent_discovery_endpoints_require_token_and_action_header(tmp_path: Path, monkeypatch) -> None:
    _install_federated_module(monkeypatch, calls=[])
    db = tmp_path / "degora.db"
    db.touch()
    server = create_server(
        db,
        port=0,
        quiet=True,
        access_token="secret-token",
        discovery_root=tmp_path / "discovery",
    )
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    host, port = server.server_address
    base = f"http://{host}:{port}"
    endpoint = f"{base}/api/discovery/searches"
    try:
        with pytest.raises(urllib.error.HTTPError) as exc_info:
            _request_json(endpoint, payload={"query": "hypoxia", "species": "human", "limit": 20}, action=True)
        assert exc_info.value.code == 401

        with pytest.raises(urllib.error.HTTPError) as exc_info:
            _request_json(
                endpoint,
                payload={"query": "hypoxia", "species": "human", "limit": 20},
                token="secret-token",
            )
        assert exc_info.value.code == 400

        status, created = _request_json(
            endpoint,
            payload={"query": "hypoxia", "species": "human", "limit": 20},
            action=True,
            token="secret-token",
        )
        assert status == 202
        job_url = f"{base}/api/discovery/jobs/{created['job_id']}"
        with pytest.raises(urllib.error.HTTPError) as exc_info:
            _request_json(job_url)
        assert exc_info.value.code == 401
        deadline = time.time() + 5
        while time.time() < deadline:
            _, job = _request_json(job_url, token="secret-token")
            if job["job"]["status"] == "complete":
                break
            time.sleep(0.05)
        else:
            raise AssertionError("authorized discovery search did not complete")
        export_url = f"{base}/api/discovery/searches/{created['search_id']}/export.xlsx"
        with pytest.raises(urllib.error.HTTPError) as exc_info:
            _request_bytes(export_url)
        assert exc_info.value.code == 401
        assert _request_bytes(export_url, token="secret-token")[0] == 200
    finally:
        _stop_server(server, thread)


def _install_progress_reporting_module(monkeypatch, *, seen: list[tuple[float, str]]) -> None:
    """A stub that accepts the optional progress callback, unlike the strict one."""

    module = types.ModuleType("degora.discovery_federated")

    def search_publications(*, query, species, limit, progress=None):
        for fraction, message in ((0.25, "Querying stub (1 of 2 sources)"), (0.75, "Inspecting linked data 5 of 20")):
            if progress is not None:
                progress(fraction, message)
                seen.append((fraction, message))
            time.sleep(0.05)
        return {
            "query": query,
            "species": {"key": species},
            "limit": limit,
            "records": [{"publication_id": "p1", "title": "paper", "year": 2020}],
            "total": 1,
        }

    module.search_publications = search_publications
    module.page_publication_snapshot = lambda snapshot, **kwargs: {
        "records": list(snapshot["records"]),
        "total": 1,
        "page": 1,
        "page_size": 20,
        "total_pages": 1,
        "sort_by": "data_readiness",
        "sort_order": "desc",
        "has_next": False,
    }
    module.resolve_publication_records = lambda records, species: list(records)
    monkeypatch.setitem(sys.modules, "degora.discovery_federated", module)


def test_discovery_job_exposes_progress_and_message(tmp_path: Path, monkeypatch) -> None:
    """The browser draws a determinate bar from these two fields."""

    reported: list[tuple[float, str]] = []
    _install_progress_reporting_module(monkeypatch, seen=reported)
    server, thread, base = _start_server(tmp_path)
    try:
        status, created = _request_json(
            f"{base}/api/discovery/searches",
            payload={"query": "hypoxia", "species": "human", "limit": 20},
            action=True,
        )
        assert status == 202
        observed: list[tuple[float, str]] = []
        deadline = time.time() + 10
        while time.time() < deadline:
            _, payload = _request_json(f"{base}/api/discovery/jobs/{created['job_id']}")
            job = payload["job"]
            assert "progress" in job and "message" in job
            if job["progress"] is not None:
                observed.append((job["progress"], job["message"]))
            if job["status"] == "complete":
                break
            time.sleep(0.02)
        else:  # pragma: no cover - only on a stalled job.
            raise AssertionError("search job did not complete")

        assert reported, "the optional progress callback was not forwarded to search_publications"
        assert observed, "no progress fractions reached the API"
        fractions = [fraction for fraction, _ in observed]
        assert fractions == sorted(fractions), f"job progress must never decrease: {fractions}"
        assert observed[-1][0] == 1.0
        assert observed[-1][1]
    finally:
        _stop_server(server, thread)


def test_discovery_job_progress_is_clamped_and_message_is_bounded(tmp_path: Path, monkeypatch) -> None:
    from degora.api import _api_job_progress, _clean_job_message

    assert _api_job_progress(1.4) == 1.0
    assert _api_job_progress(-2) == 0.0
    assert _api_job_progress(None) is None
    assert _api_job_progress("nope") is None
    assert _api_job_progress(True) is None
    assert _api_job_progress(float("nan")) is None
    assert _clean_job_message("  many\n\nspaces  ") == "many spaces"
    assert _clean_job_message(None) == ""
    assert len(_clean_job_message("x" * 500)) == 160


def test_strict_search_stub_without_progress_still_runs(tmp_path: Path, monkeypatch) -> None:
    """The forwarding guard must not break callers that take no progress kwarg."""

    calls: list[dict] = []
    _install_federated_module(monkeypatch, calls=calls)
    server, thread, base = _start_server(tmp_path)
    try:
        created = _create_search(base, "human")
        _, payload = _request_json(f"{base}/api/discovery/jobs/{created['job_id']}")
        assert payload["job"]["status"] == "complete"
        assert payload["job"]["progress"] == 1.0
    finally:
        _stop_server(server, thread)


def test_prepare_job_route_reports_progress_and_returns_the_bundle(tmp_path: Path, monkeypatch) -> None:
    """Preparation used to be a blocking request with no way to show stages."""

    calls: list[dict] = []
    _install_federated_module(monkeypatch, calls=calls)
    stages: list[tuple[float, str]] = []

    prepare_module = types.ModuleType("degora.discovery_prepare")

    def prepare_publication_records(records, species, *, query, materialize_dir, progress=None, before_publish):
        for fraction, message in ((0.2, "Downloading 1 of 3"), (0.8, "Downloading 3 of 3")):
            if progress is not None:
                progress(fraction, message)
                stages.append((fraction, message))
            time.sleep(0.05)
        Path(materialize_dir).mkdir(parents=True, exist_ok=True)
        before_publish()
        return {"bundle_id": "abc", "studies": [{"accession": "GSE1"}], "excluded_studies": []}

    prepare_module.prepare_publication_records = prepare_publication_records
    monkeypatch.setitem(sys.modules, "degora.discovery_prepare", prepare_module)

    server, thread, base = _start_server(tmp_path)
    events: list[tuple[str, str]] = []
    original_commit = server.discovery_job_manager.commit

    def commit_with_event(job_id):
        events.append(("commit", job_id))
        return original_commit(job_id)

    server.discovery_job_manager.commit = commit_with_event
    original_save_artifact = server.discovery_search_store.save_artifact

    def save_artifact_with_event(kind, artifact_id, payload):
        events.append(("save_artifact", artifact_id))
        return original_save_artifact(kind, artifact_id, payload)

    server.discovery_search_store.save_artifact = save_artifact_with_event
    try:
        created = _create_search(base, "human")
        events.clear()
        _, records_page = _request_json(f"{base}/api/discovery/searches/{created['search_id']}/records?page=1&page_size=20")
        record_ids = [row["publication_id"] for row in records_page["records"][:2]]

        status, started = _request_json(
            f"{base}/api/discovery/prepare-jobs",
            payload={
                "species": "human",
                "query": "hypoxia",
                "search_id": created["search_id"],
                "record_ids": record_ids,
            },
            action=True,
        )
        assert status == 202
        assert started["job_id"]

        observed: list[float] = []
        deadline = time.time() + 15
        while time.time() < deadline:
            _, payload = _request_json(f"{base}/api/discovery/jobs/{started['job_id']}")
            job = payload["job"]
            if job["progress"] is not None:
                observed.append(job["progress"])
            if job["status"] == "complete":
                assert job["result"], "the prepared bundle must reach the browser through the job"
                # The server assigns the bundle id; the worker's value is replaced.
                assert re.fullmatch(r"[a-f0-9]{16}", job["result"]["bundle_id"])
                assert job["result"]["studies"] == [{"accession": "GSE1"}]
                break
            if job["status"] == "failed":
                raise AssertionError(job["error"])
            time.sleep(0.02)
        else:  # pragma: no cover - only on a stalled job.
            raise AssertionError("prepare job did not complete")

        assert stages, "the progress callback was not forwarded to prepare_publication_records"
        assert observed == sorted(observed), f"prepare progress must never decrease: {observed}"
        assert observed[-1] == 1.0
        assert events[0] == ("commit", started["job_id"])
        assert events[1][0] == "save_artifact"
    finally:
        _stop_server(server, thread)


def test_prepare_job_fails_closed_when_implementation_skips_commit_barrier(tmp_path: Path) -> None:
    from degora.api import DegoraRequestHandler
    from degora.discovery_store import DiscoveryJobManager, DiscoveryStateStore

    store = DiscoveryStateStore(tmp_path / "discovery")
    manager = DiscoveryJobManager(store, max_workers=1)
    handler = object.__new__(DegoraRequestHandler)
    handler.server = type("Server", (), {"discovery_job_manager": manager})()
    handler._discovery_prepare = lambda payload, progress, before_publish: {"bundle_id": "unsafe"}

    started = handler._discovery_prepare_job({"species": "human"})
    manager.shutdown(wait=True)

    job = store.get_job(started["job_id"])
    assert job is not None
    assert job["status"] == "failed"
    assert job["result"] is None
    assert "without reaching its commit barrier" in job["error"]["message"]


def test_non_loopback_job_result_redacts_local_paths_without_mutating_store(tmp_path: Path) -> None:
    from degora.api import DegoraRequestHandler
    from degora.discovery_store import DiscoveryStateStore

    store = DiscoveryStateStore(tmp_path / "discovery")
    handler = object.__new__(DegoraRequestHandler)
    handler.server = type(
        "Server",
        (),
        {
            "discovery_search_store": store,
            "server_address": ("0.0.0.0", 0),
        },
    )()
    job = store.create_job("publication_prepare", {"species": "human"})
    local_dir = tmp_path / "discovery" / "human" / "bundles" / "abc"
    local_file = local_dir / "source.csv"
    result = {
        "bundle_id": "abc",
        "materialize_dir": str(local_dir),
        "exports": {"csv": str(local_file)},
        "studies": [{"files": [{"source_path": str(local_file), "source_url": "https://example.org/GSE1"}]}],
    }
    store.update_job(job["job_id"], status="running", progress=0.5)
    store.update_job(job["job_id"], status="completed", result=result)

    response = handler._discovery_job(job["job_id"])
    stored = store.get_job(job["job_id"])

    assert response["result"]["materialize_dir"] == "[redacted: local path]"
    assert response["result"]["exports"]["csv"] == "[redacted: local path]"
    assert response["result"]["studies"][0]["files"][0]["source_path"] == "[redacted: local path]"
    assert response["result"]["studies"][0]["files"][0]["source_url"] == "https://example.org/GSE1"
    assert stored is not None
    assert stored["result"] == result


def test_prepare_implementation_that_ignores_callback_publishes_no_bundle(tmp_path: Path, monkeypatch) -> None:
    _install_federated_module(monkeypatch, calls=[])
    prepare_module = types.ModuleType("degora.discovery_prepare")

    def prepare_publication_records(
        records,
        species,
        *,
        query,
        materialize_dir,
        progress=None,
        before_publish,
    ):
        target = Path(materialize_dir)
        target.mkdir(parents=True, exist_ok=True)
        (target / "unsafe.txt").write_text("published without commit", encoding="utf-8")
        return {"studies": [{"accession": "GSE1"}], "excluded_studies": []}

    prepare_module.prepare_publication_records = prepare_publication_records
    monkeypatch.setitem(sys.modules, "degora.discovery_prepare", prepare_module)

    server, thread, base = _start_server(tmp_path)
    saved_artifacts: list[str] = []
    original_save_artifact = server.discovery_search_store.save_artifact

    def tracking_save_artifact(kind, artifact_id, payload):
        saved_artifacts.append(artifact_id)
        return original_save_artifact(kind, artifact_id, payload)

    server.discovery_search_store.save_artifact = tracking_save_artifact
    try:
        created = _create_search(base, "human")
        _, records_page = _request_json(
            f"{base}/api/discovery/searches/{created['search_id']}/records?page=1&page_size=20"
        )
        status, started = _request_json(
            f"{base}/api/discovery/prepare-jobs",
            payload={
                "species": "human",
                "query": "hypoxia",
                "search_id": created["search_id"],
                "record_ids": [records_page["records"][0]["publication_id"]],
            },
            action=True,
        )
        assert status == 202

        deadline = time.time() + 5
        while time.time() < deadline:
            _, payload = _request_json(f"{base}/api/discovery/jobs/{started['job_id']}")
            job = payload["job"]
            if job["status"] == "failed":
                break
            time.sleep(0.02)
        else:
            raise AssertionError("unsafe prepare implementation did not fail")

        assert "skipped its required commit barrier" in job["error"]
        assert server.discovery_bundles == {}
        assert saved_artifacts == []
        bundle_root = tmp_path / "discovery" / "human" / "bundles"
        assert not bundle_root.exists() or not list(bundle_root.iterdir())
    finally:
        _stop_server(server, thread)


def test_search_cancel_after_commit_keeps_the_snapshot_and_completes(tmp_path: Path, monkeypatch) -> None:
    calls: list[dict] = []
    _install_federated_module(monkeypatch, calls=calls)
    server, thread, base = _start_server(tmp_path)
    saving = threading.Event()
    release_save = threading.Event()
    original_save_search = server.discovery_search_store.save_search

    def blocking_save_search(search_id, payload):
        if isinstance(payload, dict) and payload.get("status") == "complete":
            saving.set()
            assert release_save.wait(timeout=5)
        return original_save_search(search_id, payload)

    server.discovery_search_store.save_search = blocking_save_search
    try:
        status, created = _request_json(
            f"{base}/api/discovery/searches",
            payload={"query": "hypoxia", "species": "human", "limit": 120},
            action=True,
        )
        assert status == 202
        assert saving.wait(timeout=5), "search worker did not reach the committed save"

        status, outcome = _request_json(
            f"{base}/api/discovery/jobs/{created['job_id']}/cancel",
            payload={},
            action=True,
        )
        assert status == 200
        assert outcome["cancelled"] is False
        assert "saving" in outcome["reason"]

        release_save.set()
        deadline = time.time() + 5
        while time.time() < deadline:
            _, payload = _request_json(f"{base}/api/discovery/jobs/{created['job_id']}")
            if payload["job"]["status"] == "complete":
                break
            time.sleep(0.02)
        else:
            raise AssertionError("committed search did not complete")

        _, search = _request_json(f"{base}/api/discovery/searches/{created['search_id']}")
        assert search["search"]["status"] == "complete"
    finally:
        release_save.set()
        _stop_server(server, thread)


@pytest.mark.parametrize(
    "failure",
    [OSError("snapshot fsync failed"), KeyboardInterrupt("snapshot interrupted")],
)
def test_search_snapshot_save_failure_marks_search_failed(tmp_path: Path, monkeypatch, failure) -> None:
    calls: list[dict] = []
    _install_federated_module(monkeypatch, calls=calls)
    server, thread, base = _start_server(tmp_path)
    original_save_search = server.discovery_search_store.save_search
    failed_once = False

    def failing_complete_save(search_id, payload):
        nonlocal failed_once
        if isinstance(payload, dict) and payload.get("status") == "complete" and not failed_once:
            failed_once = True
            raise failure
        return original_save_search(search_id, payload)

    server.discovery_search_store.save_search = failing_complete_save
    try:
        status, created = _request_json(
            f"{base}/api/discovery/searches",
            payload={"query": "hypoxia", "species": "human", "limit": 120},
            action=True,
        )
        assert status == 202
        deadline = time.time() + 5
        while time.time() < deadline:
            _, payload = _request_json(f"{base}/api/discovery/jobs/{created['job_id']}")
            if payload["job"]["status"] == "failed":
                break
            time.sleep(0.02)
        else:
            raise AssertionError("search job did not fail after snapshot save failure")

        _, search = _request_json(f"{base}/api/discovery/searches/{created['search_id']}")
        assert search["search"]["status"] == "failed"
        assert str(failure) in search["search"]["error"]
    finally:
        _stop_server(server, thread)


def test_search_provider_base_exception_marks_search_failed(tmp_path: Path, monkeypatch) -> None:
    calls: list[dict] = []
    _install_federated_module(monkeypatch, calls=calls)

    def interrupted_search_publications(*, query, species, limit, progress=None):
        raise KeyboardInterrupt("provider interrupted")

    sys.modules["degora.discovery_federated"].search_publications = interrupted_search_publications
    server, thread, base = _start_server(tmp_path)
    try:
        status, created = _request_json(
            f"{base}/api/discovery/searches",
            payload={"query": "hypoxia", "species": "human", "limit": 120},
            action=True,
        )
        assert status == 202
        deadline = time.time() + 5
        while time.time() < deadline:
            _, payload = _request_json(f"{base}/api/discovery/jobs/{created['job_id']}")
            if payload["job"]["status"] == "failed":
                assert payload["job"]["error"] == "provider interrupted"
                break
            time.sleep(0.02)
        else:
            raise AssertionError("search job did not fail after provider hard failure")

        _, search = _request_json(f"{base}/api/discovery/searches/{created['search_id']}")
        assert search["search"]["status"] == "failed"
        assert search["search"]["error"] == "provider interrupted"
    finally:
        _stop_server(server, thread)


def test_prepare_cancel_after_commit_keeps_the_bundle_and_completes(tmp_path: Path, monkeypatch) -> None:
    calls: list[dict] = []
    _install_federated_module(monkeypatch, calls=calls)
    committed = threading.Event()
    release_publish = threading.Event()
    prepare_module = types.ModuleType("degora.discovery_prepare")

    def prepare_publication_records(records, species, *, query, materialize_dir, progress=None, before_publish):
        before_publish()
        committed.set()
        assert release_publish.wait(timeout=5)
        Path(materialize_dir).mkdir(parents=True, exist_ok=True)
        return {"studies": [{"accession": "GSE1"}], "excluded_studies": []}

    prepare_module.prepare_publication_records = prepare_publication_records
    monkeypatch.setitem(sys.modules, "degora.discovery_prepare", prepare_module)
    server, thread, base = _start_server(tmp_path)
    try:
        created = _create_search(base, "human")
        _, records_page = _request_json(
            f"{base}/api/discovery/searches/{created['search_id']}/records?page=1&page_size=20"
        )
        record_ids = [row["publication_id"] for row in records_page["records"][:2]]
        status, started = _request_json(
            f"{base}/api/discovery/prepare-jobs",
            payload={
                "species": "human",
                "query": "hypoxia",
                "search_id": created["search_id"],
                "record_ids": record_ids,
            },
            action=True,
        )
        assert status == 202
        assert committed.wait(timeout=5), "preparation did not reach the commit barrier"

        _, outcome = _request_json(
            f"{base}/api/discovery/jobs/{started['job_id']}/cancel",
            payload={},
            action=True,
        )
        assert outcome["cancelled"] is False
        assert "saving" in outcome["reason"]

        release_publish.set()
        deadline = time.time() + 5
        while time.time() < deadline:
            _, payload = _request_json(f"{base}/api/discovery/jobs/{started['job_id']}")
            if payload["job"]["status"] == "complete":
                result = payload["job"]["result"]
                break
            time.sleep(0.02)
        else:
            raise AssertionError("committed preparation did not complete")

        assert result["bundle_id"]
        assert server.discovery_search_store.get_artifact("bundle", result["bundle_id"])["studies"]
    finally:
        release_publish.set()
        _stop_server(server, thread)


def test_complete_search_snapshot_does_not_override_authoritative_cancelled_job(tmp_path: Path) -> None:
    """A side-table snapshot must not rewrite a terminal job state during projection."""

    from degora.api import DegoraRequestHandler

    search_id = "a" * 16
    job_id = "b" * 16

    class Store:
        def get_job(self, wanted):
            assert wanted == job_id
            return {
                "job_id": job_id,
                "kind": "publication_search",
                "status": "cancelled",
                "progress": 0.97,
                "message": "Job was cancelled by the reader.",
                "payload": {"search_id": search_id},
                "result": None,
                "error": None,
                "created_at": "now",
                "updated_at": "now",
            }

        def get_search(self, wanted):
            assert wanted == search_id
            return {
                "id": search_id,
                "query": "hypoxia",
                "species": "human",
                "limit": 20,
                "status": "complete",
                "error": "",
                "snapshot": {"records": [{"publication_id": "p1"}], "total": 1},
                "total": 1,
                "created_at": "now",
                "updated_at": "now",
            }

        def save_search(self, *_args, **_kwargs):
            raise AssertionError("a complete search must not be rewritten as cancelled")

    handler = object.__new__(DegoraRequestHandler)
    handler.server = type("Server", (), {"discovery_search_store": Store(), "discovery_job_manager": None})()

    job = handler._discovery_job(job_id)

    assert job["status"] == "cancelled"
    assert job["result"] is None


def test_duplicate_cancel_reports_already_cancelled_not_finished(tmp_path: Path) -> None:
    from degora.api import DegoraRequestHandler
    from degora.discovery_store import DiscoveryJobManager, DiscoveryStateStore

    store = DiscoveryStateStore(tmp_path / "discovery")
    manager = DiscoveryJobManager(store, max_workers=1)
    handler = object.__new__(DegoraRequestHandler)
    handler.server = type(
        "Server",
        (),
        {
            "discovery_search_store": store,
            "discovery_job_manager": manager,
            "server_address": ("127.0.0.1", 0),
        },
    )()
    job = store.create_job("publication_search", {"search_id": "a" * 16})

    first = handler._discovery_cancel_job(job["job_id"])
    second = handler._discovery_cancel_job(job["job_id"])

    assert first["cancelled"] is True
    assert second["cancelled"] is False
    assert second["job"]["status"] == "cancelled"
    assert "already cancelled" in second["reason"]
    assert "finished" not in second["reason"]
    manager.shutdown(wait=True)


def test_job_result_is_withheld_until_the_job_completes(tmp_path: Path, monkeypatch) -> None:
    calls: list[dict] = []
    _install_federated_module(monkeypatch, calls=calls)
    server, thread, base = _start_server(tmp_path)
    try:
        status, created = _request_json(
            f"{base}/api/discovery/searches",
            payload={"query": "hypoxia", "species": "human", "limit": 20},
            action=True,
        )
        assert status == 202
        _, payload = _request_json(f"{base}/api/discovery/jobs/{created['job_id']}")
        job = payload["job"]
        if job["status"] != "complete":
            assert job["result"] is None
    finally:
        _stop_server(server, thread)


def _install_blocking_federated_module(monkeypatch, *, started, release) -> None:
    """A search that reports progress, then waits until the test lets it go."""

    module = types.ModuleType("degora.discovery_federated")

    def search_publications(*, query, species, limit, progress=None):
        if progress:
            progress(0.1, "Querying public repositories.")
        started.set()
        release.wait(timeout=10)
        if progress:
            progress(0.9, "Ranking candidates.")  # must raise once cancelled
        return {"records": [], "total": 0, "provider_status": "complete", "provider_errors": []}

    module.search_publications = search_publications
    module.page_publication_snapshot = lambda **kwargs: {"records": [], "total": 0}
    module.filter_publication_records = lambda records, text_filter="": list(records)
    monkeypatch.setitem(sys.modules, "degora.discovery_federated", module)


def test_a_running_search_can_be_stopped_from_the_browser(tmp_path: Path, monkeypatch) -> None:
    """The reader pressed stop while the search was querying public repositories."""

    started = threading.Event()
    release = threading.Event()
    _install_blocking_federated_module(monkeypatch, started=started, release=release)
    server, thread, base = _start_server(tmp_path)
    try:
        _, created = _request_json(
            f"{base}/api/discovery/searches",
            payload={"query": "hypoxia", "species": "human", "limit": 20},
            action=True,
        )
        assert started.wait(timeout=10)

        status, cancelled = _request_json(
            f"{base}/api/discovery/jobs/{created['job_id']}/cancel", payload={}, action=True
        )
        release.set()

        assert status == 200
        assert cancelled["cancelled"] is True
        assert cancelled["job"]["status"] == "cancelled"
        # The reader is told what stopping did and did not undo.
        assert "downloaded" in cancelled["reason"]

        # The job stays cancelled, and never acquires a result.
        deadline = time.time() + 5
        while time.time() < deadline:
            _, polled = _request_json(f"{base}/api/discovery/jobs/{created['job_id']}")
            if polled["job"]["result"] is not None:
                raise AssertionError("a cancelled job recorded a result")
            if polled["job"]["status"] != "cancelled":
                raise AssertionError(f"status drifted to {polled['job']['status']}")
            time.sleep(0.05)

        # And the snapshot record does not claim to still be queued.
        _, search = _request_json(f"{base}/api/discovery/searches/{created['search_id']}")
        assert search["search"]["status"] == "cancelled"
    finally:
        release.set()
        _stop_server(server, thread)


def test_stopping_a_finished_search_says_it_was_too_late(tmp_path: Path, monkeypatch) -> None:
    page_calls: list[dict] = []
    _install_federated_module(monkeypatch, calls=page_calls)
    server, thread, base = _start_server(tmp_path)
    try:
        created = _create_search(base, "human")
        status, outcome = _request_json(
            f"{base}/api/discovery/jobs/{created['job_id']}/cancel", payload={}, action=True
        )

        assert status == 200
        assert outcome["cancelled"] is False
        assert "already finished" in outcome["reason"]
        # The completed result is untouched: the reader can still read it.
        assert outcome["job"]["status"] == "complete"
        _, search = _request_json(f"{base}/api/discovery/searches/{created['search_id']}")
        assert search["search"]["status"] == "complete"
    finally:
        _stop_server(server, thread)


def test_cancelling_an_unknown_job_is_a_not_found(tmp_path: Path, monkeypatch) -> None:
    page_calls: list[dict] = []
    _install_federated_module(monkeypatch, calls=page_calls)
    server, thread, base = _start_server(tmp_path)
    try:
        with pytest.raises(urllib.error.HTTPError) as excinfo:
            _request_json(f"{base}/api/discovery/jobs/{'0' * 16}/cancel", payload={}, action=True)
        assert excinfo.value.code == 404
    finally:
        _stop_server(server, thread)


def test_the_cancel_route_requires_an_action_header(tmp_path: Path, monkeypatch) -> None:
    """Cancelling changes server state, so it carries the same guard as the rest."""

    page_calls: list[dict] = []
    _install_federated_module(monkeypatch, calls=page_calls)
    server, thread, base = _start_server(tmp_path)
    try:
        created = _create_search(base, "human")
        with pytest.raises(urllib.error.HTTPError) as excinfo:
            _request_json(f"{base}/api/discovery/jobs/{created['job_id']}/cancel", payload={})
        assert excinfo.value.code in {400, 401, 403}
    finally:
        _stop_server(server, thread)
