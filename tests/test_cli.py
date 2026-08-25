from __future__ import annotations

from decimal import Decimal
import json
import sqlite3
from zipfile import ZipFile

import pandas as pd
import pytest
from openpyxl import load_workbook

from degora import SCORE_VERSION, __version__
from degora.cli import DISCOVERY_PAGE_SIZE, _int_setting, _print_run_warnings, main
from degora.excel_template import TEMPLATE_SHEETS
from degora.slice_runner import DegoraConfigError


def test_legacy_geo_discover_exports_species_scoped_result_page(tmp_path, monkeypatch, capsys) -> None:
    import degora.discovery as discovery

    captured = {}

    def fake_search(query, species, *, page, page_size, assess_files, global_rank, global_limit):
        captured.update(
            query=query,
            species=species,
            page=page,
            page_size=page_size,
            assess_files=assess_files,
            global_rank=global_rank,
            global_limit=global_limit,
        )
        return {
            "query": query,
            "species": {"key": species},
            "page": page,
            "page_size": page_size,
            "total_pages": 3,
            "total_hits": 42,
            "evaluated_studies": 42,
            "studies": [
                {
                    "species": species,
                    "accession": "GSE1",
                    "paper_title": "Paper",
                    "authors": ["Kim K"],
                    "pubmed_ids": ["1"],
                }
            ],
        }

    monkeypatch.setattr(discovery, "search_geo", fake_search)
    output = tmp_path / "search"
    assert main(
        [
            "discover",
            "hypoxia",
            "--source",
            "geo",
            "--species",
            "human",
            "--page",
            "2",
            "--output-dir",
            str(output),
        ]
    ) == 0
    assert captured == {
        "query": "hypoxia",
        "species": "human",
        "page": 2,
        "page_size": 10,
        "assess_files": True,
        "global_rank": True,
        "global_limit": 1000,
    }
    assert (output / "geo_search_page.json").is_file()
    assert (output / "geo_search_page.csv").is_file()
    assert "globally ranked exact human page 2" in capsys.readouterr().out


def test_discover_command_rejects_both_species_at_parser_boundary(tmp_path) -> None:
    with pytest.raises(SystemExit) as exc_info:
        main(["discover", "hypoxia", "--species", "both", "--output-dir", str(tmp_path)])
    assert exc_info.value.code == 2


def _write_source(path, genes, lfc_scale: float) -> None:
    pd.DataFrame(
        {
            "gene": genes,
            "log2FoldChange": [2.0 * lfc_scale, 1.5 * lfc_scale, 0.1],
            "pvalue": [1e-6, 1e-4, 0.8],
            "padj": [1e-5, 1e-3, 0.9],
        }
    ).to_csv(path, index=False)


def _write_config(path, source_a, source_b) -> None:
    project = pd.DataFrame(
        {
            "field": ["output_dir", "harmonized_dir", "min_studies"],
            "value": [str(path.parent / "results"), str(path.parent / "harmonized"), 1],
        }
    )
    contrasts = pd.DataFrame(
        {
            "study_id": ["S1_4h", "S2_4h"],
            "source_unit_id": ["P1", "P2"],
            "source_path": [str(source_a), str(source_b)],
            "gene_column": ["gene", "gene"],
            "lfc_column": ["log2FoldChange", "log2FoldChange"],
            "p_column": ["pvalue", "pvalue"],
            "padj_column": ["padj", "padj"],
            "include": ["yes", "yes"],
        }
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        project.to_excel(writer, sheet_name="Project", index=False)
        contrasts.to_excel(writer, sheet_name="Contrasts", index=False)


def test_template_command_writes_beginner_workbook(tmp_path) -> None:
    output = tmp_path / "DEGORA_template.xlsx"

    assert main(["template", str(output)]) == 0

    workbook = load_workbook(output)
    assert workbook.sheetnames == TEMPLATE_SHEETS
    assert workbook["Contrasts"]["A1"].value.startswith("#")
    assert workbook["Contrasts"]["A2"].value == "study_id"
    assert workbook["ColumnGuide"]["A1"].value.startswith("#")
    assert workbook["ColumnGuide"]["A2"].value == "column"
    assert workbook["ColumnGuide"]["C2"].value == "checked_where"
    assert workbook["Contrasts"].freeze_panes == "A3"
    contrast_headers = [cell.value for cell in workbook["Contrasts"][2]]
    guide_columns = [cell.value for cell in workbook["ColumnGuide"]["A"]]
    guide_required = {row[0].value: row[1].value for row in workbook["ColumnGuide"].iter_rows(min_row=3)}
    gold_headers = [cell.value for cell in workbook["GoldPanel"][2]]
    locked_column = gold_headers.index("locked") + 1
    assert workbook["GoldPanel"].cell(row=3, column=locked_column).value == "no"
    assert "time_course_mode" in contrast_headers
    assert "time_course_mode" in guide_columns
    assert {"n_ctrl", "n_treat", "source_url"}.issubset(contrast_headers)
    assert {"n_ctrl", "n_treat", "source_url"}.issubset(guide_columns)
    contrast_column_index = {name: index + 1 for index, name in enumerate(contrast_headers)}
    for row in range(3, 6):
        assert workbook["Contrasts"].cell(row=row, column=contrast_column_index["n_ctrl"]).value is None
        assert workbook["Contrasts"].cell(row=row, column=contrast_column_index["n_treat"]).value is None
    assert guide_required["p_column"] == "yes"
    assert guide_required["padj_column"] == "no; checked if filled"
    advanced = {
        row[0]: row[1]
        for row in workbook["AdvancedSettings"].iter_rows(min_row=3, values_only=True)
        if row[0]
    }
    assert advanced["score_version"] == SCORE_VERSION


def test_version_option_prints_package_version(capsys) -> None:
    with pytest.raises(SystemExit) as exc_info:
        main(["--version"])

    assert exc_info.value.code == 0
    assert f"DEGORA {__version__}" in capsys.readouterr().out


@pytest.mark.parametrize("bad_value", ["2.5", 2.5, Decimal("2.5"), "nan", Decimal("NaN"), True])
def test_min_studies_setting_rejects_non_integer_values(bad_value) -> None:
    with pytest.raises(DegoraConfigError, match="numeric setting is invalid"):
        _int_setting(bad_value, 2)


@pytest.mark.parametrize("value", ["2", "2.0", 2, 2.0, Decimal("2")])
def test_min_studies_setting_accepts_exact_integer_values(value) -> None:
    assert _int_setting(value, 1) == 2


def test_demo_command_writes_runnable_workspace(tmp_path, capsys) -> None:
    demo_dir = tmp_path / "demo"

    assert main(["demo", str(demo_dir)]) == 0

    config = demo_dir / "degora_demo_config.xlsx"
    assert config.exists()
    assert (demo_dir / "deg_tables" / "demo_ifn_a_4h.csv").exists()
    assert (demo_dir / "README.md").exists()
    demo_config = load_workbook(config)
    advanced = {
        row[0]: row[1]
        for row in demo_config["AdvancedSettings"].iter_rows(min_row=3, values_only=True)
        if row[0]
    }
    assert advanced["score_version"] == SCORE_VERSION
    project = {
        row[0]: row[1]
        for row in demo_config["Project"].iter_rows(min_row=3, values_only=True)
        if row[0]
    }
    assert project["demo_search_keyword"] == "hypoxia normoxia renal epithelial"
    assert project["demo_search_species"] == "human"
    assert main(["validate", str(config)]) == 0
    assert main(["run", str(config)]) == 0
    captured = capsys.readouterr()
    assert f"- DEGORA version: {__version__}" in captured.out
    assert "non-fatal input warnings" not in captured.err

    db = demo_dir / "results" / "degora_scores.db"
    assert db.exists()
    workbook_path = demo_dir / "results" / "DEGORA_output.xlsx"
    assert workbook_path.exists()
    workbook = load_workbook(workbook_path)
    assert workbook.sheetnames[:2] == ["Workbook_guide", "Column_dictionary"]
    assert {"Run_summary", "Gene_scores", "Gene_evidence", "Source_units"}.issubset(set(workbook.sheetnames))
    guide_rows = list(workbook["Workbook_guide"].iter_rows(min_row=2, values_only=True))
    assert any(row[0] == "Gene_scores" and "main prioritized gene list" in row[3] for row in guide_rows)
    summary_rows = {row[0]: row[1] for row in workbook["Run_summary"].iter_rows(min_row=2, values_only=True)}
    assert summary_rows["degora_version"] == __version__
    dictionary_rows = list(workbook["Column_dictionary"].iter_rows(min_row=2, values_only=True))
    assert any(
        row[0] == "Gene_scores"
        and row[2] == "quality_weighted_degora_score"
        and "relative index, not a probability" in row[3]
        for row in dictionary_rows
    )
    assert any(
        row[0] == "Gene_scores"
        and row[2] == "direction_posterior_mean"
        and "not a calibrated posterior probability" in row[3]
        for row in dictionary_rows
    )
    assert any(
        row[0] == "Gene_scores"
        and row[2] == "evidence_reliability_score"
        and row[4] == "0-100, higher is stronger"
        and "evidence_reliability_components_used" in row[5]
        for row in dictionary_rows
    )
    assert any(
        row[0] == "Gene_scores"
        and row[2] == "evidence_reliability_components_used"
        and "LOO was unavailable" in row[5]
        for row in dictionary_rows
    )
    assert workbook["Gene_scores"].freeze_panes == "A2"
    gene_score_headers = {cell.value: cell for cell in workbook["Gene_scores"][1]}
    assert gene_score_headers["quality_weighted_degora_score"].comment is not None
    assert "Values:" in gene_score_headers["quality_weighted_degora_score"].comment.text
    assert workbook["Gene_scores"].cell(
        row=2,
        column=gene_score_headers["loo_rank_stability_score"].column,
    ).value is None
    assert workbook["Gene_scores"].cell(
        row=2,
        column=gene_score_headers["evidence_reliability_components_used"].column,
    ).value == 3
    assert workbook["Column_dictionary"]["A1"].comment is not None
    metadata = json.loads((demo_dir / "results" / "degora_score_metadata.json").read_text())
    summary = json.loads((demo_dir / "results" / "degora_score_db_summary.json").read_text())
    assert metadata["degora_version"] == __version__
    assert summary["degora_version"] == __version__
    with sqlite3.connect(db) as connection:
        top_genes = [
            row[0]
            for row in connection.execute("SELECT gene_symbol FROM genes ORDER BY degora_rank LIMIT 4").fetchall()
        ]
        source_units = connection.execute("SELECT COUNT(DISTINCT source_unit_id) FROM studies").fetchone()[0]
        db_meta = dict(
            connection.execute(
                "SELECT key, value FROM meta WHERE key IN ('demo_search_keyword', 'demo_search_species')"
            ).fetchall()
        )
    assert top_genes[0] == "ISG15"
    assert {"IFIT1", "MX1"}.issubset(set(top_genes))
    assert source_units == 2
    assert db_meta == {
        "demo_search_keyword": "hypoxia normoxia renal epithelial",
        "demo_search_species": "human",
    }


def test_demo_command_accepts_custom_live_search_seed(tmp_path) -> None:
    demo_dir = tmp_path / "mouse-demo"
    assert main(["demo", str(demo_dir), "--keyword", "liver fibrosis", "--species", "mouse"]) == 0
    workbook = load_workbook(demo_dir / "degora_demo_config.xlsx")
    project = {
        row[0]: row[1]
        for row in workbook["Project"].iter_rows(min_row=3, values_only=True)
        if row[0]
    }
    assert project["demo_search_keyword"] == "liver fibrosis"
    assert project["demo_search_species"] == "mouse"
    readme = (demo_dir / "README.md").read_text()
    assert "Mouse tab is suggested" in readme


def test_validate_command_accepts_excel_config(tmp_path, capsys) -> None:
    source = tmp_path / "source.csv"
    _write_source(source, ["ISG15", "IFIT1", "RPL13A"], 1.0)
    config = tmp_path / "config.xlsx"
    _write_config(config, source, source)

    assert main(["validate", str(config)]) == 0
    captured = capsys.readouterr()
    assert "Required Contrasts columns" in captured.out
    assert "source_unit_id (or paper_id)" in captured.out
    assert "Required DEG-table mappings" in captured.out
    assert "Optional DEG-table mappings checked when filled" in captured.out
    assert "padj_column -> adjusted p-value/FDR" in captured.out


def test_run_command_builds_score_database_from_excel_config(tmp_path) -> None:
    source_a = tmp_path / "source_a.csv"
    source_b = tmp_path / "source_b.csv"
    _write_source(source_a, ["ISG15", "IFIT1", "RPL13A"], 1.0)
    _write_source(source_b, ["ISG15", "IFIT1", "RPL13A"], 0.8)
    config = tmp_path / "config.xlsx"
    _write_config(config, source_a, source_b)
    db = tmp_path / "results" / "degora_scores.db"

    assert main(["run", str(config), "--db", str(db)]) == 0

    assert db.exists()
    assert (tmp_path / "results" / "degora_gene_scores.csv").exists()
    assert (tmp_path / "results" / "degora_score_metadata.json").exists()
    assert (tmp_path / "results" / "DEGORA_output.xlsx").exists()
    metadata = json.loads((tmp_path / "results" / "degora_score_metadata.json").read_text())
    assert metadata["degora_version"] == __version__
    with sqlite3.connect(db) as connection:
        top_gene = connection.execute("SELECT gene_symbol FROM genes ORDER BY degora_rank LIMIT 1").fetchone()[0]
        source_units = connection.execute("SELECT COUNT(DISTINCT source_unit_id) FROM studies").fetchone()[0]
        sqlite_meta = dict(connection.execute("SELECT key, value FROM meta").fetchall())
    assert top_gene == "ISG15"
    assert source_units == 2
    assert sqlite_meta["degora_version"] == __version__


def test_run_command_can_skip_default_excel_export(tmp_path) -> None:
    source_a = tmp_path / "source_a.csv"
    source_b = tmp_path / "source_b.csv"
    _write_source(source_a, ["ISG15", "IFIT1", "RPL13A"], 1.0)
    _write_source(source_b, ["ISG15", "IFIT1", "RPL13A"], 0.8)
    config = tmp_path / "config.xlsx"
    _write_config(config, source_a, source_b)

    assert main(["run", str(config), "--no-excel"]) == 0

    assert (tmp_path / "results" / "degora_scores.db").exists()
    assert not (tmp_path / "results" / "DEGORA_output.xlsx").exists()


def test_run_returns_clean_failure_when_no_gene_meets_minimum_sources(tmp_path, capsys) -> None:
    source_a = tmp_path / "source_a.csv"
    source_b = tmp_path / "source_b.csv"
    _write_source(source_a, ["A1", "A2", "A3"], 1.0)
    _write_source(source_b, ["B1", "B2", "B3"], 1.0)
    config = tmp_path / "config.xlsx"
    _write_config(config, source_a, source_b)

    assert main(["run", str(config), "--min-studies", "2"]) == 2
    message = capsys.readouterr().err
    assert "zero genes" in message
    assert "min_studies=2" in message


def test_run_fails_before_writing_artifacts_when_source_units_cannot_satisfy_min_studies(tmp_path, capsys) -> None:
    source = tmp_path / "one_source.csv"
    _write_source(source, ["ISG15", "IFIT1", "RPL13A"], 1.0)
    config = tmp_path / "one_source_config.csv"
    pd.DataFrame(
        [
            {
                "study_id": "S1_4h",
                "source_unit_id": "P1",
                "source_path": str(source),
                "gene_column": "gene",
                "lfc_column": "log2FoldChange",
                "p_column": "pvalue",
                "padj_column": "padj",
            }
        ]
    ).to_csv(config, index=False)
    output = tmp_path / "early_results"
    harmonized = tmp_path / "early_harmonized"
    db = tmp_path / "early_results" / "degora_scores.db"

    assert (
        main(
            [
                "run",
                str(config),
                "--min-studies",
                "2",
                "--output-dir",
                str(output),
                "--harmonized-dir",
                str(harmonized),
                "--db",
                str(db),
            ]
        )
        == 2
    )

    message = capsys.readouterr().err
    assert "cannot run at min_studies=2" in message
    assert "only 1 independent source unit" in message
    assert not db.exists()
    assert not (output / "slice_harmonized.csv").exists()
    assert not (output / "degora_gene_scores.csv").exists()
    assert not any(harmonized.glob("*.csv")) if harmonized.exists() else True


def test_run_accepts_source_unit_count_equal_to_min_studies(tmp_path) -> None:
    source_a = tmp_path / "source_a.csv"
    source_b = tmp_path / "source_b.csv"
    _write_source(source_a, ["ISG15", "IFIT1", "RPL13A"], 1.0)
    _write_source(source_b, ["ISG15", "IFIT1", "RPL13A"], 0.8)
    config = tmp_path / "config.xlsx"
    _write_config(config, source_a, source_b)

    assert main(["run", str(config), "--min-studies", "2", "--no-excel"]) == 0
    assert (tmp_path / "results" / "degora_scores.db").is_file()


def test_init_returns_clean_error_when_no_table_is_confirmed(tmp_path, monkeypatch, capsys) -> None:
    source = tmp_path / "source.csv"
    _write_source(source, ["ISG15", "IFIT1", "RPL13A"], 1.0)
    output = tmp_path / "config.csv"

    monkeypatch.setattr("builtins.input", lambda _prompt: (_ for _ in ()).throw(EOFError))

    assert main(["init", str(output), "--deg-dir", str(tmp_path)]) == 2

    message = capsys.readouterr().err
    assert "no table was confirmed" in message
    assert not output.exists()


def test_init_cli_cannot_publish_rank_values_as_gene_symbols(tmp_path, monkeypatch, capsys) -> None:
    deg_dir = tmp_path / "deg"
    deg_dir.mkdir()
    for index in (1, 2):
        pd.DataFrame(
            {
                "rank": list(range(1, 121)),
                "log2FoldChange": [1.2 - row / 100 for row in range(120)],
                "pvalue": [0.001 + row / 10000 for row in range(120)],
            }
        ).to_csv(deg_dir / f"rank_results_{index}.csv", index=False)
    output = tmp_path / "config.csv"

    def answer(prompt: str) -> str:
        if "Which species" in prompt:
            return "human"
        return pytest.fail(f"non-gene tables must be skipped before prompting: {prompt}")

    monkeypatch.setattr("builtins.input", answer)

    assert main(["init", str(output), "--deg-dir", str(deg_dir)]) == 2
    assert "no table was confirmed" in capsys.readouterr().err
    assert not output.exists()


def test_init_does_not_hide_unexpected_value_errors(monkeypatch, tmp_path) -> None:
    import degora.beginner as beginner

    monkeypatch.setattr(beginner, "run_init", lambda *args, **kwargs: (_ for _ in ()).throw(ValueError("bug")))

    with pytest.raises(ValueError, match="bug"):
        main(["init", str(tmp_path / "config.csv"), "--deg-dir", str(tmp_path)])


def test_run_returns_clean_failure_when_default_excel_export_fails(tmp_path, monkeypatch, capsys) -> None:
    import degora.excel_export as excel_export

    source_a = tmp_path / "source_a.csv"
    source_b = tmp_path / "source_b.csv"
    _write_source(source_a, ["ISG15", "IFIT1", "RPL13A"], 1.0)
    _write_source(source_b, ["ISG15", "IFIT1", "RPL13A"], 0.8)
    config = tmp_path / "config.xlsx"
    _write_config(config, source_a, source_b)
    monkeypatch.setattr(excel_export, "export_run_workbook", lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("synthetic export failure")))

    assert main(["run", str(config)]) == 2
    message = capsys.readouterr().err
    assert "Excel workbook export failed" in message
    assert "--no-excel" in message


def test_run_warns_when_gold_panel_cannot_support_recall_metrics(tmp_path, capsys) -> None:
    source_a = tmp_path / "source_a.csv"
    source_b = tmp_path / "source_b.csv"
    _write_source(source_a, ["ISG15", "IFIT1", "RPL13A"], 1.0)
    _write_source(source_b, ["ISG15", "IFIT1", "RPL13A"], 0.8)
    config = tmp_path / "config.xlsx"
    _write_config(config, source_a, source_b)
    with pd.ExcelWriter(config, mode="a", engine="openpyxl") as writer:
        pd.DataFrame({"marker": ["ISG15"]}).to_excel(writer, sheet_name="GoldPanel", index=False)

    assert main(["run", str(config)]) == 0

    message = capsys.readouterr().err
    assert "GoldPanel is missing the required gene_symbol column" in message
    assert "curated recall was not calculated" in message
    assert "DEGORA_output.manifest.json" in message
    manifest = json.loads((tmp_path / "results" / "DEGORA_output.manifest.json").read_text())
    assert manifest["gold_panel"]["status"] == "invalid"


def test_validate_missing_config_returns_clean_error(tmp_path, capsys) -> None:
    exit_code = main(["validate", str(tmp_path / "does_not_exist.xlsx")])

    # Beginner-facing error contract: a clear message and exit code 2, not a raw traceback.
    assert exit_code == 2
    assert "config file was not found" in capsys.readouterr().err


def test_serve_duplicate_workspace_returns_clean_error(tmp_path, monkeypatch, capsys) -> None:
    import degora.api as api

    db = tmp_path / "degora_scores.db"
    db.write_bytes(b"placeholder")
    monkeypatch.setattr(
        api,
        "serve",
        lambda *args, **kwargs: (_ for _ in ()).throw(
            api.DiscoveryWorkspaceInUseError("another DEGORA server is already using this discovery workspace")
        ),
    )

    assert main(["serve", str(db)]) == 2
    assert "already using this discovery workspace" in capsys.readouterr().err


def test_validate_rejects_unsupported_score_version(tmp_path, capsys) -> None:
    source = tmp_path / "source.csv"
    _write_source(source, ["ISG15", "IFIT1", "RPL13A"], 1.0)
    config = tmp_path / "config.xlsx"
    _write_config(config, source, source)
    with pd.ExcelWriter(config, mode="a", engine="openpyxl") as writer:
        pd.DataFrame(
            {"setting": ["score_version"], "value": ["degora_score_v1"]}
        ).to_excel(writer, sheet_name="AdvancedSettings", index=False)

    assert main(["validate", str(config)]) == 2
    message = capsys.readouterr().err
    assert "score_version is not supported" in message
    assert SCORE_VERSION in message


def test_validate_checks_score_version_before_source_files(tmp_path, capsys) -> None:
    missing = tmp_path / "missing.csv"
    config = tmp_path / "config.xlsx"
    _write_config(config, missing, missing)
    with pd.ExcelWriter(config, mode="a", engine="openpyxl") as writer:
        pd.DataFrame(
            {"setting": ["score_version"], "value": ["degora_score_v1"]}
        ).to_excel(writer, sheet_name="AdvancedSettings", index=False)

    assert main(["validate", str(config)]) == 2
    message = capsys.readouterr().err
    assert "score_version is not supported" in message
    assert "source DEG table file was not found" not in message


def test_run_warnings_are_printed_to_stderr(tmp_path, capsys) -> None:
    metrics_path = tmp_path / "slice_metrics.json"

    _print_run_warnings(
        {
            "warnings": ["S1 produced zero harmonized rows", "S1 produced zero harmonized rows"],
            "rank_universe_warnings": ["DEG-only table without rank_universe_size"],
            "pvalue_clipped_rows": 2,
        },
        metrics_path=metrics_path,
    )

    captured = capsys.readouterr()
    assert captured.out == ""
    assert "DEGORA completed with non-fatal input warnings" in captured.err
    assert captured.err.count("S1 produced zero harmonized rows") == 1
    assert "DEG-only table without rank_universe_size" in captured.err
    assert "2 row(s) reported pvalue < 1e-300" in captured.err
    assert str(metrics_path) in captured.err


def test_run_reports_each_phase_with_elapsed_time(tmp_path, capsys, monkeypatch) -> None:
    """A five-source corpus takes minutes; the command used to print nothing."""

    from degora.cli import _RunProgress

    progress = _RunProgress()
    progress.start("Harmonizing source tables")
    progress.done("16,852 harmonized rows")
    out = capsys.readouterr().out
    assert "Harmonizing source tables..." in out
    assert "done in" in out
    assert "16,852 harmonized rows" in out
    # Every line carries the elapsed time so a slow phase is identifiable.
    assert all(line.startswith("[") for line in out.splitlines() if line.strip())


def test_run_progress_can_be_silenced() -> None:
    from degora.cli import _RunProgress

    progress = _RunProgress(enabled=False)
    progress.start("phase")
    progress.done("detail")  # must not raise


def test_validate_reports_row_progress_for_large_catalogs(tmp_path, monkeypatch, capsys) -> None:
    import degora.slice_runner as slice_runner

    config = tmp_path / "config.csv"
    config.write_text("placeholder\n", encoding="utf-8")

    def fake_validate(path, *, progress=None):
        if progress is not None:
            progress(0, 50_000, "")
            progress(1_000, 50_000, "S1000")
            progress(50_000, 50_000, "S50000")
        return {
            "active_contrasts": 50_000,
            "excluded_contrasts": 0,
            "source_units": 50_000,
            "required_contrasts_columns": [],
            "required_source_table_mappings": [],
            "optional_source_table_mappings": [],
            "warnings": [],
        }

    monkeypatch.setattr(slice_runner, "validate_catalog_inputs", fake_validate)

    assert main(["validate", str(config)]) == 0
    out = capsys.readouterr().out
    assert "checked 0/50,000 active row(s)" in out
    assert "checked 50,000/50,000 active row(s) (S50000)" in out
    assert "DEGORA config OK" in out


def test_zero_gene_error_names_the_identifier_mismatch(tmp_path) -> None:
    """Two sources that share no identifiers is the commonest cause by far."""

    import pandas as pd

    from degora.cli import _zero_gene_diagnostic

    harmonized = tmp_path / "slice_harmonized.csv"
    pd.DataFrame(
        {
            "gene_symbol": ["YAL001C", "YAL002W", "GENE-A1CF|A1CF", "GENE-A2M|A2M"],
            "source_unit_id": ["U_YEAST", "U_YEAST", "U_COD", "U_COD"],
        }
    ).to_csv(harmonized, index=False)

    message = _zero_gene_diagnostic(harmonized, min_studies=2)

    assert "No identifier is shared by all 2 source units" in message
    assert "U_YEAST (2)" in message and "U_COD (2)" in message
    assert "same identifier space" in message


def test_zero_gene_error_reports_too_few_contributing_units(tmp_path) -> None:
    import pandas as pd

    from degora.cli import _zero_gene_diagnostic

    harmonized = tmp_path / "slice_harmonized.csv"
    pd.DataFrame({"gene_symbol": ["TP53"], "source_unit_id": ["ONLY"]}).to_csv(harmonized, index=False)

    message = _zero_gene_diagnostic(harmonized, min_studies=2)
    assert "Only 1 source unit(s)" in message
    assert "min_studies" in message


def test_zero_gene_error_says_when_overlap_is_not_the_problem(tmp_path) -> None:
    import pandas as pd

    from degora.cli import _zero_gene_diagnostic

    harmonized = tmp_path / "slice_harmonized.csv"
    pd.DataFrame(
        {"gene_symbol": ["TP53", "TP53", "MYC", "MYC"], "source_unit_id": ["A", "B", "A", "B"]}
    ).to_csv(harmonized, index=False)

    message = _zero_gene_diagnostic(harmonized, min_studies=2)
    assert "overlap is not the problem" in message
    assert "contrast direction" in message


def test_zero_gene_diagnostic_never_masks_the_real_error(tmp_path) -> None:
    from degora.cli import _zero_gene_diagnostic

    assert _zero_gene_diagnostic(tmp_path / "does-not-exist.csv", min_studies=2) == ""


def test_discover_help_states_the_real_page_size(capsys) -> None:
    """Help text must be derived from DISCOVERY_PAGE_SIZE, not restated by hand.

    The constant and the prose disagreed once already; deriving both from the same
    value is what keeps a documented page size from drifting away from the code.
    """

    with pytest.raises(SystemExit):
        main(["discover", "--help"])

    # argparse hard-wraps help text, so compare on collapsed whitespace.
    text = " ".join(capsys.readouterr().out.split())
    assert f"{DISCOVERY_PAGE_SIZE} rows per globally sorted page" in text
    assert f"{DISCOVERY_PAGE_SIZE} rows per page" in text
    assert "20 rows per" not in text


def test_malformed_workbook_config_reports_a_config_error(tmp_path, capsys) -> None:
    """A .xlsx that is a valid ZIP but not a workbook must not surface a traceback.

    The optional Project/AdvancedSettings sheets are read before the catalog, and
    pandas raises engine-selection errors there that are neither ValueError nor a
    DEGORA error, so they escaped the beginner-readable error contract.
    """

    config = tmp_path / "broken.xlsx"
    with ZipFile(config, "w") as archive:
        archive.writestr("hello.txt", "not a workbook")

    assert main(["validate", str(config)]) == 2

    message = capsys.readouterr().err
    assert "DEGORA config error: config file could not be read" in message
    assert "valid CSV or Excel (.xlsx) workbook" in message


def test_validate_says_a_config_cannot_meet_the_replication_rule(tmp_path, capsys) -> None:
    """One source unit against min_studies=2 scores zero genes, guaranteed.

    validate reported "DEGORA config OK" for such a config, and the fact only
    surfaced after a full run. Both numbers are already known at validate time.
    """

    import pandas as pd

    from degora.cli import main

    deg = tmp_path / "one_study.csv"
    pd.DataFrame(
        {
            "gene": [f"G{index}" for index in range(120)],
            "log2FoldChange": [2.0 - index * 0.01 for index in range(120)],
            "pvalue": [1e-5 if index < 40 else 0.5 for index in range(120)],
        }
    ).to_csv(deg, index=False)

    config = tmp_path / "config.csv"
    pd.DataFrame(
        [
            {
                "study_id": "S1",
                "source_unit_id": "P1",
                "source_path": deg.name,
                "gene_column": "gene",
                "lfc_column": "log2FoldChange",
                "p_column": "pvalue",
                "include_in_analysis": "yes",
            }
        ]
    ).to_csv(config, index=False)

    assert main(["validate", str(config)]) == 0

    captured = capsys.readouterr()
    assert "DEGORA config OK" in captured.out
    assert "1 independent source unit(s) but min_studies is 2" in captured.err
    assert "scores zero genes" in captured.err
