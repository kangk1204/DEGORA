from __future__ import annotations

import json
import sqlite3
from pathlib import Path
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook

from degora.excel_export import (
    COLUMN_DEFINITIONS,
    _cap_evidence_for_sheet,
    _evidence_row_cap,
    _force_formula_like_text,
    export_run_workbook,
)
from degora.score_db import HETEROGENEITY_RULE, RANDOM_EFFECTS_STOUFFER_RULE


def test_evidence_sheet_cap_keeps_top_ranked_genes_and_flags() -> None:
    # 60 evidence rows across 6 genes; cap to 20 must keep the lowest-rank (best) genes' evidence.
    evidence = pd.DataFrame(
        {
            "gene_symbol": [f"G{i % 6}" for i in range(60)],
            "study_id": [f"S{i}" for i in range(60)],
        }
    )
    genes = pd.DataFrame(
        {
            "gene_symbol": [f"G{i}" for i in range(6)],
            "quality_weighted_degora_rank": [3, 1, 5, 2, 6, 4],
        }
    )
    capped, was_capped = _cap_evidence_for_sheet(evidence, genes, cap=20)
    assert was_capped is True
    assert len(capped) == 20
    # The two best-ranked genes are G1 (rank 1) and G3 (rank 2); their evidence must survive the cap.
    assert {"G1", "G3"} <= set(capped["gene_symbol"])
    # The worst-ranked gene G4 (rank 6) must be dropped first.
    assert "G4" not in set(capped["gene_symbol"])


def test_evidence_sheet_cap_noop_when_small_or_disabled() -> None:
    evidence = pd.DataFrame({"gene_symbol": ["G0", "G1"], "study_id": ["S0", "S1"]})
    genes = pd.DataFrame({"gene_symbol": ["G0", "G1"], "quality_weighted_degora_rank": [1, 2]})
    sheet, was_capped = _cap_evidence_for_sheet(evidence, genes, cap=100)
    assert was_capped is False and len(sheet) == 2
    sheet0, capped0 = _cap_evidence_for_sheet(evidence, genes, cap=0)
    assert capped0 is False and len(sheet0) == 2


def test_evidence_row_cap_reads_env(monkeypatch) -> None:
    monkeypatch.setenv("DEGORA_EXCEL_EVIDENCE_ROW_CAP", "5000")
    assert _evidence_row_cap() == 5000
    monkeypatch.setenv("DEGORA_EXCEL_EVIDENCE_ROW_CAP", "0")
    assert _evidence_row_cap() == 0  # 0 disables the cap
    monkeypatch.setenv("DEGORA_EXCEL_EVIDENCE_ROW_CAP", "not-an-int")
    assert _evidence_row_cap() > 0  # invalid value falls back to the default


def test_export_caps_large_evidence_in_sql_but_reports_uncapped_total(tmp_path, monkeypatch) -> None:
    result_dir = tmp_path / "results"
    result_dir.mkdir()
    db_path = result_dir / "degora_scores.db"
    genes = pd.DataFrame(
        {
            "gene_symbol": ["TOP", "LOW"],
            "quality_weighted_degora_rank": [1, 2],
            "quality_weighted_degora_score": [0.9, 0.1],
        }
    )
    evidence = pd.DataFrame(
        {
            "gene_symbol": ["LOW"] + ["TOP"] * 100_000,
            "study_id": ["LOW_SOURCE"] + [f"S{i % 3}" for i in range(100_000)],
        }
    )
    studies = pd.DataFrame({"study_id": ["S0"], "source_unit_id": ["P0"]})
    with sqlite3.connect(db_path) as connection:
        genes.to_sql("genes", connection, index=False)
        evidence.to_sql("gene_evidence", connection, index=False)
        studies.to_sql("studies", connection, index=False)
        pd.DataFrame({"key": ["corpus"], "value": ["large-test"]}).to_sql(
            "meta", connection, index=False
        )
    (result_dir / "degora_score_metadata.json").write_text("{}\n", encoding="utf-8")
    monkeypatch.setenv("DEGORA_EXCEL_EVIDENCE_ROW_CAP", "100")

    exported = export_run_workbook(result_dir, command="pytest large evidence export")

    assert exported["rows_gene_evidence"] == 100_001
    workbook = load_workbook(exported["output"], read_only=True, data_only=False)
    assert workbook["Gene_evidence"].max_row == 101
    assert {
        row[0].value: row[1].value
        for row in workbook["Run_summary"].iter_rows(min_row=2, values_only=False)
    }["gene_evidence_rows_total"] == 100_001
    summary = {
        row[0].value: row[1].value
        for row in workbook["Run_summary"].iter_rows(min_row=2, values_only=False)
    }
    assert "complete evidence for all genes is in degora_scores.db" in summary["gene_evidence_note"]
    assert "not source-level evidence" in summary["gene_evidence_note"]
    assert summary["path_base"] == "workbook_directory"
    assert summary["result_dir"] == "."
    exported_symbols = {
        row[0].value for row in workbook["Gene_evidence"].iter_rows(min_row=2, values_only=False)
    }
    assert exported_symbols == {"TOP"}
    manifest_path = result_dir / "DEGORA_output.manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert manifest["path_base"] == "manifest_directory"
    assert manifest["inputs"] == ["degora_scores.db", "degora_score_metadata.json"]
    assert manifest["outputs"] == [
        "DEGORA_output.xlsx",
        "DEGORA_output.manifest.json",
        "DEGORA_output.validation.txt",
    ]
    assert str(tmp_path) not in json.dumps(manifest)
    assert str(tmp_path) not in (result_dir / "DEGORA_output.xlsx.source").read_text(encoding="utf-8")


def test_formula_like_strings_are_written_as_literal_xlsx_text(tmp_path) -> None:
    output = tmp_path / "formula_safe.xlsx"
    values = ["=2+2", "+cmd", "-2+3", "@SUM(A1:A2)", " \t=2+2", "#N/A", "#REF!", "#GETTING_DATA"]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame({"gene_symbol": values}).to_excel(writer, sheet_name="Gene_scores", index=False)
        _force_formula_like_text(writer)

    workbook = load_workbook(output, data_only=False)
    cells = list(workbook["Gene_scores"].iter_rows(min_row=2, max_col=1))
    assert [row[0].value for row in cells] == values
    assert all(row[0].data_type == "s" for row in cells)
    with ZipFile(output) as archive:
        sheet_xml = archive.read("xl/worksheets/sheet1.xml")
    assert b"<f" not in sheet_xml


def test_workbook_metadata_uses_canonical_heterogeneity_rule_text(tmp_path) -> None:
    result_dir = _minimal_result_dir(tmp_path)
    exported = export_run_workbook(result_dir, command="pytest workbook metadata")

    workbook = load_workbook(exported["output"], read_only=False, data_only=False)
    all_cell_text = [
        str(cell.value)
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    all_comment_text = [
        str(cell.comment.text)
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.comment is not None
    ]
    workbook_text = "\n".join([*all_cell_text, *all_comment_text])

    assert COLUMN_DEFINITIONS["heterogeneity_i2"][0] == HETEROGENEITY_RULE
    assert RANDOM_EFFECTS_STOUFFER_RULE in COLUMN_DEFINITIONS["re_stouffer_padj"][0]
    assert "no calibrated bias direction" in workbook_text
    assert "positively biased" not in workbook_text
    assert "small-k-biased" not in workbook_text
    assert "small-k positive bias" not in workbook_text


def test_workbook_summary_preserves_dirty_revision_provenance(tmp_path, monkeypatch) -> None:
    import degora.excel_export as excel_export

    monkeypatch.setattr(
        excel_export,
        "runtime_version_info",
        lambda: {
            "degora_version": "0.4.14",
            "degora_code_revision": "abc1234-dirty",
            "degora_code_dirty": "true",
        },
    )
    result_dir = _minimal_result_dir(tmp_path)

    exported = export_run_workbook(result_dir, command="pytest dirty revision provenance")

    workbook = load_workbook(exported["output"], read_only=True, data_only=True)
    summary = {
        row[0]: row[1]
        for row in workbook["Run_summary"].iter_rows(min_row=2, values_only=True)
        if row[0]
    }
    assert summary["degora_code_revision"] == "abc1234-dirty"


def _minimal_result_dir(root, gene_symbol: str = "TOP"):
    result_dir = root / "results"
    result_dir.mkdir(parents=True)
    db_path = result_dir / "degora_scores.db"
    with sqlite3.connect(db_path) as connection:
        pd.DataFrame(
            {
                "gene_symbol": [gene_symbol],
                "quality_weighted_degora_rank": [1],
                "quality_weighted_degora_score": [0.9],
                "heterogeneity_i2": [0.0],
                "re_stouffer_z": [1.2],
                "re_stouffer_p": [0.23],
                "re_stouffer_padj": [0.23],
                "re_stouffer_shrinkage_factor": [1.0],
            }
        ).to_sql("genes", connection, index=False)
        pd.DataFrame({"gene_symbol": [gene_symbol], "study_id": ["S0"]}).to_sql(
            "gene_evidence", connection, index=False
        )
        pd.DataFrame({"study_id": ["S0"], "source_unit_id": ["P0"]}).to_sql(
            "studies", connection, index=False
        )
        pd.DataFrame({"key": ["corpus"], "value": ["repro-test"]}).to_sql(
            "meta", connection, index=False
        )
    (result_dir / "degora_score_metadata.json").write_text("{}\n", encoding="utf-8")
    return result_dir


def test_workbook_bytes_are_identical_for_identical_inputs(tmp_path) -> None:
    """The workbook must be reproducible like the CSV and SQLite outputs beside it.

    openpyxl stamps the save time into docProps/core.xml and into every ZIP member,
    so two runs over the same inputs used to differ in bytes while every sheet was
    identical. That made the recorded artifact sha256 unusable for verification.
    """

    first = export_run_workbook(_minimal_result_dir(tmp_path / "a"), command="pytest repro")
    second = export_run_workbook(_minimal_result_dir(tmp_path / "b"), command="pytest repro")

    first_bytes = Path(first["output"]).read_bytes()
    second_bytes = Path(second["output"]).read_bytes()
    assert first_bytes == second_bytes

    with ZipFile(first["output"]) as archive:
        stamps = {info.date_time for info in archive.infolist()}
        core = archive.read("docProps/core.xml").decode("utf-8")
    assert stamps == {(2000, 1, 1, 0, 0, 0)}
    assert "2000-01-01T00:00:00Z" in core

    # The workbook must still be a readable OOXML file with its data intact.
    workbook = load_workbook(first["output"], read_only=True, data_only=True)
    assert workbook["Gene_scores"].max_row == 2


def test_demo_and_template_workbooks_are_reproducible(tmp_path) -> None:
    """The inputs a reader starts from have to be reproducible like the outputs.

    v0.4.7 pinned the timestamps in the generated workbook. The config workbooks
    `degora demo` and `degora template` write were still stamped with the clock,
    so two demo runs produced inputs with different checksums -- and every
    provenance sidecar recording an input hash differed with them.
    """

    from degora.demo import write_demo_workspace
    from degora.excel_template import write_template

    first_template = write_template(tmp_path / "a" / "DEGORA_template.xlsx")
    second_template = write_template(tmp_path / "b" / "DEGORA_template.xlsx")
    assert Path(first_template).read_bytes() == Path(second_template).read_bytes()

    first_demo = write_demo_workspace(tmp_path / "demo-a")
    second_demo = write_demo_workspace(tmp_path / "demo-b")
    first_config = Path(first_demo["demo_dir"]) / "degora_demo_config.xlsx"
    second_config = Path(second_demo["demo_dir"]) / "degora_demo_config.xlsx"
    assert first_config.read_bytes() == second_config.read_bytes()

    with ZipFile(first_config) as archive:
        assert {info.date_time for info in archive.infolist()} == {(2000, 1, 1, 0, 0, 0)}
