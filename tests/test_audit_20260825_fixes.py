"""Regression tests for the 2026-08-25 code audit of the v0.4.17 branch.

Each test is named after what the audit ran into. Fixtures are kept small and
offline; the one live-GEO fact the audit established (the Series record lists
organisms as ``!Series_sample_organism``) is pinned with the bytes GEO returned.
"""

from __future__ import annotations

import json
import os
import sqlite3
import stat
import threading
import urllib.request
from pathlib import Path
from urllib.error import HTTPError

import numpy as np
import pandas as pd
import pytest
from scipy import stats

import degora.api as api
from degora.aggregate import duration_hours, time_course_selection_report
from degora.beginner import infer_source_table, run_init
from degora.discovery import _matrix_rows_from_text, classify_header, parse_geo_soft
from degora.discovery_federated import _readiness, _safe_provider_error, _species_decision
from degora.discovery import normalize_species
from degora.excel_io import locked_panel_mask
from degora.harmonize import (
    TableMapping,
    canonical_gene_symbol,
    duplicate_source_headers,
    harmonize_frame,
    read_deg_table,
)
from degora.provenance import redact_secrets_in_text
from degora.reanalysis import welch_with_variance_floor
from degora.score_db import degora_score_table, direction_conflict_warnings
from degora.slice_runner import (
    DegoraConfigError,
    _read_locked_gold_panel,
    run_slice,
    validate_catalog_inputs,
)


# --------------------------------------------------------------------------- helpers


def _table(genes, lfc, p, *, extra=None) -> pd.DataFrame:
    frame = pd.DataFrame({"gene": genes, "log2FoldChange": lfc, "pvalue": p})
    if extra:
        for key, value in extra.items():
            frame[key] = value
    return frame


def _two_unit_catalog(tmp_path: Path, first: pd.DataFrame, second: pd.DataFrame, **overrides) -> Path:
    deg = tmp_path / "deg"
    deg.mkdir(exist_ok=True)
    first.to_csv(deg / "a.csv", index=False)
    second.to_csv(deg / "b.csv", index=False)
    rows = []
    for study, unit, name in (("S1", "U1", "a.csv"), ("S2", "U2", "b.csv")):
        row = {
            "study_id": study,
            "source_unit_id": unit,
            "source_path": f"deg/{name}",
            "gene_column": "gene",
            "lfc_column": "log2FoldChange",
            "p_column": "pvalue",
        }
        row.update(overrides.get(study, {}))
        rows.append(row)
    config = tmp_path / "config.csv"
    pd.DataFrame(rows).to_csv(config, index=False)
    return config


def _clean_pair(n: int = 40, seed: int = 0) -> tuple[pd.DataFrame, pd.DataFrame]:
    rng = np.random.default_rng(seed)
    genes = [f"G{i:03d}" for i in range(n)]
    lfc = rng.normal(0, 1.5, n)
    p = rng.uniform(1e-6, 0.05, n)
    return _table(genes, lfc, p), _table(genes, lfc + rng.normal(0, 0.2, n), p * 1.5)


# --------------------------------------------------------------------------- H-1


def test_a_linear_fold_change_column_is_refused_before_it_calls_every_gene_up(tmp_path) -> None:
    rng = np.random.default_rng(1)
    genes = [f"G{i:03d}" for i in range(120)]
    log2fc = rng.normal(0, 1.5, 120)
    linear = _table(genes, 2.0**log2fc, rng.uniform(1e-6, 0.05, 120)).rename(columns={"log2FoldChange": "FoldChange"})
    good = _table(genes, log2fc, rng.uniform(1e-6, 0.05, 120))
    config = _two_unit_catalog(tmp_path, linear, good, S1={"lfc_column": "FoldChange"})

    with pytest.raises(DegoraConfigError) as exc_info:
        validate_catalog_inputs(config)

    message = str(exc_info.value)
    assert "not on a log2 scale" in message
    assert "linear fold change" in message
    assert "every gene in this table would be called up" in message


def test_an_up_only_log2_table_is_warned_about_not_refused(tmp_path) -> None:
    genes = [f"G{i:03d}" for i in range(60)]
    up_only = _table(genes, np.linspace(0.6, 4.0, 60), np.linspace(1e-6, 0.04, 60))
    partner, _ = _clean_pair(60)
    config = _two_unit_catalog(tmp_path, up_only, partner)

    summary = validate_catalog_inputs(config)

    assert summary["active_contrasts"] == 2
    assert not any("no negative values" in warning for warning in summary["warnings"]), (
        "a column named log2FoldChange with values >= 0.6 is an up-only list, not a linear ratio"
    )


def test_absurd_log2_values_are_warned_about(tmp_path) -> None:
    first, second = _clean_pair(30)
    first.loc[0, "log2FoldChange"] = 1e10
    config = _two_unit_catalog(tmp_path, first, second)

    summary = validate_catalog_inputs(config)

    assert any("|log2FC| > 30" in warning for warning in summary["warnings"])


# --------------------------------------------------------------------------- H-2


def test_a_p_value_written_as_a_bound_is_refused_not_dropped(tmp_path) -> None:
    first, second = _clean_pair(300)
    first["pvalue"] = first["pvalue"].astype(object)
    for index in range(8):
        first.loc[index, "pvalue"] = "<1E-16"
    config = _two_unit_catalog(tmp_path, first, second)

    with pytest.raises(DegoraConfigError) as exc_info:
        validate_catalog_inputs(config)

    message = str(exc_info.value)
    assert "written as a bound" in message
    assert "<1E-16" in message
    assert "most significant rows" in message


def test_every_dropped_row_count_reaches_the_run_warnings_whatever_its_share(tmp_path) -> None:
    from degora.cli import _run_warning_messages

    messages = _run_warning_messages(
        {
            "warnings": [],
            "unusable_row_counts": {"S1": 8},
            "input_row_counts": {"S1": 300},
        }
    )

    assert any("S1: 8 of 300 row(s)" in message for message in messages)


# --------------------------------------------------------------------------- H-3


@pytest.mark.parametrize("label", ["30min", "4h", "24 h", "", "inf"])
def test_a_duration_that_is_not_a_plain_number_is_not_a_duration(label) -> None:
    assert np.isnan(duration_hours(label))


@pytest.mark.parametrize("label,hours", [("24", 24.0), ("0.5", 0.5), (" 6 ", 6.0), ("1e1", 10.0)])
def test_plain_numeric_durations_parse(label, hours) -> None:
    assert duration_hours(label) == hours


def test_early_mode_refuses_a_unit_with_text_durations(tmp_path) -> None:
    first, second = _clean_pair(20)
    third = first.copy()
    third["log2FoldChange"] = -third["log2FoldChange"]
    deg = tmp_path / "deg"
    deg.mkdir()
    first.to_csv(deg / "t30.csv", index=False)
    third.to_csv(deg / "t4.csv", index=False)
    second.to_csv(deg / "b.csv", index=False)
    rows = [
        {"study_id": "P1_30min", "source_unit_id": "P1", "source_path": "deg/t30.csv", "duration_h": "30min", "time_course_mode": "early"},
        {"study_id": "P1_4h", "source_unit_id": "P1", "source_path": "deg/t4.csv", "duration_h": "4h", "time_course_mode": "early"},
        {"study_id": "B", "source_unit_id": "U2", "source_path": "deg/b.csv", "duration_h": "", "time_course_mode": "mean"},
    ]
    for row in rows:
        row.update({"gene_column": "gene", "lfc_column": "log2FoldChange", "p_column": "pvalue"})
    config = tmp_path / "config.csv"
    pd.DataFrame(rows).to_csv(config, index=False)

    with pytest.raises(DegoraConfigError) as exc_info:
        validate_catalog_inputs(config)

    message = str(exc_info.value)
    assert "time-course durations are not numeric" in message
    assert "'30min'" in message and "'4h'" in message
    assert "'B'" not in message, "a mean unit does not select by duration and must not be asked for one"


# --------------------------------------------------------------------------- H-4 / M-16


GEO_SERIES_SOFT = """^SERIES = GSE312638
!Series_title = Mixed human and mouse melanoma series
!Series_platform_organism = Homo sapiens
!Series_platform_organism = Mus musculus
!Series_sample_organism = Homo sapiens
!Series_sample_organism = Mus musculus
!Series_supplementary_file = ftp://ftp.ncbi.nlm.nih.gov/geo/series/GSE312nnn/GSE312638/suppl/GSE312638_counts.csv.gz
"""


def test_geo_series_record_organisms_are_read_from_the_keys_geo_actually_emits() -> None:
    parsed = parse_geo_soft(GEO_SERIES_SOFT)

    assert parsed["taxa"] == ["Homo sapiens", "Mus musculus"]
    pure = GEO_SERIES_SOFT.replace("!Series_sample_organism = Mus musculus\n", "").replace(
        "!Series_platform_organism = Mus musculus\n", ""
    )
    assert parse_geo_soft(pure)["taxa"] == ["Homo sapiens"]


def test_per_sample_organism_keys_are_still_read() -> None:
    assert parse_geo_soft("^SAMPLE = GSM1\n!Sample_organism_ch1 = Mus musculus\n!Sample_organism_ch2 = Homo sapiens\n")["taxa"] == [
        "Homo sapiens",
        "Mus musculus",
    ]


def test_a_geo_record_whose_samples_are_the_requested_species_reaches_verified_ready() -> None:
    spec = normalize_species("human")
    record = {
        "provider": "ncbi_geo",
        "record_type": "publication",
        "accession": "GSE319526",
        "species_evidence": {
            "requested": "Homo sapiens",
            "observed_taxa": ["Homo sapiens"],
            "status": "exact",
            "basis": "GEO SOFT observed taxa",
        },
        "quarantined": False,
        "target_species_verified": True,
        "target_species_evidence": "GEO SOFT organism Homo sapiens (per-record check)",
        "supplementary_file_candidates": [
            {"name": "GSE319526_counts.csv.gz", "url": "https://x/GSE319526_counts.csv.gz", "role": "count_matrix", "tier": "tabular"}
        ],
        "detail_assessment": "complete",
    }

    assert _species_decision(record, spec) == "target_species_verified"
    assert _readiness(record, spec)["tier"] == "verified_ready"


def test_the_federated_geo_provider_marks_exact_organism_records_verified() -> None:
    import urllib.parse

    from degora.discovery import NcbiRequestConfig
    from degora.discovery_sources import NcbiGeoProvider
    from test_discovery_sources import Response, public_transport

    search_url = (
        "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?"
        + urllib.parse.urlencode(
            {
                "db": "gds",
                "term": '("hypoxia"[All Fields]) AND gse[ETYP] AND "Homo sapiens"[Organism]',
                "retmax": 1,
                "retmode": "json",
                "sort": "relevance",
                "usehistory": "y",
                "tool": "degora-test",
            }
        )
    )
    summary_url = (
        "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi?"
        + urllib.parse.urlencode({"db": "gds", "id": "999", "retmode": "json", "tool": "degora-test"})
    )
    soft_url = (
        "https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?"
        + urllib.parse.urlencode({"acc": "GSE9", "targ": "self", "form": "text", "view": "brief"})
    )
    soft = b"^SERIES = GSE9\n!Series_title = Pure human\n!Series_sample_organism = Homo sapiens\n!Series_platform_organism = Homo sapiens\n"
    provider = NcbiGeoProvider(
        transport=public_transport(
            {
                search_url: Response(search_url, b'{"esearchresult":{"idlist":["999"]}}'),
                summary_url: Response(summary_url, b'{"result":{"999":{"accession":"GSE9","pubmedids":["111"]}}}'),
                soft_url: Response(soft_url, soft),
            }
        ),
        request_config=NcbiRequestConfig(tool="degora-test"),
        sleep=lambda _x: None,
    )

    records = provider.search("hypoxia", normalize_species("human"), 1)

    assert records[0]["species_evidence"]["status"] == "exact"
    assert records[0]["target_species_verified"] is True
    assert records[0]["quarantined"] is False


# --------------------------------------------------------------------------- H-5 / M-8 / M-17 / L-18 / L-19


def test_the_species_confirmation_box_drives_run_eligibility() -> None:
    assert '$("speciesConfirmed").addEventListener("change", updateAnalysisEligibility)' in api.INDEX_HTML


def test_a_new_search_empties_the_previous_snapshot_when_submitted() -> None:
    script = api.INDEX_HTML
    submit = script.index("state.query = query;")
    assert "state.studies = [];" in script[submit : submit + 800]


def test_the_browser_opens_the_workspace_the_demo_prefilled() -> None:
    assert "if (preferredDiscoverySpecies !== activeSpecies) setSpecies(preferredDiscoverySpecies);" in api.INDEX_HTML


def test_the_result_filter_listens_for_input_and_the_stop_button_is_updated_in_place() -> None:
    assert 'if (event.target.id === "resultFilter") applyResultFilter(event);' in api.INDEX_HTML
    assert 'querySelector(".search-progress")' in api.INDEX_HTML


# --------------------------------------------------------------------------- H-6 / L-17 / F1


def _scored_db(tmp_path: Path, directory_name: str) -> Path:
    from degora.score_db import write_score_database

    harmonized = pd.DataFrame(
        {
            "study_id": ["S1", "S2"],
            "paper_id": ["P1", "P2"],
            "gene_symbol": ["VEGFA", "VEGFA"],
            "lfc": [2.0, 1.8],
            "signed_z": [5.0, 4.5],
            "pvalue": [1e-7, 1e-6],
            "padj": [1e-5, 1e-4],
            "normalized_rank": [0.02, 0.03],
            "n_genes_in_study": [1000, 1000],
        }
    )
    folder = tmp_path / directory_name
    folder.mkdir()
    harmonized_path = folder / "harmonized.csv"
    harmonized.to_csv(harmonized_path, index=False)
    write_score_database(harmonized_path, folder, db_path=folder / "degora_scores.db")
    return folder / "degora_scores.db"


@pytest.mark.parametrize("directory_name", ["exp#3", "q?x", "p%41ct", "50% O2"])
def test_serve_accepts_a_database_whose_path_carries_uri_syntax(tmp_path, directory_name) -> None:
    db_path = _scored_db(tmp_path, directory_name)
    before = sorted(path.name for path in tmp_path.iterdir())

    api._require_degora_score_database(db_path)

    assert sorted(path.name for path in tmp_path.iterdir()) == before, "the preflight must not create files"


def test_serve_refuses_a_database_without_the_studies_table(tmp_path) -> None:
    db_path = _scored_db(tmp_path, "run")
    with sqlite3.connect(db_path) as connection:
        connection.execute("DROP TABLE studies")

    with pytest.raises(api.ScoreDatabaseError, match="missing studies"):
        api._require_degora_score_database(db_path)


def test_serve_refuses_a_database_missing_a_column_the_dashboard_reads(tmp_path) -> None:
    db_path = _scored_db(tmp_path, "run")
    with sqlite3.connect(db_path) as connection:
        connection.execute("CREATE TABLE studies_trimmed AS SELECT study_id FROM studies")
        connection.execute("DROP TABLE studies")
        connection.execute("ALTER TABLE studies_trimmed RENAME TO studies")

    with pytest.raises(api.ScoreDatabaseError, match="studies.source_unit_id"):
        api._require_degora_score_database(db_path)


# --------------------------------------------------------------------------- H-7


def test_one_table_declared_under_two_source_units_is_refused(tmp_path) -> None:
    first, _ = _clean_pair(20)
    deg = tmp_path / "deg"
    deg.mkdir()
    first.to_csv(deg / "a.csv", index=False)
    rows = []
    for study, unit in (("S1", "U1"), ("S2", "U2")):
        rows.append(
            {"study_id": study, "source_unit_id": unit, "source_path": "deg/a.csv", "gene_column": "gene", "lfc_column": "log2FoldChange", "p_column": "pvalue"}
        )
    config = tmp_path / "config.csv"
    pd.DataFrame(rows).to_csv(config, index=False)

    with pytest.raises(DegoraConfigError, match="declared as two independent source units"):
        validate_catalog_inputs(config)


def test_byte_identical_copies_under_two_source_units_are_refused(tmp_path) -> None:
    first, _ = _clean_pair(20)
    config = _two_unit_catalog(tmp_path, first, first.copy())

    with pytest.raises(DegoraConfigError) as exc_info:
        validate_catalog_inputs(config)

    message = str(exc_info.value)
    assert "byte-identical" in message
    assert "S1 (deg/a.csv) and S2 (deg/b.csv)" in message, "the catalog's own path text is reported, not a resolved absolute path"


def test_one_workbook_read_twice_inside_one_source_unit_is_fine(tmp_path) -> None:
    first, second = _clean_pair(20)
    first["log2FoldChange_24h"] = second["log2FoldChange"]
    first["pvalue_24h"] = second["pvalue"]
    partner = second.copy()
    partner["log2FoldChange"] += 0.3
    deg = tmp_path / "deg"
    deg.mkdir()
    first.to_csv(deg / "a.csv", index=False)
    partner.to_csv(deg / "b.csv", index=False)
    rows = [
        {"study_id": "S1_4h", "source_unit_id": "U1", "source_path": "deg/a.csv", "gene_column": "gene", "lfc_column": "log2FoldChange", "p_column": "pvalue"},
        {"study_id": "S1_24h", "source_unit_id": "U1", "source_path": "deg/a.csv", "gene_column": "gene", "lfc_column": "log2FoldChange_24h", "p_column": "pvalue_24h"},
        {"study_id": "S2", "source_unit_id": "U2", "source_path": "deg/b.csv", "gene_column": "gene", "lfc_column": "log2FoldChange", "p_column": "pvalue"},
    ]
    config = tmp_path / "config.csv"
    pd.DataFrame(rows).to_csv(config, index=False)

    summary = validate_catalog_inputs(config)

    assert summary["source_units"] == 2
    assert not any("same table" in warning for warning in summary["warnings"])


# --------------------------------------------------------------------------- H-8


def test_a_paper_id_only_catalog_reports_time_course_selection_per_paper() -> None:
    rows = []
    for study, paper, duration in (("A_2h", "PA", 2), ("A_24h", "PA", 24), ("B_6h", "PB", 6), ("B_48h", "PB", 48)):
        for gene in ("G1", "G2"):
            rows.append(
                {
                    "study_id": study,
                    "paper_id": paper,
                    "source_unit_id": "",
                    "gene_symbol": gene,
                    "lfc": 1.0,
                    "pvalue": 0.01,
                    "signed_z": 2.5,
                    "normalized_rank": 0.1,
                    "n_genes_in_study": 2,
                    "duration_h": duration,
                    "time_course_mode": "early",
                }
            )
    harmonized = pd.DataFrame(rows)

    report = time_course_selection_report(harmonized)

    assert [entry["source_unit_id"] for entry in report] == ["PA", "PB"]
    assert all(entry["rows_before"] == 4 and entry["rows_after"] == 2 for entry in report)


def test_a_paper_id_only_catalog_with_mixed_modes_no_longer_crashes_the_report() -> None:
    rows = []
    for study, paper, duration, mode in (("A_2h", "PA", 2, "early"), ("A_24h", "PA", 24, "early"), ("B_1", "PB", "", "mean")):
        rows.append(
            {
                "study_id": study,
                "paper_id": paper,
                "source_unit_id": "",
                "gene_symbol": "G1",
                "lfc": 1.0,
                "pvalue": 0.01,
                "signed_z": 2.5,
                "normalized_rank": 0.1,
                "n_genes_in_study": 1,
                "duration_h": duration,
                "time_course_mode": mode,
            }
        )

    report = time_course_selection_report(pd.DataFrame(rows))

    assert [entry["source_unit_id"] for entry in report] == ["PA"]


# --------------------------------------------------------------------------- M-1 / F2


def test_gold_panel_locked_flags_typed_as_numbers_are_read_as_written() -> None:
    flags = pd.Series([1.0, 1.0, np.nan, 0.0, "yes", "no"])

    assert locked_panel_mask(flags).tolist() == [True, True, True, False, True, False]


def test_a_gold_panel_workbook_with_numeric_locked_flags_keeps_the_marked_rows(tmp_path) -> None:
    contrasts = pd.DataFrame(
        [{"study_id": "S1", "source_unit_id": "U1", "source_path": "a.csv", "gene_column": "gene", "lfc_column": "log2FoldChange", "p_column": "pvalue"}]
    )
    gold = pd.DataFrame({"gene_symbol": ["TP53", "EGFR", "MYC"], "locked": [1, 0, np.nan]})
    config = tmp_path / "config.xlsx"
    with pd.ExcelWriter(config, engine="openpyxl") as writer:
        contrasts.to_excel(writer, sheet_name="Contrasts", index=False)
        gold.to_excel(writer, sheet_name="GoldPanel", index=False)

    panel = _read_locked_gold_panel(config)

    assert panel["genes"] == ["MYC", "TP53"]


# --------------------------------------------------------------------------- M-2 / F5


@pytest.mark.parametrize(
    "label,expected",
    [
        ("NKX2.5", "NKX2.5"),
        ("Nkx2.1", "NKX2.1"),
        ("ENSG00000141510.16", "ENSG00000141510"),
        ("ENSMUSG00000015627.8", "ENSMUSG00000015627"),
        ("NM_000546.5", "NM_000546"),
        ("7157.0", "7157"),
        (7157.0, "7157"),
        ("NA", ""),
        ("<NA>", ""),
        ("N/A", ""),
        ("NULL", ""),
        ("SEPT9", "SEPTIN9"),
    ],
)
def test_version_suffixes_are_stripped_only_from_accessions(label, expected) -> None:
    assert canonical_gene_symbol(label) == expected


def test_the_vector_and_scalar_gene_rules_agree_including_missing_sentinels() -> None:
    frame = pd.DataFrame(
        {
            "gene": ["NKX2.5", "NKX2.1", "NA", "<NA>", "ENSG00000141510.16", "SEPT9"],
            "log2FoldChange": [1.0, -1.0, 1.0, 1.0, 1.0, 1.0],
            "pvalue": [0.01] * 6,
        }
    )

    out = harmonize_frame(frame, TableMapping("gene", "log2FoldChange", "pvalue"), {"study_id": "S"})

    assert sorted(out["gene_symbol"]) == ["ENSG00000141510", "NKX2.1", "NKX2.5", "SEPTIN9"]
    assert int(out["n_rows_dropped_unusable"].iloc[0]) == 2


# --------------------------------------------------------------------------- M-3


def test_a_repeated_header_cannot_be_mapped_by_its_bare_name(tmp_path) -> None:
    path = tmp_path / "two_contrasts.csv"
    path.write_text("gene,logFC,P.Value,logFC,P.Value\nG1,1.0,0.01,-1.0,0.02\nG2,0.5,0.02,-0.5,0.03\n", encoding="utf-8")

    frame = read_deg_table(path, TableMapping("gene", "logFC", "P.Value"))

    assert duplicate_source_headers(frame) == {"logFC": 2, "P.Value": 2}
    partner = tmp_path / "b.csv"
    _clean_pair(10)[0].to_csv(partner, index=False)
    config = tmp_path / "config.csv"
    pd.DataFrame(
        [
            {"study_id": "S1", "source_unit_id": "U1", "source_path": path.name, "gene_column": "gene", "lfc_column": "logFC", "p_column": "P.Value"},
            {"study_id": "S2", "source_unit_id": "U2", "source_path": partner.name, "gene_column": "gene", "lfc_column": "log2FoldChange", "p_column": "pvalue"},
        ]
    ).to_csv(config, index=False)

    with pytest.raises(DegoraConfigError) as exc_info:
        validate_catalog_inputs(config)

    message = str(exc_info.value)
    assert "repeated column headers" in message
    assert "'logFC.1'" in message


def test_an_explicit_pandas_name_selects_the_later_block(tmp_path) -> None:
    path = tmp_path / "two_contrasts.csv"
    path.write_text("gene,logFC,P.Value,logFC,P.Value\nG1,1.0,0.01,-1.0,0.02\nG2,0.5,0.02,-0.5,0.03\n", encoding="utf-8")

    frame = read_deg_table(path, TableMapping("gene", "logFC.1", "P.Value.1"))
    out = harmonize_frame(frame, TableMapping("gene", "logFC.1", "P.Value.1"), {"study_id": "S"})

    assert out.set_index("gene_symbol").loc["G1", "lfc"] == -1.0


# --------------------------------------------------------------------------- M-4


def test_identical_replicates_no_longer_yield_p_zero_and_other_genes_match_scipy() -> None:
    rng = np.random.default_rng(0)
    treatment = rng.normal(5, 1, (300, 3))
    control = rng.normal(5, 1, (300, 3))
    treatment[0] = [5.1, 5.1, 5.1]
    control[0] = [5.0, 5.0, 5.0]

    result = welch_with_variance_floor(treatment, control)
    reference = stats.ttest_ind(treatment, control, axis=1, equal_var=False)

    assert result["pvalue"][0] > 0.01, "a 0.1 log2 difference between identical replicates is not p = 0"
    floored = (treatment.var(axis=1, ddof=1) < result["variance_floor"]) | (control.var(axis=1, ddof=1) < result["variance_floor"])
    untouched = ~floored
    assert untouched.sum() > 250
    assert np.allclose(result["pvalue"][untouched], reference.pvalue[untouched])
    assert result["n_zero_variance_rows"] == 1


# --------------------------------------------------------------------------- M-5


def test_an_oversized_field_does_not_end_a_matrix_inspection() -> None:
    text = "colA,colB,colC\n" + "X" * 200 * 1024 + ",1,2\nfoo,3,4\n"

    rows = _matrix_rows_from_text(text)

    assert isinstance(rows, list)


# --------------------------------------------------------------------------- M-6


def test_a_source_whose_direction_runs_against_the_corpus_is_flagged_without_changing_ranks() -> None:
    rng = np.random.default_rng(2)
    genes = [f"G{i:03d}" for i in range(200)]
    truth = rng.normal(0, 1.5, 200)
    rows = []
    for unit, sign in (("U1", 1), ("U2", 1), ("U3", 1), ("U4", -1)):
        lfc = sign * (truth + rng.normal(0, 0.3, 200))
        z = lfc / 0.3
        for gene, effect, zed in zip(genes, lfc, z):
            rows.append(
                {
                    "study_id": unit,
                    "paper_id": unit,
                    "source_unit_id": unit,
                    "gene_symbol": gene,
                    "lfc": effect,
                    "signed_z": zed,
                    "pvalue": float(2 * stats.norm.sf(abs(zed))),
                    "padj": np.nan,
                    "normalized_rank": 0.5,
                    "n_genes_in_study": 200,
                    "n_ctrl": 3,
                    "n_treat": 3,
                    "table_scope": "full_results",
                    "source_input_type": "author_deg_table",
                }
            )
    harmonized = pd.DataFrame(rows)

    scores, _evidence, metadata = degora_score_table(harmonized, min_studies=2)
    diagnostics = pd.DataFrame(metadata["source_quality_diagnostics"])

    flagged = diagnostics.loc[diagnostics["source_direction_conflict_flag"], "source_unit_id"].tolist()
    assert flagged == ["U4"]
    assert metadata["n_source_direction_conflicts"] == 1
    warnings = direction_conflict_warnings(metadata["source_quality_diagnostics"])
    assert len(warnings) == 1 and "'U4'" in warnings[0]
    assert diagnostics.set_index("source_unit_id").loc["U4", "source_reliability_weight"] == pytest.approx(
        diagnostics.set_index("source_unit_id").loc["U1", "source_reliability_weight"]
    ), "the flag is advisory: weights are unchanged"


# --------------------------------------------------------------------------- M-7 / L-16 / L-21


@pytest.mark.parametrize("value", ["1%/21% O2", "hypoxia (1%/21% O2) vs normoxia", "log2FC>1 &/or padj<0.05", "ratio (A)/(B)"])
def test_experimental_metadata_is_not_mistaken_for_a_local_path(value) -> None:
    assert api._contains_local_path(value) is False


@pytest.mark.parametrize("value", ["/home/user/deg/table.csv", "file:///tmp/x.csv", "C:\\Users\\me\\deg.csv", "/etc/passwd"])
def test_real_local_paths_are_still_redacted(value) -> None:
    assert api._contains_local_path(value) is True


def test_loopback_aliases_are_loopback_everywhere() -> None:
    assert api._is_loopback_host("127.0.0.2")
    assert api._is_loopback_host("::1")
    assert not api._is_loopback_host("0.0.0.0")


def test_non_finite_values_inside_tuples_serialise_as_null() -> None:
    assert json.dumps(api._jsonable({"ci": (float("nan"), 1.0), "set": {float("inf")}}), sort_keys=True) == '{"ci": [null, 1.0], "set": [null]}'


def test_search_timestamps_are_iso8601() -> None:
    assert api._iso_timestamp(0.0) == "1970-01-01T00:00:00Z"
    assert api._iso_timestamp("2026-08-25T00:00:00Z") == "2026-08-25T00:00:00Z"
    assert api._iso_timestamp(None) is None


# --------------------------------------------------------------------------- M-9 / M-11 / M-12 / L-15


def _start(db_path: Path, **kwargs):
    server = api.create_server(db_path, port=0, quiet=True, **kwargs)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    return server, thread


def _stop(server, thread) -> None:
    server.shutdown()
    server.server_close()
    thread.join(timeout=5)


def test_a_non_ascii_token_is_a_401_not_a_dropped_connection(tmp_path) -> None:
    db_path = _scored_db(tmp_path, "run")
    server, thread = _start(db_path, access_token="secret-token")
    host, port = server.server_address[:2]
    try:
        request = urllib.request.Request(f"http://{host}:{port}/api/health", headers={"X-DEGORA-Token": "abc\xe9"})
        with pytest.raises(HTTPError) as exc_info:
            urllib.request.urlopen(request, timeout=5)
        assert exc_info.value.code == 401
    finally:
        _stop(server, thread)


def test_infinity_in_a_json_limit_is_a_400(tmp_path) -> None:
    db_path = _scored_db(tmp_path, "run")
    server, thread = _start(db_path)
    host, port = server.server_address[:2]
    try:
        body = b'{"query":"hypoxia","species":"human","limit":Infinity}'
        request = urllib.request.Request(
            f"http://{host}:{port}/api/discovery/searches",
            data=body,
            method="POST",
            headers={"Content-Type": "application/json", "X-DEGORA-Action": "1"},
        )
        with pytest.raises(HTTPError) as exc_info:
            urllib.request.urlopen(request, timeout=5)
        assert exc_info.value.code == 400
    finally:
        _stop(server, thread)


def test_an_unwritable_results_folder_is_reported_as_a_workspace_problem(tmp_path, monkeypatch) -> None:
    db_path = _scored_db(tmp_path, "run")

    def refuse(self, *args, **kwargs):
        raise PermissionError(13, "Permission denied")

    monkeypatch.setattr(Path, "mkdir", refuse)
    with pytest.raises(api.DiscoveryWorkspaceError, match="discovery workspace"):
        api.create_server(db_path, port=0, quiet=True)


def test_ipv6_loopback_binds_and_prints_a_bracketed_url(tmp_path) -> None:
    db_path = _scored_db(tmp_path, "run")
    try:
        server = api.create_server(db_path, host="::1", port=0, quiet=True)
    except OSError as exc:  # pragma: no cover - no IPv6 loopback on this host
        pytest.skip(f"IPv6 loopback unavailable: {exc}")
    try:
        assert server.address_family == __import__("socket").AF_INET6
    finally:
        server.server_close()


def test_the_server_reuses_its_address_on_posix() -> None:
    assert api.DegoraHttpServer.allow_reuse_address == (os.name != "nt")


# --------------------------------------------------------------------------- F4


@pytest.mark.parametrize(
    "text",
    [
        "Authorization: Bearer SYNTHETIC_BEARER_REVIEW_7F3A",
        "authorization: Basic dXNlcjpwYXNz",
        "https://user:hunter2@example.org/x?api_key=ABC123&db=gds",
        'request failed {"api_key": "ABC123", "token": "T0K"}',
        "X-API-Key: k-123 rejected",
    ],
)
def test_credentials_in_provider_errors_are_redacted(text) -> None:
    cleaned = redact_secrets_in_text(text)
    for secret in ("SYNTHETIC_BEARER_REVIEW_7F3A", "dXNlcjpwYXNz", "hunter2", "ABC123", "T0K", "k-123"):
        assert secret not in cleaned
    assert "SYNTHETIC" not in _safe_provider_error(RuntimeError("Authorization: Bearer SYNTHETIC_BEARER_REVIEW_7F3A"))


# --------------------------------------------------------------------------- M-14 / L-7 / L-8 / L-11 (init)


def test_init_finds_a_table_on_a_later_sheet_below_a_title_row(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    path = deg / "supplementary.xlsx"
    table = _table([f"G{i}" for i in range(30)], np.linspace(-2, 2, 30), np.linspace(1e-5, 0.04, 30))
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame({"README": ["Supplementary Table S3. Differential expression."]}).to_excel(writer, sheet_name="README", index=False)
        table.to_excel(writer, sheet_name="DEGs", index=False, startrow=2)
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        pd.DataFrame([["Table S3: hypoxia vs normoxia"]]).to_excel(writer, sheet_name="DEGs", index=False, header=False, startrow=0)

    inference = infer_source_table(path)

    assert inference.looks_like_a_deg_table
    assert inference.sheet_name == "DEGs"
    assert inference.header_row == 3


def test_init_recognises_single_cell_marker_tables_without_a_question() -> None:
    header = classify_header(["gene", "p_val", "avg_log2FC", "pct.1", "pct.2", "p_val_adj"])

    assert header["mapping"] == {"gene_column": "gene", "lfc_column": "avg_log2FC", "p_column": "p_val", "padj_column": "p_val_adj"}
    assert header["status"] == "candidate_header"


def test_init_skips_degora_output_tables(tmp_path) -> None:
    path = tmp_path / "degora_gene_scores.csv"
    pd.DataFrame({"degora_rank": [1], "degora_score": [90.0], "quality_weighted_degora_rank": [1], "gene_symbol": ["A"]}).to_csv(path, index=False)

    inference = infer_source_table(path)

    assert "DEGORA output table" in inference.problem


def test_init_asks_whether_a_fold_change_column_is_log2_and_excludes_a_no(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    rng = np.random.default_rng(3)
    genes = [f"G{i:03d}" for i in range(60)]
    pd.DataFrame({"Gene symbol": genes, "FoldChange": 2.0 ** rng.normal(0, 1, 60), "P.Value": rng.uniform(1e-5, 0.04, 60)}).to_csv(
        deg / "linear.csv", index=False
    )
    lines: list[str] = []
    # species; accept the default FoldChange column; "no" to log2; "yes" to direction;
    # comparison; paper; "n=3" is refused and asked again; treated count.
    answers = iter(["human", "", "no", "yes", "drug vs vehicle", "PAPER_A", "n=3", "3", "3"])

    summary = run_init(tmp_path / "config.csv", deg, ask=lambda question, default="": next(answers), echo=lines.append)

    config = pd.read_csv(tmp_path / "config.csv")
    assert summary["n_excluded_not_log2"] == 1
    assert config.loc[0, "include_in_analysis"] == "no"
    assert str(config.loc[0, "sign_convention"]).startswith("NOT LOG2")
    assert any("shape of a LINEAR fold change" in line for line in lines)
    assert str(config.loc[0, "n_ctrl"]) == "3", "'n=3' was rejected and the reader asked again"


# --------------------------------------------------------------------------- L-9 / L-10 / L-14


def test_every_run_artifact_has_the_umask_permissions_of_a_plain_write(tmp_path) -> None:
    first, second = _clean_pair(15)
    config = _two_unit_catalog(tmp_path, first, second)
    previous = os.umask(0o022)
    try:
        run_slice(config, tmp_path / "out", tmp_path / "harmonized", min_studies=2)
    finally:
        os.umask(previous)

    modes = {path.name: stat.S_IMODE(path.stat().st_mode) for path in (tmp_path / "out").iterdir() if path.is_file() and not path.name.startswith(".")}
    assert set(modes.values()) == {0o644}, modes


def test_two_runs_with_output_folders_of_the_same_name_keep_both_harmonized_copies(tmp_path) -> None:
    first, second = _clean_pair(15)
    config = _two_unit_catalog(tmp_path, first, second)
    harmonized = tmp_path / "harmonized"
    run_slice(config, tmp_path / "a" / "results", harmonized, min_studies=2)
    run_slice(config, tmp_path / "b" / "results", harmonized, min_studies=2)
    run_slice(config, tmp_path / "a" / "results", harmonized, min_studies=2)

    copies = sorted(path.name for path in harmonized.glob("*_harmonized.csv"))
    assert len(copies) == 2 and "results_harmonized.csv" in copies


def test_a_config_that_fails_validation_leaves_no_output_folder_behind(tmp_path) -> None:
    from degora.cli import main

    deg = tmp_path / "deg"
    deg.mkdir()
    _clean_pair(15)[0].to_csv(deg / "a.csv", index=False)
    config = tmp_path / "config.csv"
    pd.DataFrame(
        [{"study_id": "S1", "source_unit_id": "U1", "source_path": "deg/a.csv", "gene_column": "gene", "lfc_column": "log2FoldChange", "p_column": "pvalue"}]
    ).to_csv(config, index=False)
    output = tmp_path / "out_fail"

    assert main(["run", str(config), "--output-dir", str(output), "--no-excel"]) == 2
    assert not output.exists()


# --------------------------------------------------------------------------- L-1 / L-3 / L-5


def test_template_refuses_a_non_xlsx_name_before_writing(tmp_path) -> None:
    from degora.cli import main

    assert main(["template", str(tmp_path / "my_config.csv")]) == 2
    assert not (tmp_path / "my_config.csv").exists()


def test_a_demo_keyword_that_looks_like_a_formula_is_stored_as_text(tmp_path) -> None:
    from openpyxl import load_workbook

    from degora.cli import main, read_excel_settings

    assert main(["demo", str(tmp_path / "kw"), "--keyword", "=1+1"]) == 0
    assert read_excel_settings(tmp_path / "kw" / "degora_demo_config.xlsx")["demo_search_keyword"] == "=1+1"
    sheet = load_workbook(tmp_path / "kw" / "degora_demo_config.xlsx")["Project"]
    assert all(cell.data_type != "f" for row in sheet.iter_rows() for cell in row)


def test_min_studies_from_the_command_line_is_reported_as_such(tmp_path, capsys) -> None:
    from degora.cli import main

    first, second = _clean_pair(10)
    config = _two_unit_catalog(tmp_path, first, second)

    assert main(["run", str(config), "--min-studies", "0", "--output-dir", str(tmp_path / "o"), "--no-excel"]) == 2
    assert "--min-studies option" in capsys.readouterr().err


# --------------------------------------------------------------------------- F3 / M-15 / L-13 (quickstart, gitattributes)


def test_shell_scripts_are_pinned_to_lf_line_endings() -> None:
    attributes = Path(".gitattributes").read_text(encoding="utf-8")
    assert "*.sh text eol=lf" in attributes


def test_quickstart_resolves_an_absolute_demo_dir_and_needs_git_only_to_clone() -> None:
    script = Path("scripts/degora_quickstart.sh").read_text(encoding="utf-8")
    assert 'DB_PATH="$REPO_ROOT/$DEMO_DIR' not in script
    assert "need_git()" in script
    assert 'command -v git >/dev/null 2>&1 || die "git is required. Install it' not in script
    assert "results are written beside it, not inside the checkout" in script


# --------------------------------------------------------------------------- ablation CLI and vectorised lanes


def _corpus_for_ablation(seed: int = 5) -> pd.DataFrame:
    rng = np.random.default_rng(seed)
    genes = [f"G{i:03d}" for i in range(80)]
    truth = rng.normal(0, 1.5, 80)
    rows = []
    for unit, n in (("U1", 3), ("U2", 4), ("U3", 6)):
        lfc = truth + rng.normal(0, 0.4, 80)
        z = lfc / (0.6 / np.sqrt(n))
        for gene, effect, zed in zip(genes, lfc, z):
            rows.append(
                {
                    "study_id": unit,
                    "paper_id": unit,
                    "source_unit_id": unit,
                    "gene_symbol": gene,
                    "lfc": effect,
                    "signed_z": zed,
                    "pvalue": float(2 * stats.norm.sf(abs(zed))),
                    "padj": np.nan,
                    "normalized_rank": 0.5,
                    "n_genes_in_study": 80,
                    "n_ctrl": n,
                    "n_treat": n,
                    "table_scope": "full_results",
                    "source_input_type": "author_deg_table",
                }
            )
    frame = pd.DataFrame(rows)
    frame["normalized_rank"] = frame.groupby("study_id")["signed_z"].rank(ascending=False) / 80
    return frame


def test_ablate_writes_a_summary_and_the_default_variant_equals_the_run(tmp_path, capsys) -> None:
    from degora.ablation import run_ablations
    from degora.cli import main

    harmonized = _corpus_for_ablation()
    results = tmp_path / "results"
    results.mkdir()
    harmonized.to_csv(results / "slice_harmonized.csv", index=False)
    (tmp_path / "gold.txt").write_text("gene_symbol\nG001\nG002\nSEPT9\n", encoding="utf-8")

    assert main(["ablate", str(results), "--gold-panel", str(tmp_path / "gold.txt"), "--weights", "equal=support=1,direction=1,evidence=1,rank=1,effect=1"]) == 0

    summary = pd.read_csv(results / "ablation" / "degora_ablation_summary.csv")
    assert summary["ablation"].tolist()[:2] == ["full", "without_support_score"]
    assert "equal" in summary["ablation"].tolist()
    full = summary.set_index("ablation").loc["full"]
    assert full["spearman_vs_full"] == 1.0 and full["top50_overlap_with_full"] == 1.0
    assert "recall_at_50" in summary.columns
    ranks = pd.read_csv(results / "ablation" / "degora_ablation_ranks.csv")
    assert set(ranks.columns) >= {"gene_symbol", "full", "without_effect_score", "equal"}
    shipped, _evidence, _metadata = degora_score_table(harmonized, min_studies=2)
    full_ranks = ranks.set_index("gene_symbol")["full"]
    assert full_ranks.equals(shipped.set_index("gene_symbol")["quality_weighted_degora_rank"].reindex(full_ranks.index).astype(full_ranks.dtype))
    out = capsys.readouterr().out
    assert "spearman_vs_full" in out
    _summary, per_variant = run_ablations(harmonized, min_studies=2)
    assert set(per_variant) >= {"full", "without_source_quality_weighting", "without_sample_size_weighting"}


def test_ablate_rejects_a_malformed_weight_spec(tmp_path, capsys) -> None:
    from degora.cli import main

    results = tmp_path / "results"
    results.mkdir()
    _corpus_for_ablation().to_csv(results / "slice_harmonized.csv", index=False)

    assert main(["ablate", str(results), "--weights", "bad=support=x"]) == 2
    assert "--weights" in capsys.readouterr().err


def test_vectorised_rra_and_effect_meta_match_a_per_gene_loop() -> None:
    from scipy.stats import beta, t as t_dist

    from degora.score_db import _effect_meta_layer, _rra_beta_layer, _score_ready_harmonized, study_gene_evidence

    evidence = study_gene_evidence(_score_ready_harmonized(_corpus_for_ablation())[0])
    n_lists = int(evidence["source_unit_id"].nunique())

    rra = _rra_beta_layer(evidence, total_source_units=n_lists, min_studies=2).set_index("gene_symbol")
    for gene, group in evidence.groupby("gene_symbol"):
        ranks = np.sort(group["normalized_rank"].to_numpy(dtype=float))
        expected = min(float(beta.logcdf(r, k, n_lists - k + 1)) for k, r in enumerate(ranks, start=1))
        assert rra.loc[gene, "rra_neglog10_rho"] == pytest.approx(-expected / np.log(10.0), abs=1e-9)

    meta = _effect_meta_layer(evidence, min_studies=2).set_index("gene_symbol")
    for gene, group in evidence.groupby("gene_symbol"):
        y = group["lfc"].to_numpy(dtype=float)
        vi = (group["lfc"].abs() / group["signed_z"].abs()).to_numpy(dtype=float) ** 2
        w = 1.0 / vi
        fixed = np.sum(w * y) / np.sum(w)
        q = float(np.sum(w * (y - fixed) ** 2))
        df = len(y) - 1
        c = float(np.sum(w) - np.sum(w**2) / np.sum(w))
        tau2 = max(0.0, (q - df) / c)
        w_re = 1.0 / (vi + tau2)
        pooled = float(np.sum(w_re * y) / np.sum(w_re))
        pooled_se = float(np.sqrt(1.0 / np.sum(w_re)))
        q_hksj = float(np.sum(w_re * (y - pooled) ** 2) / ((len(y) - 1) * np.sum(w_re)))
        se_ci = max(np.sqrt(q_hksj), pooled_se)
        crit = float(t_dist.ppf(0.975, len(y) - 1))
        assert meta.loc[gene, "effect_meta_log2fc_re"] == pytest.approx(pooled, abs=1e-9)
        assert meta.loc[gene, "effect_meta_tau2"] == pytest.approx(tau2, abs=1e-9)
        assert meta.loc[gene, "effect_meta_ci_high"] == pytest.approx(pooled + crit * se_ci, abs=1e-9)
        assert meta.loc[gene, "effect_meta_i2"] == pytest.approx(max(0.0, (q - df) / q) if q > 0 else 0.0, abs=1e-9)
