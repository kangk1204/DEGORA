from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from degora.beginner import (
    BeginnerInitError,
    _write_catalog_atomic,
    build_catalog,
    catalog_row,
    default_study_id,
    describe_inference,
    find_source_tables,
    infer_source_table,
    run_init,
)
from degora.slice_runner import read_catalog, validate_catalog_inputs


def _clean_table(path):
    pd.DataFrame(
        {
            "gene": [f"G{index}" for index in range(200)],
            "log2FoldChange": [2.0 - index * 0.01 for index in range(200)],
            "pvalue": [1e-5 if index < 50 else 0.5 for index in range(200)],
            "padj": [1e-4 if index < 50 else 0.9 for index in range(200)],
        }
    ).to_csv(path, index=False)


def _ambiguous_table(path):
    """Two plausible effect columns and no log2 in either name."""

    pd.DataFrame(
        {
            "Gene symbol": [f"G{index}" for index in range(60)],
            "FC": [1.5] * 60,
            "ratio": [1.2] * 60,
            "P.Value": [0.01] * 60,
            "adj.P.Val": [0.05] * 60,
        }
    ).to_csv(path, sep="\t", index=False)


def _not_a_deg_table(path):
    pd.DataFrame({"sample": ["A", "B"], "batch": [1, 2]}).to_csv(path, index=False)


def test_a_clean_table_needs_no_questions(tmp_path) -> None:
    """Everything a DESeq2 export states about itself is read off it."""

    path = tmp_path / "clean_deseq2.csv"
    _clean_table(path)

    inference = infer_source_table(path)

    assert inference.readable
    assert inference.looks_like_a_deg_table
    assert inference.needs_a_question == ()
    assert inference.mapping == {
        "gene_column": "gene",
        "lfc_column": "log2FoldChange",
        "p_column": "pvalue",
        "padj_column": "padj",
    }
    # The scope is inferred too, so the reader is not asked to classify their table.
    assert inference.table_scope == "full_results"
    assert inference.n_rows == 200


def test_a_missing_optional_column_is_not_a_question(tmp_path) -> None:
    """DEGORA runs without an adjusted p-value; asking for an absent one is noise."""

    path = tmp_path / "no_padj.csv"
    pd.DataFrame(
        {"Symbol": [f"G{i}" for i in range(40)], "logFC": [2.0] * 40, "PValue": [1e-6] * 40}
    ).to_csv(path, index=False)

    inference = infer_source_table(path)

    # `mapping` reports the columns that were found, so an absent optional one is
    # simply not a key - and it is not a question either.
    assert "padj_column" not in inference.mapping
    assert inference.mapping["p_column"] == "PValue"
    assert [choice.role for choice in inference.needs_a_question] == []


def test_beginner_infers_hyphenated_stat_headers_without_collisions(tmp_path) -> None:
    path = tmp_path / "hyphenated.csv"
    pd.DataFrame(
        {
            "Gene Symbol": ["TP53", "BRCA1", "EGFR"],
            "fold-change": [1.2, -0.4, 0.8],
            "p-value": [0.01, 0.02, 0.03],
            "q-value": [0.03, 0.04, 0.05],
        }
    ).to_csv(path, index=False)

    inference = infer_source_table(path)

    assert inference.looks_like_a_deg_table
    assert inference.mapping == {
        "gene_column": "Gene Symbol",
        "lfc_column": "fold-change",
        "p_column": "p-value",
        "padj_column": "q-value",
    }
    assert {choice.role for choice in inference.needs_a_question} == {"lfc_column"}


def test_an_ambiguous_effect_column_is_a_question_not_a_guess(tmp_path) -> None:
    """Picking between two effect columns is the reader's call, not a coin toss."""

    path = tmp_path / "ambiguous.tsv"
    _ambiguous_table(path)

    inference = infer_source_table(path)

    assert inference.looks_like_a_deg_table
    assert "lfc_column" in {choice.role for choice in inference.needs_a_question}


def test_a_file_that_is_not_a_deg_table_is_recognised(tmp_path) -> None:
    """A sample sheet in the same folder must not become a question-by-question walk."""

    path = tmp_path / "sample_sheet.csv"
    _not_a_deg_table(path)

    assert not infer_source_table(path).looks_like_a_deg_table


def test_unnamed_sample_metadata_is_not_promoted_by_plausible_values(tmp_path) -> None:
    path = tmp_path / "metadata.csv"
    pd.DataFrame(
        {
            "sample": [f"S{i}" for i in range(120)],
            "group": [i % 2 for i in range(120)],
            "qc_score": [(i % 10) / 10.0 for i in range(120)],
        }
    ).to_csv(path, index=False)

    inference = infer_source_table(path)

    assert inference.plausible["gene_column"] == ()
    assert inference.plausible["lfc_column"]
    assert inference.plausible["p_column"] == ("qc_score",)
    assert not inference.looks_like_a_deg_table


def test_beginner_init_does_not_auto_accept_sample_id_as_gene_column(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    pd.DataFrame(
        {
            "sample_id": [f"S{i}" for i in range(120)],
            "log2FoldChange": [1.2 - i / 100 for i in range(120)],
            "pvalue": [0.001 + i / 10000 for i in range(120)],
        }
    ).to_csv(deg / "sample_id_results.csv", index=False)

    lines: list[str] = []
    with pytest.raises(BeginnerInitError, match="no table was confirmed"):
        run_init(tmp_path / "config.csv", deg, ask=lambda *_args, **_kwargs: "human", echo=lines.append)

    assert not (tmp_path / "config.csv").exists()
    assert any("this does not look like a DEG results table" in line for line in lines)


def test_beginner_init_does_not_treat_gsm_values_as_gene_symbols_under_sample_id(tmp_path) -> None:
    path = tmp_path / "sample_accessions.tsv"
    pd.DataFrame(
        {
            "sample_id": [f"GSM{100000 + i}" for i in range(120)],
            "log2FoldChange": [1.2 - i / 100 for i in range(120)],
            "pvalue": [0.001 + i / 10000 for i in range(120)],
        }
    ).to_csv(path, sep="\t", index=False)

    inference = infer_source_table(path)

    assert inference.plausible["gene_column"] == ()
    assert inference.mapping.get("gene_column", "") == ""
    assert not inference.looks_like_a_deg_table


def test_beginner_init_rejects_sample_id_even_when_values_look_like_gene_symbols(tmp_path) -> None:
    path = tmp_path / "sample_ids_with_symbol_like_values.tsv"
    pd.DataFrame(
        {
            "sample_id": [f"GENE{i}" for i in range(120)],
            "log2FoldChange": [1.2 - i / 100 for i in range(120)],
            "pvalue": [0.001 + i / 10000 for i in range(120)],
        }
    ).to_csv(path, sep="\t", index=False)

    inference = infer_source_table(path)

    assert inference.plausible["gene_column"] == ()
    assert inference.mapping.get("gene_column", "") == ""
    assert not inference.looks_like_a_deg_table


@pytest.mark.parametrize(
    ("column", "values"),
    [
        ("rank", list(range(1, 121))),
        ("row_number", list(range(1, 121))),
        ("baseMean", list(range(100, 220))),
        ("mean_count", list(range(100, 220))),
        ("stat", list(range(1, 121))),
        ("score", list(range(1, 121))),
        ("index", list(range(1, 121))),
        ("position", list(range(1, 121))),
        ("pathway", [f"PATHWAY_{i}" for i in range(120)]),
        ("pathway_id", [f"PATHWAY_{i}" for i in range(120)]),
        ("metabolite", [f"METABOLITE_{i}" for i in range(120)]),
        ("cell_line", [f"CELL_{i}" for i in range(120)]),
        ("compound", [f"COMPOUND_{i}" for i in range(120)]),
        ("compound_id", [f"COMPOUND_{i}" for i in range(120)]),
        ("cluster_id", list(range(1, 121))),
        ("taxon_id", list(range(9606, 9726))),
    ],
)
def test_beginner_does_not_auto_accept_non_gene_row_labels(
    tmp_path,
    column: str,
    values: list[object],
) -> None:
    path = tmp_path / "non_gene_labels.tsv"
    pd.DataFrame(
        {
            column: values,
            "log2FoldChange": [1.2 - i / 100 for i in range(120)],
            "pvalue": [0.001 + i / 10000 for i in range(120)],
        }
    ).to_csv(path, sep="\t", index=False)

    inference = infer_source_table(path)

    assert inference.plausible["gene_column"] == ()
    assert inference.mapping.get("gene_column", "") == ""
    assert not inference.looks_like_a_deg_table


def test_beginner_init_cannot_publish_rank_values_as_gene_symbols(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    for index in (1, 2):
        pd.DataFrame(
            {
                "rank": list(range(1, 121)),
                "log2FoldChange": [1.2 - i / 100 for i in range(120)],
                "pvalue": [0.001 + i / 10000 for i in range(120)],
            }
        ).to_csv(deg / f"rank_results_{index}.csv", index=False)

    output = tmp_path / "config.csv"
    with pytest.raises(BeginnerInitError, match="no table was confirmed"):
        run_init(
            output,
            deg,
            ask=lambda question, default="": (
                "human"
                if "species" in question.lower()
                else pytest.fail(f"non-gene tables must be skipped before asking: {question}")
            ),
            echo=lambda _line: None,
        )

    assert not output.exists()


@pytest.mark.parametrize(
    ("identifier", "values", "space"),
    [
        ("ID", [f"ENSG{i:011d}.{i % 9 + 1}" for i in range(120)], "Ensembl ID"),
        ("ID", [100000 + i for i in range(120)], "Entrez ID"),
        ("feature", [100000 + i for i in range(120)], "Entrez ID"),
        ("feature_id", [100000 + i for i in range(120)], "Entrez ID"),
        ("identifier", [f"{1000 + i}_at" for i in range(120)], "Affymetrix probe ID"),
        ("identifier", [200000 + i for i in range(120)], "Entrez ID"),
    ],
)
def test_beginner_asks_before_using_value_recognised_identifiers_with_generic_headers(
    tmp_path,
    identifier: str,
    values: list[object],
    space: str,
) -> None:
    path = tmp_path / "generic_identifier.tsv"
    pd.DataFrame(
        {
            identifier: values,
            "log2FoldChange": [1.2 - i / 100 for i in range(120)],
            "pvalue": [0.001 + i / 10000 for i in range(120)],
        }
    ).to_csv(path, sep="\t", index=False)

    inference = infer_source_table(path)

    assert inference.looks_like_a_deg_table
    assert inference.mapping.get("gene_column", "") == ""
    assert inference.plausible["gene_column"] == (identifier,)
    assert inference.identifier_space_for(identifier) == space
    assert [choice.role for choice in inference.needs_a_question] == ["gene_column"]


@pytest.mark.parametrize("identifier", ["gene_id", "entrez_id", "ensembl_gene_id", "probe_id", "transcript_id", "ID_REF"])
def test_beginner_preserves_known_gene_identifier_headers(tmp_path, identifier: str) -> None:
    path = tmp_path / "known_ids.tsv"
    pd.DataFrame(
        {
            identifier: [f"G{i}" for i in range(120)],
            "log2FoldChange": [1.2 - i / 100 for i in range(120)],
            "pvalue": [0.001 + i / 10000 for i in range(120)],
        }
    ).to_csv(path, sep="\t", index=False)

    inference = infer_source_table(path)

    assert inference.looks_like_a_deg_table
    assert inference.mapping["gene_column"] == identifier


def test_an_unreadable_file_does_not_end_the_walk(tmp_path) -> None:
    path = tmp_path / "broken.xlsx"
    path.write_text("not a workbook", encoding="utf-8")

    inference = infer_source_table(path)

    assert not inference.readable
    assert inference.problem
    assert not inference.looks_like_a_deg_table


def test_find_source_tables_skips_dotfiles(tmp_path) -> None:
    _clean_table(tmp_path / "real.csv")
    (tmp_path / ".hidden.csv").write_text("gene,lfc\n", encoding="utf-8")

    assert [path.name for path in find_source_tables(tmp_path)] == ["real.csv"]


def test_generated_degora_catalog_is_not_treated_as_a_source_table(tmp_path) -> None:
    catalog = tmp_path / "old_config.csv"
    pd.DataFrame(
        {
            "study_id": ["S1"],
            "source_path": ["source.csv"],
            "gene_column": ["gene"],
            "lfc_column": ["log2FoldChange"],
            "p_column": ["pvalue"],
            "padj_column": ["padj"],
            "n_ctrl": [3],
            "n_treat": [3],
        }
    ).to_csv(catalog, index=False)

    inference = infer_source_table(catalog)

    assert not inference.looks_like_a_deg_table
    assert "config catalog" in inference.problem


def test_study_ids_are_readable_and_unique(tmp_path) -> None:
    assert default_study_id(Path("GSE123 hypoxia (4h).csv"), []) == "GSE123_hypoxia_4h"
    assert default_study_id(Path("a.csv"), ["a"]) == "a_2"
    assert default_study_id(Path("a.csv"), ["a", "a_2"]) == "a_3"


def test_a_reversed_direction_is_excluded_rather_than_flipped(tmp_path) -> None:
    """DEGORA never reverses an effect column, so it must not pretend to here.

    A table whose positive values mean "up in the control group" is unusable as it
    stands. Flipping the sign for the reader would put a correction DEGORA cannot
    verify into the middle of the evidence chain, so the row is written excluded
    with the reason on it.
    """

    from degora.beginner import ContrastAnswers

    path = tmp_path / "reversed.csv"
    _clean_table(path)
    inference = infer_source_table(path)

    row = catalog_row(
        inference,
        ContrastAnswers(positive_means_up_in_treated=False, source_unit_id="P1"),
        study_id="S1",
        catalog_dir=tmp_path,
    )

    assert row["include_in_analysis"] == "no"
    assert "REVERSED" in row["sign_convention"]
    assert "correct the table at source" in row["notes"]
    # The mapping is preserved, so fixing the table is the only remaining step.
    assert row["lfc_column"] == "log2FoldChange"


def test_a_confirmed_direction_is_recorded_not_assumed(tmp_path) -> None:
    from degora.beginner import ContrastAnswers

    path = tmp_path / "forward.csv"
    _clean_table(path)

    row = catalog_row(
        infer_source_table(path),
        ContrastAnswers(positive_means_up_in_treated=True, source_unit_id="P1"),
        study_id="S1",
        catalog_dir=tmp_path,
    )

    assert row["include_in_analysis"] == "yes"
    assert row["sign_convention"] == "confirmed_treatment_minus_control"


def test_the_guided_flow_produces_a_config_that_validates(tmp_path) -> None:
    """The whole point: what init writes has to be something run accepts."""

    deg = tmp_path / "deg"
    deg.mkdir()
    _clean_table(deg / "clean_deseq2.csv")
    _ambiguous_table(deg / "ambiguous.tsv")
    _not_a_deg_table(deg / "sample_sheet.csv")

    answers = iter(
        [
            "human",  # species, asked once
            # ambiguous.tsv - the reader picks the effect column, then the direction
            "FC",
            "yes",
            "drug vs vehicle",
            "PAPER_A",
            "3",
            "3",
            # clean_deseq2.csv
            "yes",
            "hypoxia vs normoxia",
            "PAPER_B",
            "4",
            "4",
        ]
    )
    config = tmp_path / "config.csv"

    summary = run_init(config, deg, ask=lambda question, default="": next(answers), echo=lambda _line: None)

    assert summary["n_contrasts"] == 2
    assert summary["n_source_units"] == 2
    assert [entry["path"] for entry in summary["skipped"]] == ["sample_sheet.csv"]

    validation = validate_catalog_inputs(config)
    assert validation["active_contrasts"] == 2
    assert validation["source_units"] == 2


def test_beginner_rejects_probability_column_as_effect_size(tmp_path) -> None:
    """A p-value column must not be accepted as the effect column."""

    deg = tmp_path / "deg"
    deg.mkdir()
    _ambiguous_table(deg / "ambiguous.tsv")
    _clean_table(deg / "clean.csv")

    lines: list[str] = []
    answers = iter(
        [
            "human",
            "P.Value",
            "FC",
            "yes",
            "a vs b",
            "P1",
            "3",
            "3",
            "yes",
            "a vs b",
            "P2",
            "3",
            "3",
        ]
    )

    summary = run_init(
        tmp_path / "config.csv",
        deg,
        ask=lambda question, default="": next(answers),
        echo=lines.append,
    )

    assert summary["n_contrasts"] == 2
    assert any("cannot be used as both effect-size column and p-value column" in line for line in lines)
    config = pd.read_csv(tmp_path / "config.csv")
    ambiguous = config.loc[config["source_path"].eq("deg/ambiguous.tsv")].iloc[0]
    assert ambiguous["lfc_column"] == "FC"
    assert ambiguous["p_column"] == "P.Value"


def test_beginner_rejects_a_distinct_probability_candidate_as_effect_size(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    path = deg / "ambiguous.tsv"
    pd.DataFrame(
        {
            "Gene symbol": [f"G{index}" for index in range(60)],
            "FC": [1.5] * 60,
            "ratio": [1.2] * 60,
            "P.Value": [0.01] * 60,
            "p_backup": [0.02] * 60,
            "adj.P.Val": [0.05] * 60,
        }
    ).to_csv(path, sep="\t", index=False)
    lines: list[str] = []
    answers = iter(["human", "p_backup", "FC", "yes", "a vs b", "P1", "3", "3"])

    summary = run_init(
        tmp_path / "config.csv",
        deg,
        ask=lambda question, default="": next(answers),
        echo=lines.append,
    )

    assert summary["n_contrasts"] == 1
    assert any("not one of the available lfc column choices" in line for line in lines)
    config = pd.read_csv(tmp_path / "config.csv")
    assert config.loc[0, "lfc_column"] == "FC"
    assert config.loc[0, "p_column"] == "P.Value"


def test_beginner_skips_after_three_bad_column_answers(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    _ambiguous_table(deg / "ambiguous.tsv")
    _clean_table(deg / "clean.csv")

    answers = iter(
        [
            "human",
            "not_a_column",
            "also_missing",
            "still_missing",
            "yes",
            "a vs b",
            "P2",
            "3",
            "3",
        ]
    )

    summary = run_init(
        tmp_path / "config.csv",
        deg,
        ask=lambda question, default="": next(answers),
        echo=lambda _line: None,
    )

    assert summary["skipped"] == [{"path": "ambiguous.tsv", "reason": "column mapping was not usable"}]
    config = pd.read_csv(tmp_path / "config.csv")
    assert config["source_path"].tolist() == ["deg/clean.csv"]


def test_an_unsure_reader_skips_the_table_rather_than_defaulting(tmp_path) -> None:
    """Pressing enter must never stand in for "yes" on the direction question."""

    deg = tmp_path / "deg"
    deg.mkdir()
    _clean_table(deg / "clean_deseq2.csv")
    _clean_table(deg / "second.csv")

    # "" is unsure, then "yes" to the skip prompt; the second table is confirmed.
    answers = iter(["human", "", "yes", "yes", "hypoxia vs normoxia", "P1", "3", "3"])
    config = tmp_path / "config.csv"

    summary = run_init(config, deg, ask=lambda question, default="": next(answers), echo=lambda _line: None)

    assert summary["n_contrasts"] == 1
    assert summary["skipped"] == [
        {"path": "clean_deseq2.csv", "reason": "contrast direction not confirmed"}
    ]


def test_invalid_direction_answers_are_bounded_and_skip_the_table(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    _clean_table(deg / "first.csv")
    _clean_table(deg / "second.csv")
    answers = iter(["human", "maybe", "maybe", "maybe", "yes", "a vs b", "P2", "3", "3"])
    lines: list[str] = []

    summary = run_init(
        tmp_path / "config.csv",
        deg,
        ask=lambda question, default="": next(answers),
        echo=lines.append,
    )

    assert summary["n_contrasts"] == 1
    assert summary["skipped"] == [{"path": "first.csv", "reason": "contrast direction not confirmed"}]
    assert any("not confirmed after three tries" in line for line in lines)


def test_nothing_confirmed_writes_nothing(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    _not_a_deg_table(deg / "sample_sheet.csv")
    config = tmp_path / "config.csv"

    with pytest.raises(ValueError, match="nothing to write"):
        run_init(config, deg, ask=lambda question, default="": "human", echo=lambda _line: None)

    assert not config.exists()


def test_all_bad_column_answers_write_nothing(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    _ambiguous_table(deg / "ambiguous.tsv")

    answers = iter(["human", "missing_1", "missing_2", "missing_3"])
    config = tmp_path / "config.csv"

    with pytest.raises(ValueError, match="nothing to write"):
        run_init(config, deg, ask=lambda question, default="": next(answers), echo=lambda _line: None)

    assert not config.exists()


def test_an_existing_config_is_not_overwritten(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    _clean_table(deg / "clean.csv")
    config = tmp_path / "config.csv"
    config.write_text("mine\n", encoding="utf-8")

    with pytest.raises(FileExistsError):
        run_init(config, deg, ask=lambda question, default="": "yes", echo=lambda _line: None)

    assert config.read_text(encoding="utf-8") == "mine\n"


def test_build_catalog_keeps_a_stable_column_order() -> None:
    frame = build_catalog([{"study_id": "S1"}])

    assert list(frame.columns)[:3] == ["study_id", "source_unit_id", "source_path"]
    assert "sign_convention" in frame.columns


def test_r_row_labels_are_recognised_as_the_gene_column(tmp_path) -> None:
    """DEGORA renames R's unnamed index to row_name, then did not know that name.

    `write.csv(results, file)` writes the gene identifiers as an unnamed index.
    read_deg_table recovers them under `row_name`, which is DEGORA's own
    convention - and the header classifier had no entry for it, so the guided flow
    asked which column held the gene names for a file whose gene column it had
    just built. Nine contrasts in a real corpus were affected.
    """

    path = tmp_path / "r_export.csv"
    # An R write.csv export: one fewer header field than the data rows.
    path.write_text(
        "log2FoldChange,pvalue,padj\n"
        + "".join(f"GENE{index},{2.0 - index * 0.1},{0.001 * index},{0.01 * index}\n" for index in range(1, 8)),
        encoding="utf-8",
    )

    inference = infer_source_table(path)

    assert inference.mapping["gene_column"] == "row_name"
    assert "gene_column" not in {choice.role for choice in inference.needs_a_question}


def test_a_real_gene_column_still_outranks_the_recovered_row_label(tmp_path) -> None:
    """row_name is a fallback identifier, not a preferred one."""

    from degora.discovery import classify_header

    assert classify_header(["row_name", "log2FoldChange", "pvalue"])["mapping"]["gene_column"] == "row_name"
    assert (
        classify_header(["row_name", "gene_symbol", "log2FoldChange", "pvalue"])["mapping"]["gene_column"]
        == "gene_symbol"
    )


def test_the_fallback_option_list_is_not_the_whole_spreadsheet(tmp_path) -> None:
    """A real table had 43 columns, 32 of them per-sample expression values.

    When the header classifier recognises no gene column, the reader used to be
    offered every column in the file on one line. Values narrow that honestly:
    a column of CPM numbers cannot be a gene name.
    """

    path = tmp_path / "wide.csv"
    frame = pd.DataFrame({"NAME": [f"G{i}" for i in range(300)], "Feature ID": [f"F{i}" for i in range(300)]})
    for sample in range(32):
        frame[f"T{sample} - CPM"] = [float(sample + i) for i in range(300)]
    frame["logFC_T1"] = [0.5] * 300
    frame["P-value_T1"] = [0.01] * 300
    frame.to_csv(path, index=False)

    inference = infer_source_table(path)

    assert inference.plausible["gene_column"] == ("NAME", "Feature ID")
    # The expression columns are numeric, so they are offered for a fold change
    # but never for a gene name.
    assert "T0 - CPM" in inference.plausible["lfc_column"]
    assert "T0 - CPM" not in inference.plausible["gene_column"]


def test_probability_columns_are_not_lfc_fallback_options_when_effects_exist(tmp_path) -> None:
    path = tmp_path / "unnamed_effects.tsv"
    pd.DataFrame(
        {
            "NAME": [f"G{i}" for i in range(120)],
            "FC": [1.5] * 120,
            "ratio": [1.2] * 120,
            "P.Value": [0.01] * 120,
            "adj.P.Val": [0.05] * 120,
        }
    ).to_csv(path, sep="\t", index=False)

    inference = infer_source_table(path)

    assert inference.plausible["lfc_column"] == ("FC", "ratio")
    assert "P.Value" not in inference.plausible["lfc_column"]
    assert "adj.P.Val" not in inference.plausible["lfc_column"]


def test_treatment_label_p_in_effect_name_is_not_mistaken_for_a_pvalue(tmp_path) -> None:
    path = tmp_path / "treatment_p.tsv"
    pd.DataFrame(
        {
            "NAME": [f"G{i}" for i in range(120)],
            "ratio_P": [1.2 + i / 1000 for i in range(120)],
            "FC (P vs C)": [0.5 + i / 1000 for i in range(120)],
            "P.Value": [0.001 + i / 10000 for i in range(120)],
        }
    ).to_csv(path, sep="\t", index=False)

    inference = infer_source_table(path)

    assert "ratio_P" in inference.plausible["lfc_column"]
    assert "FC (P vs C)" in inference.plausible["lfc_column"]
    assert "P.Value" not in inference.plausible["lfc_column"]


def test_numeric_entrez_identifiers_remain_gene_column_options(tmp_path) -> None:
    path = tmp_path / "numeric_ids.tsv"
    pd.DataFrame(
        {
            "feature": [1000 + i for i in range(120)],
            "log2FoldChange": [1.2 + i / 1000 for i in range(120)],
            "P.Value": [0.001 + i / 10000 for i in range(120)],
        }
    ).to_csv(path, sep="\t", index=False)

    inference = infer_source_table(path)

    assert inference.looks_like_a_deg_table
    assert "feature" in inference.plausible["gene_column"]
    assert inference.identifier_space_for("feature") == "Entrez ID"


def test_padj_spellings_are_not_lfc_fallback_options(tmp_path) -> None:
    path = tmp_path / "padj_spellings.tsv"
    pd.DataFrame(
        {
            "NAME": [f"G{i}" for i in range(120)],
            "effect": [0.5] * 120,
            "padj": [0.01] * 120,
            "p_adj": [0.02] * 120,
            "p.adjust": [0.03] * 120,
        }
    ).to_csv(path, sep="\t", index=False)

    inference = infer_source_table(path)

    assert "effect" in inference.plausible["lfc_column"]
    assert "padj" not in inference.plausible["lfc_column"]
    assert "p_adj" not in inference.plausible["lfc_column"]
    assert "p.adjust" not in inference.plausible["lfc_column"]


def test_unit_interval_effect_values_stay_lfc_fallback_options(tmp_path) -> None:
    path = tmp_path / "small_effects.tsv"
    pd.DataFrame(
        {
            "NAME": [f"G{i}" for i in range(120)],
            "baseMean": [100.0 + i for i in range(120)],
            "effect": [0.1 + (i % 8) * 0.1 for i in range(120)],
            "P.Value": [0.01] * 120,
        }
    ).to_csv(path, sep="\t", index=False)

    inference = infer_source_table(path)

    assert "effect" in inference.plausible["lfc_column"]
    assert "baseMean" in inference.plausible["lfc_column"]
    assert "P.Value" not in inference.plausible["lfc_column"]


def test_binary_significance_flag_is_not_inferred_as_a_probability_column(tmp_path) -> None:
    path = tmp_path / "thresholded_flags.tsv"
    pd.DataFrame(
        {
            "gene": [f"G{i}" for i in range(120)],
            "log2FoldChange": [1.0 if i % 2 else -1.0 for i in range(120)],
            "significant": [i % 2 for i in range(120)],
        }
    ).to_csv(path, sep="\t", index=False)

    inference = infer_source_table(path)

    assert "significant" not in inference.plausible["p_column"]
    assert not inference.looks_like_a_deg_table


def test_binary_values_under_a_pvalue_header_are_rejected_during_init_inference(tmp_path) -> None:
    path = tmp_path / "thresholded_pvalue.tsv"
    pd.DataFrame(
        {
            "gene": [f"G{i}" for i in range(120)],
            "log2FoldChange": [1.0 if i % 2 else -1.0 for i in range(120)],
            "pvalue": [i % 2 for i in range(120)],
        }
    ).to_csv(path, sep="\t", index=False)

    inference = infer_source_table(path)

    assert not inference.looks_like_a_deg_table
    assert "p_column" not in inference.mapping
    assert "binary significance flag" in inference.problem
    assert "unrounded gene-level p-values" in "\n".join(describe_inference(inference))


def test_probability_vector_with_zero_one_and_interior_values_remains_usable(tmp_path) -> None:
    path = tmp_path / "real_pvalues.tsv"
    pd.DataFrame(
        {
            "gene": [f"G{i}" for i in range(120)],
            "log2FoldChange": [1.0 if i % 2 else -1.0 for i in range(120)],
            "pvalue": [0.0, 1.0, 0.5] * 40,
        }
    ).to_csv(path, sep="\t", index=False)

    inference = infer_source_table(path)

    assert inference.looks_like_a_deg_table
    assert inference.mapping["p_column"] == "pvalue"
    assert inference.problem == ""


def test_probability_only_numeric_columns_still_remain_lfc_fallback_options(tmp_path) -> None:
    path = tmp_path / "probability_only.tsv"
    pd.DataFrame(
        {
            "NAME": [f"G{i}" for i in range(120)],
            "P.Value": [0.01] * 120,
            "adj.P.Val": [0.05] * 120,
        }
    ).to_csv(path, sep="\t", index=False)

    inference = infer_source_table(path)

    assert inference.plausible["lfc_column"] == ("P.Value", "adj.P.Val")


def test_lfc_answer_is_not_rejected_before_missing_pvalue_is_asked(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    pd.DataFrame(
        {
            "Gene symbol": [f"G{i}" for i in range(120)],
            "FC": [1.2] * 120,
            "ratio": [0.8] * 120,
            "adj.P.Val": [0.01] * 120,
        }
    ).to_csv(deg / "padj_only_ambiguous_lfc.tsv", sep="\t", index=False)

    answers = iter(["human", "ratio", "adj.P.Val", "yes", "a vs b", "P1", "3", "3"])

    summary = run_init(
        tmp_path / "config.csv",
        deg,
        ask=lambda question, default="": next(answers),
        echo=lambda _line: None,
    )

    assert summary["n_contrasts"] == 1
    assert summary["skipped"] == []
    config = pd.read_csv(tmp_path / "config.csv")
    row = config.iloc[0]
    assert row["lfc_column"] == "ratio"
    assert row["p_column"] == "adj.P.Val"
    assert row["padj_column"] == "adj.P.Val"


def test_empty_required_column_answer_retries_instead_of_skipping_immediately(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    pd.DataFrame(
        {
            "Gene symbol": [f"G{i}" for i in range(120)],
            "FC": [1.2] * 120,
            "ratio": [0.8] * 120,
            "adj.P.Val": [0.01] * 120,
        }
    ).to_csv(deg / "padj_only.tsv", sep="\t", index=False)

    lines: list[str] = []
    answers = iter(["human", "", "ratio", "adj.P.Val", "yes", "a vs b", "P1", "3", "3"])
    summary = run_init(
        tmp_path / "config.csv",
        deg,
        ask=lambda question, default="": next(answers),
        echo=lines.append,
    )

    assert summary["n_contrasts"] == 1
    assert any("lfc column is required" in line and "Try again" in line for line in lines)
    assert pd.read_csv(tmp_path / "config.csv").loc[0, "lfc_column"] == "ratio"


def test_init_rejects_legacy_xls_output_before_prompting(tmp_path) -> None:
    with pytest.raises(BeginnerInitError, match=r"cannot write the legacy \.xls format"):
        run_init(
            tmp_path / "config.xls",
            tmp_path / "missing-input-dir",
            ask=lambda *_args, **_kwargs: pytest.fail("init should reject .xls before prompting"),
        )


@pytest.mark.parametrize("name", ["config", "config.json", "config.tsv", "config.csv.gz"])
def test_init_rejects_unsupported_output_suffix_before_prompting(tmp_path, name) -> None:
    with pytest.raises(BeginnerInitError, match=r"must end in \.csv or \.xlsx"):
        run_init(
            tmp_path / name,
            tmp_path / "missing-input-dir",
            ask=lambda *_args, **_kwargs: pytest.fail("init should reject the suffix before prompting"),
        )


def test_init_force_keeps_existing_config_when_csv_write_fails(tmp_path, monkeypatch) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    _clean_table(deg / "clean.csv")
    output = tmp_path / "config.csv"
    original = "original,config\n1,2\n"
    output.write_text(original)
    answers = iter(["human", "yes", "a vs b", "P1", "3", "3"])

    def fail_to_csv(self, path, *args, **kwargs):
        if Path(path).name.startswith(".config.csv."):
            raise OSError("simulated writer failure")
        return original_to_csv(self, path, *args, **kwargs)

    original_to_csv = pd.DataFrame.to_csv
    monkeypatch.setattr(pd.DataFrame, "to_csv", fail_to_csv)

    with pytest.raises(OSError, match="simulated writer failure"):
        run_init(output, deg, ask=lambda question, default="": next(answers), echo=lambda _line: None, force=True)

    assert output.read_text() == original
    assert not list(tmp_path.glob(".config.csv.*"))


def test_beginner_csv_config_formula_guard_is_reversible_and_provenanced(tmp_path) -> None:
    output = tmp_path / "config.csv"
    catalog = pd.DataFrame(
        {
            "study_id": ["S1"],
            "source_unit_id": ["U1"],
            "source_path": ["source.csv"],
            "gene_column": ["gene"],
            "lfc_column": ["logFC"],
            "p_column": ["pvalue"],
            "hypoxia_modality": ["-Dox versus +Dox"],
        }
    )

    _write_catalog_atomic(catalog, output)

    assert pd.read_csv(output).loc[0, "hypoxia_modality"] == "'-Dox versus +Dox"
    assert read_catalog(output).loc[0, "hypoxia_modality"] == "-Dox versus +Dox"
    provenance = json.loads(Path(str(output) + ".provenance.json").read_text())
    assert provenance["metadata"]["csv_formula_guard"] == "reversible_apostrophe_prefix_v1"


def test_plain_generated_csv_config_remains_editable_with_a_stale_digest(tmp_path) -> None:
    output = tmp_path / "config.csv"
    catalog = pd.DataFrame(
        {
            "study_id": ["S1"],
            "source_unit_id": ["U1"],
            "source_path": ["source.csv"],
            "gene_column": ["gene"],
            "lfc_column": ["logFC"],
            "p_column": ["pvalue"],
            "notes": ["before edit"],
        }
    )
    _write_catalog_atomic(catalog, output)
    edited = pd.read_csv(output)
    edited.loc[0, "notes"] = "after edit"
    edited.to_csv(output, index=False)

    assert read_catalog(output).loc[0, "notes"] == "after edit"


def test_init_rejects_directory_output_even_with_force_before_prompting(tmp_path) -> None:
    output = tmp_path / "config_dir"
    output.mkdir()

    with pytest.raises(BeginnerInitError, match="output must be a file path"):
        run_init(
            output,
            tmp_path / "missing-input-dir",
            force=True,
            ask=lambda *_args, **_kwargs: pytest.fail("init should reject a directory before prompting"),
        )


def test_init_rejects_zero_row_tables_before_prompting_or_writing(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    pd.DataFrame(columns=["gene", "log2FoldChange", "pvalue"]).to_csv(deg / "empty.csv", index=False)
    config = tmp_path / "config.csv"

    with pytest.raises(BeginnerInitError, match="no data rows"):
        run_init(
            config,
            deg,
            ask=lambda *_args, **_kwargs: pytest.fail("init should reject empty tables before prompting"),
        )

    assert not config.exists()


@pytest.mark.parametrize("suffix", [".csv", ".xlsx"])
def test_forced_init_preserves_existing_config_when_atomic_replace_fails(tmp_path, monkeypatch, suffix) -> None:
    import degora.beginner as beginner

    deg = tmp_path / "deg"
    deg.mkdir()
    _clean_table(deg / "valid.csv")
    output = tmp_path / f"config{suffix}"
    original = b"ORIGINAL CONFIG\n"
    output.write_bytes(original)
    answers = iter(["human", "yes", "a vs b", "P1", "3", "3"])

    def fail_replace(source, target):
        assert Path(target) == output
        assert Path(source).is_file()
        raise OSError("synthetic atomic replace failure")

    monkeypatch.setattr(beginner.os, "replace", fail_replace)
    with pytest.raises(OSError, match="synthetic atomic replace failure"):
        run_init(
            output,
            deg,
            ask=lambda question, default="": next(answers),
            echo=lambda _line: None,
            force=True,
        )

    assert output.read_bytes() == original
    assert list(tmp_path.glob(f".{output.name}.*")) == []


def test_init_skips_an_empty_table_when_a_valid_deg_table_is_present(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    pd.DataFrame(columns=["gene", "log2FoldChange", "pvalue"]).to_csv(deg / "empty.csv", index=False)
    _clean_table(deg / "valid.csv")

    answers = iter(["human", "yes", "a vs b", "P1", "3", "3"])
    summary = run_init(
        tmp_path / "config.csv",
        deg,
        ask=lambda question, default="": next(answers),
        echo=lambda _line: None,
    )

    assert summary["n_contrasts"] == 1
    assert summary["skipped"] == [
        {"path": "empty.csv", "reason": "this table has column headers but no data rows"}
    ]


def test_only_columns_inside_zero_and_one_are_offered_as_a_p_value(tmp_path) -> None:
    """Offering baseMean as a p-value candidate is offering nonsense."""

    path = tmp_path / "deseq_no_pvalue.tsv"
    pd.DataFrame(
        {
            "row_name": [f"G{i}" for i in range(120)],
            "baseMean": [100.0 + i for i in range(120)],
            "log2 fold change": [2.0 - i * 0.01 for i in range(120)],
            "padj": [i / 200.0 for i in range(120)],
        }
    ).to_csv(path, sep="\t", index=False)

    inference = infer_source_table(path)

    assert inference.plausible["p_column"] == ("padj",)


def test_a_table_with_only_adjusted_p_values_says_so(tmp_path) -> None:
    """The reader should choose that knowingly, not discover it later."""

    path = tmp_path / "padj_only.tsv"
    pd.DataFrame(
        {
            "row_name": [f"G{i}" for i in range(120)],
            "baseMean": [100.0 + i for i in range(120)],
            "log2 fold change": [2.0 - i * 0.01 for i in range(120)],
            "padj": [i / 200.0 for i in range(120)],
        }
    ).to_csv(path, sep="\t", index=False)

    notes = [choice.note for choice in infer_source_table(path).needs_a_question]

    assert any("no unadjusted p-value" in note for note in notes)
    assert any("already adjusted" in note for note in notes)


def test_a_long_option_list_is_truncated_with_a_way_out(tmp_path) -> None:
    from degora.beginner import MAX_OPTIONS_SHOWN

    deg = tmp_path / "deg"
    deg.mkdir()
    frame = pd.DataFrame({f"label_{index}": [f"V{index}_{row}" for row in range(200)] for index in range(30)})
    frame["logFC"] = [1.0] * 200
    frame["PValue"] = [0.01] * 200
    frame.to_csv(deg / "many_labels.csv", index=False)

    lines: list[str] = []
    answers = iter(["human", "label_0", "yes", "x vs y", "P1", "3", "3"])
    run_init(
        tmp_path / "config.csv",
        deg,
        ask=lambda question, default="": next(answers),
        echo=lines.append,
    )

    options_line = next(line for line in lines if line.strip().startswith("Options:"))
    assert options_line.count(",") <= MAX_OPTIONS_SHOWN
    assert "type another listed choice" in options_line


def test_an_ensembl_table_and_a_symbol_table_are_flagged_before_the_run(tmp_path) -> None:
    """A run mixing identifier spaces scores zero genes, and said so only at the end.

    Two real GEO series downloaded for the same keyword wrote their gene columns
    in different conventions: one Ensembl IDs, one symbols. DEGORA matches genes
    on the identifier itself, so those two share nothing. The config validated,
    the run took its full time, and reported zero genes scored. Every table was
    already read while the config was being built, so it can be said there.
    """

    deg = tmp_path / "deg"
    deg.mkdir()
    pd.DataFrame(
        {
            "gene": [f"ENSG{index:011d}" for index in range(150)],
            "log2FoldChange": [1.0] * 150,
            "pvalue": [0.001] * 150,
        }
    ).to_csv(deg / "ensembl_study.csv", index=False)
    pd.DataFrame(
        {
            "gene": [f"GENE{index}" for index in range(150)],
            "log2FoldChange": [1.0] * 150,
            "pvalue": [0.001] * 150,
        }
    ).to_csv(deg / "symbol_study.csv", index=False)

    lines: list[str] = []
    answers = iter(["human", "yes", "a vs b", "P1", "3", "3", "yes", "a vs b", "P2", "3", "3"])
    summary = run_init(
        tmp_path / "config.csv",
        deg,
        ask=lambda question, default="": next(answers),
        echo=lines.append,
    )

    assert summary["identifier_warning"]
    assert "ensembl_study.csv" in summary["identifier_spaces"]["Ensembl ID"]
    assert "symbol_study.csv" in summary["identifier_spaces"]["gene symbol"]
    # And the reader is told, not just the return value.
    assert any("WARNING" in line and "identifier space" in line for line in lines)


def test_unrecognised_identifier_space_is_flagged_before_writing_config(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    pd.DataFrame(
        {
            "gene": [f"chr1:{index}-{index + 1}" for index in range(150)],
            "log2FoldChange": [1.0] * 150,
            "pvalue": [0.001] * 150,
        }
    ).to_csv(deg / "unknown_ids.csv", index=False)
    _clean_table(deg / "symbols.csv")

    config = tmp_path / "config.csv"
    lines: list[str] = []
    warning_saw_no_output = False

    def echo(line: str) -> None:
        nonlocal warning_saw_no_output
        if "WARNING" in line and "identifier space" in line:
            warning_saw_no_output = not config.exists()
        lines.append(line)

    answers = iter(["human", "yes", "a vs b", "P1", "3", "3", "yes", "a vs b", "P2", "3", "3"])
    summary = run_init(config, deg, ask=lambda question, default="": next(answers), echo=echo)

    assert config.exists()
    assert warning_saw_no_output
    assert summary["identifier_warning"]
    assert "unrecognised identifiers" in summary["identifier_warning"]
    assert "unknown_ids.csv" in summary["identifier_spaces"]["unrecognised identifiers"]
    assert any("WARNING" in line and "identifier space" in line for line in lines)


def test_one_identifier_space_raises_no_warning(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    for name in ("first.csv", "second.csv"):
        _clean_table(deg / name)

    answers = iter(["human", "yes", "a vs b", "P1", "3", "3", "yes", "a vs b", "P2", "3", "3"])
    summary = run_init(
        tmp_path / "config.csv",
        deg,
        ask=lambda question, default="": next(answers),
        echo=lambda _line: None,
    )

    assert summary["identifier_warning"] == ""
    assert list(summary["identifier_spaces"]) == ["gene symbol"]


def test_identifier_space_names_the_common_conventions() -> None:
    from degora.beginner import UNKNOWN_IDENTIFIER_SPACE, identifier_space

    assert identifier_space(["ENSG00000141510", "ENSG00000012048.7"]) == "Ensembl ID"
    assert identifier_space(["TP53", "BRCA1", "A1BG-AS1"]) == "gene symbol"
    assert identifier_space(["1007_s_at", "1053_at"]) == "Affymetrix probe ID"
    assert identifier_space(["NM_000546.6", "NR_003286.2"]) == "RefSeq ID"
    assert identifier_space(["7157", "672"]) == "Entrez ID"
    # Too mixed to name is reported as such rather than guessed at.
    assert identifier_space(["TP53", "ENSG00000141510", "1007_s_at", "7157"]) == UNKNOWN_IDENTIFIER_SPACE
    assert identifier_space([]) == UNKNOWN_IDENTIFIER_SPACE


def test_the_identifier_space_follows_the_column_the_reader_picked(tmp_path) -> None:
    """The reader overrides the gene column; the space must describe that column.

    The check read the auto-detected column instead. A table offering both an
    Ensembl column and a symbol column reported whichever DEGORA guessed, and a
    table where nothing was auto-detected reported nothing at all - so the one
    case that most needs the mixed-space check was silently left out of it.
    """

    deg = tmp_path / "deg"
    deg.mkdir()
    # Two candidate gene columns in different spaces, so the pick decides the answer.
    pd.DataFrame(
        {
            "gene_id": [f"ENSG{index:011d}" for index in range(150)],
            "gene_symbol": [f"GENE{index}" for index in range(150)],
            "log2FoldChange": [1.0] * 150,
            "pvalue": [0.001] * 150,
        }
    ).to_csv(deg / "both_spaces.csv", index=False)
    # A second study in symbols, so picking Ensembl above makes the pair mixed.
    _clean_table(deg / "symbols.csv")

    def build(gene_pick: str) -> dict:
        answers = iter(["human", gene_pick, "yes", "a vs b", "P1", "3", "3", "yes", "a vs b", "P2", "3", "3"])
        config = tmp_path / f"config_{gene_pick}.csv"
        return run_init(
            config, deg, ask=lambda question, default="": next(answers), echo=lambda _line: None
        )

    picked_ensembl = build("gene_id")
    assert "both_spaces.csv" in picked_ensembl["identifier_spaces"]["Ensembl ID"]
    assert picked_ensembl["identifier_warning"], "mixing Ensembl with symbols must be flagged"

    picked_symbol = build("gene_symbol")
    assert "both_spaces.csv" in picked_symbol["identifier_spaces"]["gene symbol"]
    assert picked_symbol["identifier_warning"] == "", "both tables are symbols; nothing to warn about"


def test_identifier_warning_ignores_excluded_reversed_tables(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    pd.DataFrame(
        {
            "gene": [f"ENSG{index:011d}" for index in range(150)],
            "log2FoldChange": [1.0] * 150,
            "pvalue": [0.001] * 150,
        }
    ).to_csv(deg / "excluded_ensembl.csv", index=False)
    _clean_table(deg / "symbols_a.csv")
    _clean_table(deg / "symbols_b.csv")

    answers = iter(
        [
            "human",
            "no",
            "a vs b",
            "P0",
            "3",
            "3",
            "yes",
            "a vs b",
            "P1",
            "3",
            "3",
            "yes",
            "a vs b",
            "P2",
            "3",
            "3",
        ]
    )
    summary = run_init(
        tmp_path / "config.csv",
        deg,
        ask=lambda question, default="": next(answers),
        echo=lambda _line: None,
    )

    assert summary["identifier_warning"] == ""
    assert summary["n_source_units"] == 2
    assert summary["n_excluded_reversed_direction"] == 1
    assert "excluded_ensembl.csv" in summary["identifier_spaces"]["Ensembl ID"]


def test_tables_from_one_study_cannot_meet_the_replication_rule(tmp_path) -> None:
    """Five tables from one GEO series is a config that scores nothing."""

    deg = tmp_path / "deg"
    deg.mkdir()
    for name in ("first.csv", "second.csv"):
        _clean_table(deg / name)

    lines: list[str] = []
    # Same answer to "which paper or dataset" for both: one source unit.
    answers = iter(["human", "yes", "a vs b", "GSE1", "3", "3", "yes", "a vs b", "GSE1", "3", "3"])
    summary = run_init(
        tmp_path / "config.csv",
        deg,
        ask=lambda question, default="": next(answers),
        echo=lines.append,
    )

    assert summary["n_source_units"] == 1
    assert "scores zero genes" in summary["replication_warning"]
    assert any("WARNING" in line and "independent source unit" in line for line in lines)


def test_two_source_units_raise_no_replication_warning(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    for name in ("first.csv", "second.csv"):
        _clean_table(deg / name)

    answers = iter(["human", "yes", "a vs b", "GSE1", "3", "3", "yes", "a vs b", "GSE2", "3", "3"])
    summary = run_init(
        tmp_path / "config.csv",
        deg,
        ask=lambda question, default="": next(answers),
        echo=lambda _line: None,
    )

    assert summary["replication_warning"] == ""


def test_pvalue_override_resets_inferred_table_scope_to_auto(tmp_path) -> None:
    from degora.beginner import ContrastAnswers

    path = tmp_path / "scope.csv"
    pd.DataFrame(
        {
            "gene": [f"G{i}" for i in range(200)],
            "log2FoldChange": [1.0] * 200,
            "pvalue": [0.001 if i < 20 else 0.9 for i in range(200)],
            "pvalue_alt": [0.001] * 200,
        }
    ).to_csv(path, index=False)
    inference = infer_source_table(path)

    row = catalog_row(
        inference,
        ContrastAnswers(
            positive_means_up_in_treated=True,
            source_unit_id="P1",
            overrides={"p_column": "pvalue_alt"},
        ),
        study_id="S1",
        catalog_dir=tmp_path,
    )

    assert inference.table_scope == "full_results"
    assert row["p_column"] == "pvalue_alt"
    assert row["table_scope"] == "auto"


def test_init_xlsx_stores_formula_like_text_as_literals(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    _clean_table(deg / "clean.csv")

    answers = iter(["human", "yes", "-Dox vs +Dox", "=paper", "3", "3"])
    output = tmp_path / "config.xlsx"

    run_init(output, deg, ask=lambda question, default="": next(answers), echo=lambda _line: None)

    workbook = load_workbook(output, data_only=False)
    sheet = workbook["Contrasts"]
    rows = list(sheet.iter_rows(values_only=False))
    headers = [cell.value for cell in rows[0]]
    values = {header: rows[1][index] for index, header in enumerate(headers)}
    assert values["condition"].value == "-Dox vs +Dox"
    assert values["condition"].data_type == "s"
    assert values["source_unit_id"].value == "=paper"
    assert values["source_unit_id"].data_type == "s"


def test_identifier_warning_is_emitted_before_config_is_written(tmp_path) -> None:
    deg = tmp_path / "deg"
    deg.mkdir()
    _clean_table(deg / "symbols.csv")
    pd.DataFrame(
        {
            "gene": ["A-1", "B:2", "C/3", "D 4"] * 50,
            "log2FoldChange": [1.0] * 200,
            "pvalue": [0.01] * 200,
        }
    ).to_csv(deg / "unknown_ids.csv", index=False)
    output = tmp_path / "config.csv"
    answers = iter(["human", "yes", "a vs b", "P1", "3", "3", "yes", "a vs b", "P2", "3", "3"])
    warning_seen_before_write = False

    def echo(line: str) -> None:
        nonlocal warning_seen_before_write
        if line.startswith("WARNING:") and "identifier" in line:
            warning_seen_before_write = not output.exists()

    summary = run_init(output, deg, ask=lambda question, default="": next(answers), echo=echo)

    assert warning_seen_before_write is True
    assert output.exists()
    assert "unrecognised identifiers" in summary["identifier_warning"]
